from datetime import datetime, date
from venv import logger
import logging
from django.shortcuts import render, redirect
from django.db.models import Count, Q, Sum, Case, When, IntegerField
from django.core.mail import EmailMultiAlternatives
from django.http import HttpResponse
from .models import BMI, Account, Barangay, Parent, Temperature, VaccinationSchedule, BHW, BNS, Announcement, Midwife, Nurse
from .models import get_enhanced_vaccine_status, get_vaccine_eligibility
from .services import PushNotificationService  # Single clean import
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.models import User
from rest_framework.permissions import IsAuthenticated
from django.utils.crypto import get_random_string
from django.contrib.auth.hashers import make_password
from django.contrib.auth import authenticate, login as auth_login
from django.contrib.auth.decorators import login_required #abang lang
from django.shortcuts import render, redirect, get_object_or_404
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from django.utils.timezone import now
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.contrib.auth.password_validation import validate_password
from .models import PasswordResetOTP
from django.core.serializers.json import DjangoJSONEncoder
from .models import Account, ProfilePhoto
from django.views.decorators.csrf import csrf_exempt,csrf_protect
from django.http import JsonResponse
from django.urls import reverse
from django.template.loader import render_to_string
from weasyprint import HTML
from .models import Preschooler, Account, Barangay, BHW, BMI, Parent, PasswordResetOTP, NutritionService
from .models import ParentActivityLog, PreschoolerActivityLog
from django.db.models import Prefetch
from django.db.models import Q, F
import json
from calendar import monthrange
from django.http import HttpResponseForbidden
from django.contrib.auth.hashers import check_password
from django.views.decorators.http import require_POST
from .models import Preschooler, BMI, Temperature, Barangay, BHW, BNS
from django.utils import timezone
from django.db.models.functions import Lower
from django.utils.timesince import timesince
from datetime import timedelta
from django.db import IntegrityError
from django.core.paginator import Paginator
import calendar
import random
import string
from collections import defaultdict
from django.db.models.functions import TruncMonth
from .models import Account, Parent, Barangay, BHW
from django.db.models import Count, Q
from django.views.decorators.http import require_GET
from django.utils.html import strip_tags
from django.utils.dateparse import parse_date
from django.contrib.auth.decorators import login_required
from rest_framework.views import APIView
from WebApp.modelserializers import AccountSerializer
from .modelserializers import *
from rest_framework.response import Response
from rest_framework import status
from django.views.decorators.csrf import csrf_exempt
from rest_framework.authtoken.models import Token
from django.utils.decorators import method_decorator
from dateutil.relativedelta import relativedelta
#added for hardware - start
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt                                   
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from .serializers import ESP32DataSerializer, ESP32ResponseSerializer
from django.core.files.base import ContentFile
import cloudinary.uploader

from django.contrib.auth.password_validation import (
    MinimumLengthValidator,
    NumericPasswordValidator,
    UserAttributeSimilarityValidator
)
import json
from datetime import datetime, timedelta, date
from .decorators import admin_required
OFFLINE_THRESHOLD = timedelta(minutes=5)

#REPORT
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

ESP32_DATA_CACHE = {}

# Device tracking for status monitoring
DEVICE_STATUS = {}

@csrf_exempt
@require_http_methods(["POST"])
def receive_esp32_data_simple(request):
    """
    Multi-device ESP32 data reception with device ID separation
    """
    global ESP32_DATA_CACHE, DEVICE_STATUS
    try:
        # Parse JSON data
        data = json.loads(request.body)
        
        # Use serializer for validation
        serializer = ESP32DataSerializer(data=data)
        
        if serializer.is_valid():
            validated_data = serializer.validated_data
            device_id = validated_data.get('device_id')
            measurement_type = validated_data.get('measurement_type', 'UNKNOWN').upper()
            
            # Validate device_id is provided
            if not device_id:
                return JsonResponse({
                    'status': 'error',
                    'message': 'device_id is required for multi-device support'
                }, status=400)
            
            # Update device status (CREATE if doesn't exist - FIXED)
            if device_id not in DEVICE_STATUS:
                print(f"DEBUG: Creating new device entry for {device_id}")
                DEVICE_STATUS[device_id] = {
                    'device_name': device_id,  # Use device_id as name initially
                    'last_seen': None,
                    'is_online': False,
                    'measurements_today': 0,
                    'last_measurement_type': None
                }

            # Update device status
            DEVICE_STATUS[device_id]['last_seen'] = timezone.now()
            DEVICE_STATUS[device_id]['is_online'] = True
            DEVICE_STATUS[device_id]['last_measurement_type'] = measurement_type
            DEVICE_STATUS[device_id]['measurements_today'] += 1
            
            print(f"DEBUG: Updated device status for {device_id}: {DEVICE_STATUS[device_id]}")
            
            # Initialize device data if not exists
            if device_id not in ESP32_DATA_CACHE:
                ESP32_DATA_CACHE[device_id] = {}
                print(f"DEBUG: Initialized new cache entry for device {device_id}")
            else:
                print(f"DEBUG: Existing cache for device {device_id}: {ESP32_DATA_CACHE[device_id]}")
            
            # Store data based on measurement type - MERGE with existing data
            if measurement_type == 'BMI':
                # Add BMI data while preserving any existing temperature data
                print(f"DEBUG: Adding BMI data to device {device_id}...")
                ESP32_DATA_CACHE[device_id].update({
                    'weight': validated_data['weight'],
                    'height': validated_data['height'],
                    'bmi': validated_data.get('bmi'),
                    'bmi_category': validated_data.get('bmi_category'),
                    'subject_type': validated_data.get('subject_type'),
                    'sensor_height': validated_data.get('sensor_height'),
                    'bmi_timestamp': str(timezone.now()),
                    'has_bmi_data': True
                })
                print(f"DEBUG: After BMI update for {device_id}: {ESP32_DATA_CACHE[device_id]}")
                
            elif measurement_type == 'TEMPERATURE':
                # Add temperature data while preserving any existing BMI data
                print(f"DEBUG: Adding temperature data to device {device_id}...")
                ESP32_DATA_CACHE[device_id].update({
                    'temperature': validated_data['temperature'],
                    'temperature_status': validated_data.get('temperature_status'),
                    'temp_timestamp': str(timezone.now()),
                    'has_temperature_data': True
                })
                print(f"DEBUG: After temperature update for {device_id}: {ESP32_DATA_CACHE[device_id]}")
                
            else:
                # Legacy support - store all available data
                ESP32_DATA_CACHE[device_id].update({
                    'weight': validated_data.get('weight'),
                    'height': validated_data.get('height'),
                    'temperature': validated_data.get('temperature'),
                    'bmi': validated_data.get('bmi'),
                    'bmi_category': validated_data.get('bmi_category'),
                    'temperature_status': validated_data.get('temperature_status'),
                })
            
            # Always update common fields
            ESP32_DATA_CACHE[device_id].update({
                'device_id': device_id,
                'last_update': str(timezone.now()),
                'esp32_timestamp': validated_data.get('timestamp'),
                'measurement_type': measurement_type
            })
            
            return JsonResponse({
                'status': 'success',
                'message': f'{measurement_type} data received successfully from {device_id}',
                'data': ESP32_DATA_CACHE[device_id],
                'measurement_type': measurement_type,
                'device_id': device_id
            })
            
        else:
            print(f"DEBUG: Validation failed: {serializer.errors}")
            return JsonResponse({
                'status': 'error',
                'message': 'Validation failed',
                'errors': serializer.errors
            }, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({
            'status': 'error',
            'message': 'Invalid JSON data'
        }, status=400)
        
    except Exception as e:
        print(f"DEBUG: Exception occurred: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
        
@csrf_exempt
@require_http_methods(["GET"])
def get_esp32_data_simple(request):
    """
    Get ESP32 data with device selection support
    """
    global ESP32_DATA_CACHE
    
    # Get device_id from URL parameter (required for multi-device)
    device_id = request.GET.get('device_id')
    force_clear = request.GET.get('clear', 'false').lower() == 'true'
    
    print(f"DEBUG: get_esp32_data_simple called - device_id: {device_id}, force_clear: {force_clear}")
    
    # If no device_id specified, return available devices
    if not device_id:
        available_devices = []
        for dev_id, data in ESP32_DATA_CACHE.items():
            available_devices.append({
                'device_id': dev_id,
                'device_name': DEVICE_STATUS.get(dev_id, {}).get('device_name', dev_id),
                'last_update': data.get('last_update'),
                'has_data': True,
                'is_online': DEVICE_STATUS.get(dev_id, {}).get('is_online', False)
            })
        
        # Also include devices with no data but known status
        for dev_id, status in DEVICE_STATUS.items():
            if dev_id not in ESP32_DATA_CACHE:
                available_devices.append({
                    'device_id': dev_id,
                    'device_name': status.get('device_name', dev_id),
                    'last_update': None,
                    'has_data': False,
                    'is_online': status.get('is_online', False)
                })
        
        return JsonResponse({
            'status': 'device_list',
            'message': 'Please specify device_id parameter',
            'available_devices': available_devices,
            'example_url': '/api/esp32/get-data/?device_id=BMI_STATION_A'
        })
    
    # If force clear is requested, clear data first and return
    if force_clear:
        if device_id in ESP32_DATA_CACHE:
            cleared_data = ESP32_DATA_CACHE[device_id].copy()
            del ESP32_DATA_CACHE[device_id]
            print(f"DEBUG: FORCE CLEARED data for device {device_id}: {cleared_data}")
        else:
            print(f"DEBUG: No data to clear for device {device_id} (already empty)")
        
        return JsonResponse({
            'status': 'no_data',
            'message': f'ESP32 data cleared for device {device_id}',
            'device_id': device_id,
            'data_cleared': True,
            'waiting_for': ['BMI', 'Temperature']
        })
    
    # Normal data retrieval logic for specific device
    esp32_data = ESP32_DATA_CACHE.get(device_id, {})
    print(f"DEBUG: Retrieved ESP32 data for {device_id}: {esp32_data}")
    
    if esp32_data:
        has_bmi = esp32_data.get('has_bmi_data', False)
        has_temp = esp32_data.get('has_temperature_data', False)
        is_complete = has_bmi and has_temp
        
        response_data = {
            'status': 'success',
            'data': esp32_data.copy(),
            'device_id': device_id,
            'device_name': DEVICE_STATUS.get(device_id, {}).get('device_name', device_id),
            'has_bmi_data': has_bmi,
            'has_temperature_data': has_temp,
            'last_measurement_type': esp32_data.get('measurement_type', 'UNKNOWN'),
            'is_complete': is_complete,
            'waiting_for': []
        }
        
        # Add info about what data we're still waiting for
        if not has_bmi:
            response_data['waiting_for'].append('BMI')
        if not has_temp:
            response_data['waiting_for'].append('Temperature')
        
        # Only clear if we have BOTH BMI and temperature data
        if is_complete:
            del ESP32_DATA_CACHE[device_id]
            print(f"DEBUG: Auto-cleared COMPLETE data for device {device_id} (BMI + Temperature)")
            response_data['data_cleared'] = True
        else:
            print(f"DEBUG: Keeping PARTIAL data for device {device_id} (BMI: {has_bmi}, Temp: {has_temp})")
            response_data['data_cleared'] = False
        
        return JsonResponse(response_data)
    else:
        return JsonResponse({
            'status': 'no_data',
            'message': f'No ESP32 data available for device {device_id}',
            'device_id': device_id,
            'device_name': DEVICE_STATUS.get(device_id, {}).get('device_name', device_id),
            'waiting_for': ['BMI', 'Temperature']
        })




@csrf_exempt
@require_http_methods(["POST"])
def clear_esp32_data(request):
    """
    FORCE clear ESP32 data cache for specific device or all devices
    """
    global ESP32_DATA_CACHE
    
    try:
        # Handle different content types
        if hasattr(request, 'content_type') and 'application/json' in request.content_type:
            data = json.loads(request.body)
            device_id = data.get('device_id')
        else:
            # Handle form data or beacon data
            try:
                # Try JSON first
                data = json.loads(request.body)
                device_id = data.get('device_id')
            except:
                # Fallback to POST parameters
                device_id = request.POST.get('device_id')
    except Exception as e:
        print(f"DEBUG: Error parsing clear request: {e}")
        device_id = None
    
    # If no device_id specified, show available devices
    if not device_id:
        return JsonResponse({
            'status': 'error',
            'message': 'device_id is required for clearing data',
            'available_devices': list(ESP32_DATA_CACHE.keys()),
            'example': {'device_id': 'BMI_STATION_A'}
        }, status=400)
    
    # ALWAYS clear data for specified device
    if device_id in ESP32_DATA_CACHE:
        cleared_data = ESP32_DATA_CACHE[device_id].copy()
        del ESP32_DATA_CACHE[device_id]
        print(f"DEBUG: POST CLEAR - Forcefully cleared data for device {device_id}: {cleared_data}")
        return JsonResponse({
            'status': 'success',
            'message': f'Data forcefully cleared for device {device_id}',
            'device_id': device_id,
            'cleared_data': cleared_data
        })
    else:
        print(f"DEBUG: POST CLEAR - No data to clear for device {device_id} (cache was already empty)")
        return JsonResponse({
            'status': 'success',  # Still return success even if no data
            'message': f'No data found for device {device_id} (already clear)',
            'device_id': device_id,
            'cleared_data': None
        })


def update_device_online_status():
    global DEVICE_STATUS
    now = timezone.now()
    for device_id, status in DEVICE_STATUS.items():
        if status['last_seen'] and (now - status['last_seen']) > OFFLINE_THRESHOLD:
            DEVICE_STATUS[device_id]['is_online'] = False

def check_device_online_status():
    update_device_online_status()
    
@csrf_exempt
@require_http_methods(["GET"])
def get_device_status(request):
    update_device_online_status()  
    """
    Get status of all ESP32 devices
    """
    global ESP32_DATA_CACHE, DEVICE_STATUS
    
    print(f"DEBUG: get_device_status called - DEVICE_STATUS: {DEVICE_STATUS}")
    
    device_statuses = []
    for device_id, status in DEVICE_STATUS.items():
        has_cached_data = device_id in ESP32_DATA_CACHE
        cached_data = ESP32_DATA_CACHE.get(device_id, {})
        
        device_info = {
            'device_id': device_id,
            'device_name': status['device_name'],
            'is_online': status['is_online'],
            'last_seen': str(status['last_seen']) if status['last_seen'] else None,
            'last_measurement_type': status['last_measurement_type'],
            'measurements_today': status['measurements_today'],
            'has_cached_data': has_cached_data,
            'cached_data_preview': {
                'has_bmi': cached_data.get('has_bmi_data', False),
                'has_temperature': cached_data.get('has_temperature_data', False),
                'last_update': cached_data.get('last_update')
            } if has_cached_data else None
        }
        device_statuses.append(device_info)
    
    print(f"DEBUG: Returning {len(device_statuses)} devices")
    
    return JsonResponse({
        'status': 'success',
        'devices': device_statuses,
        'total_devices': len(device_statuses),
        'online_devices': len([d for d in device_statuses if d['is_online']]),
        'devices_with_data': len([d for d in device_statuses if d['has_cached_data']])
    })



@csrf_exempt
@require_http_methods(["GET"])
def debug_esp32_cache(request):
    """
    Debug view to see current ESP32 cache contents with device separation
    """
    global ESP32_DATA_CACHE, DEVICE_STATUS
    
    return JsonResponse({
        'status': 'debug',
        'cache_contents': ESP32_DATA_CACHE,
        'device_status': DEVICE_STATUS,
        'device_count': len(ESP32_DATA_CACHE),
        'known_devices': list(DEVICE_STATUS.keys()),
        'timestamp': str(timezone.now())
    })

@csrf_exempt
@require_http_methods(["POST"])
def clear_all_esp32_data(request):
    """
    Clear all ESP32 data for all devices (useful for testing)
    """
    global ESP32_DATA_CACHE
    
    device_count = len(ESP32_DATA_CACHE)
    cleared_data = ESP32_DATA_CACHE.copy()
    ESP32_DATA_CACHE.clear()
    
    print(f"DEBUG: Cleared all ESP32 data ({device_count} devices): {cleared_data}")
    
    return JsonResponse({
        'status': 'success',
        'message': f'Cleared data for all {device_count} devices',
        'devices_cleared': device_count,
        'cleared_devices': list(cleared_data.keys()),
        'cleared_data': cleared_data
    })


@csrf_exempt
@require_http_methods(["GET"])
def list_esp32_devices(request):
    """
    List all ESP32 devices that have sent data or are known
    """
    global ESP32_DATA_CACHE, DEVICE_STATUS
    
    devices = []
    all_device_ids = set(list(ESP32_DATA_CACHE.keys()) + list(DEVICE_STATUS.keys()))
    
    for device_id in all_device_ids:
        data = ESP32_DATA_CACHE.get(device_id, {})
        status = DEVICE_STATUS.get(device_id, {})
        
        devices.append({
            'device_id': device_id,
            'device_name': status.get('device_name', device_id),
            'last_update': data.get('last_update'),
            'has_bmi_data': data.get('has_bmi_data', False),
            'has_temperature_data': data.get('has_temperature_data', False),
            'last_measurement_type': data.get('measurement_type', status.get('last_measurement_type')),
            'is_online': status.get('is_online', False),
            'last_seen': str(status.get('last_seen')) if status.get('last_seen') else None,
            'data_preview': {
                'weight': data.get('weight'),
                'height': data.get('height'),
                'temperature': data.get('temperature')
            } if data else None
        })
    
    return JsonResponse({
        'status': 'success',
        'devices': devices,
        'device_count': len(devices),
        'cache_device_count': len(ESP32_DATA_CACHE),
        'known_device_count': len(DEVICE_STATUS)
    })


# DRF versions (if you prefer using Django REST Framework)
@api_view(['POST'])
@permission_classes([AllowAny])
def receive_esp32_data(request):
    """
    API endpoint to receive data from ESP32 using proper serializers (DRF version with multi-device)
    """
    try:
        # Use serializer to validate incoming data
        serializer = ESP32DataSerializer(data=request.data)
        
        if serializer.is_valid():
            validated_data = serializer.validated_data
            device_id = validated_data.get('device_id')
            measurement_type = validated_data.get('measurement_type', 'UNKNOWN').upper()
            
            # Update device tracking (CREATE if doesn't exist)
            if device_id not in DEVICE_STATUS:
                DEVICE_STATUS[device_id] = {
                    'device_name': device_id,
                    'last_seen': None,
                    'is_online': False,
                    'measurements_today': 0,
                    'last_measurement_type': None
                }

            DEVICE_STATUS[device_id]['last_seen'] = timezone.now()
            DEVICE_STATUS[device_id]['is_online'] = True
            DEVICE_STATUS[device_id]['last_measurement_type'] = measurement_type
            DEVICE_STATUS[device_id]['measurements_today'] += 1
            
            # Initialize device data if not exists
            if device_id not in ESP32_DATA_CACHE:
                ESP32_DATA_CACHE[device_id] = {}
            
            # Store data based on measurement type - MERGE with existing data
            if measurement_type == 'BMI':
                ESP32_DATA_CACHE[device_id].update({
                    'weight': validated_data['weight'],
                    'height': validated_data['height'],
                    'bmi': validated_data.get('bmi'),
                    'bmi_category': validated_data.get('bmi_category'),
                    'subject_type': validated_data.get('subject_type'),
                    'sensor_height': validated_data.get('sensor_height'),
                    'bmi_timestamp': str(timezone.now()),
                    'has_bmi_data': True
                })
            elif measurement_type == 'TEMPERATURE':
                ESP32_DATA_CACHE[device_id].update({
                    'temperature': validated_data['temperature'],
                    'temperature_status': validated_data.get('temperature_status'),
                    'temp_timestamp': str(timezone.now()),
                    'has_temperature_data': True
                })
            
            # Always update common fields
            ESP32_DATA_CACHE[device_id].update({
                'device_id': device_id,
                'last_update': str(timezone.now()),
                'esp32_timestamp': validated_data.get('timestamp'),
                'measurement_type': measurement_type
            })
            
            # Create response using response serializer
            response_data = {
                'status': 'success',
                'message': f'{measurement_type} data received and validated successfully from {device_id}',
                'data': validated_data,
                'server_timestamp': timezone.now(),
                'measurement_type': measurement_type,
                'device_id': device_id
            }
            
            response_serializer = ESP32ResponseSerializer(response_data)
            
            # Debug logging
            print(f"=== ESP32 DATA RECEIVED ({measurement_type}) ===")
            print(f"Device: {device_id}")
            print(f"Data: {validated_data}")
            print(f"Cache: {ESP32_DATA_CACHE[device_id]}")
            print("=" * 50)
            
            return Response(response_serializer.data, status=status.HTTP_200_OK)
        
        else:
            # Validation failed
            response_data = {
                'status': 'error',
                'message': 'Data validation failed',
                'errors': serializer.errors,
                'server_timestamp': timezone.now()
            }
            
            response_serializer = ESP32ResponseSerializer(response_data)
            
            print(f"=== VALIDATION FAILED ===")
            print(f"Errors: {serializer.errors}")
            print(f"Raw data: {request.data}")
            print("=" * 30)
            
            return Response(response_serializer.data, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        # Unexpected error
        response_data = {
            'status': 'error',
            'message': f'Server error: {str(e)}',
            'server_timestamp': timezone.now()
        }
        
        response_serializer = ESP32ResponseSerializer(response_data)
        
        print(f"=== SERVER ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Raw data: {request.data}")
        print("=" * 20)
        
        return Response(response_serializer.data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    except Exception as e:
        # Unexpected error
        response_data = {
            'status': 'error',
            'message': f'Server error: {str(e)}',
            'server_timestamp': timezone.now()
        }
        
        response_serializer = ESP32ResponseSerializer(response_data)
        
        print("=== SERVER ERROR ===")
        print("Error: {str(e)}")
        print("Raw data: {request.data}")
        print("=" * 20)
        
        return Response(response_serializer.data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
def get_esp32_data(request):
    """
    API endpoint for the webpage to get the latest ESP32 data (DRF version with device selection)
    """
    device_id = request.GET.get('device_id')
    
    if not device_id:
        # Return list of available devices
        return Response({
            'status': 'device_selection_required',
            'message': 'Please specify device_id parameter',
            'available_devices': list(DEVICE_STATUS.keys()),
            'example_url': '/api/esp32/data/?device_id=BMI_STATION_A'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    esp32_data = ESP32_DATA_CACHE.get(device_id, None)
    
    if esp32_data:
        response_data = {
            'status': 'success',
            'message': f'ESP32 data found for device {device_id}',
            'data': esp32_data,
            'device_id': device_id,
            'server_timestamp': timezone.now()
        }
        
        response_serializer = ESP32ResponseSerializer(response_data)
        return Response(response_serializer.data, status=status.HTTP_200_OK)
    else:
        response_data = {
            'status': 'no_data',
            'message': f'No ESP32 data available for device {device_id}',
            'device_id': device_id,
            'server_timestamp': timezone.now()
        }
        
        response_serializer = ESP32ResponseSerializer(response_data)
        return Response(response_serializer.data, status=status.HTTP_404_NOT_FOUND)

#added for hardware - end



#def view_vaccine_stocks(request):
    """Updated medicine/vaccine stocks view with better filtering"""
    
    # Get user's barangay if applicable
    user_barangay = None
    debug_info = []
    
    debug_info.append("Current user: {request.user}")
    debug_info.append("User email: {request.user.email}")
    debug_info.append("User is authenticated: {request.user.is_authenticated}")
    
    if request.user.is_authenticated:
        # Check each model type
        debug_info.append("Checking Account model...")
        try:
            account = Account.objects.get(email=request.user.email)
            user_barangay = account.barangay
            debug_info.append("✓ Found in Account: {account.email}, Barangay: {user_barangay}")
        except Account.DoesNotExist:
            debug_info.append("✗ Not found in Account model")
            
        if not user_barangay:
            debug_info.append("Checking BHW model...")
            try:
                bhw = BHW.objects.get(email=request.user.email)
                user_barangay = bhw.barangay
                debug_info.append("✓ Found in BHW: {bhw.email}, Barangay: {user_barangay}")
            except BHW.DoesNotExist:
                debug_info.append("✗ Not found in BHW model")
                
        if not user_barangay:
            debug_info.append("Checking BNS model...")
            try:
                bns = BNS.objects.get(email=request.user.email)
                user_barangay = bns.barangay
                debug_info.append("✓ Found in BNS: {bns.email}, Barangay: {user_barangay}")
            except BNS.DoesNotExist:
                debug_info.append("✗ Not found in BNS model")
                
        if not user_barangay:
            debug_info.append("Checking Midwife model...")
            try:
                midwife = Midwife.objects.get(email=request.user.email)
                user_barangay = midwife.barangay
                debug_info.append("✓ Found in Midwife: {midwife.email}, Barangay: {user_barangay}")
            except Midwife.DoesNotExist:
                debug_info.append("✗ Not found in Midwife model")
                
        if not user_barangay:
            debug_info.append("Checking Nurse model...")
            try:
                nurse = Nurse.objects.get(email=request.user.email)
                user_barangay = nurse.barangay
                debug_info.append("✓ Found in Nurse: {nurse.email}, Barangay: {user_barangay}")
            except Nurse.DoesNotExist:
                debug_info.append("✗ Not found in Nurse model")
                
        if not user_barangay:
            debug_info.append("Checking Parent model...")
            try:
                parent = Parent.objects.get(email=request.user.email)
                user_barangay = parent.barangay
                debug_info.append("✓ Found in Parent: {parent.email}, Barangay: {user_barangay}")
            except Parent.DoesNotExist:
                debug_info.append("✗ Not found in Parent model")

    debug_info.append("Final user_barangay: {user_barangay}")

    # Check all vaccine stocks in database first
    all_stocks = VaccineStock.objects.all()
    debug_info.append("Total vaccine stocks in database: {all_stocks.count()}")
    
    for stock in all_stocks:
        if stock.barangay:
            debug_info.append("Stock: {stock.vaccine_name} - Barangay: {stock.barangay.name}")
        else:
            debug_info.append("Stock: {stock.vaccine_name} - No barangay assigned")

    # Filter stocks by barangay if user has one
    if user_barangay:
        stocks = VaccineStock.objects.filter(barangay=user_barangay)
        debug_info.append("Filtering by barangay: {user_barangay}")
        debug_info.append("Found {stocks.count()} stocks for this barangay")
    else:
        stocks = VaccineStock.objects.all()
        debug_info.append("No barangay filter applied. Showing all {stocks.count()} stocks")

    # Print all debug info
    for info in debug_info:
        print(info)

    # Statistics
    total_vaccines = stocks.count()
    total_stock = sum(stock.total_stock for stock in stocks)
    available_stock = sum(stock.available_stock for stock in stocks)
    low_stock_count = sum(1 for stock in stocks if stock.available_stock < 10)

    context = {
        'vaccine_stocks': stocks,
        'total_vaccines': total_vaccines,
        'total_stock': total_stock,
        'available_stock': available_stock,
        'low_stock_count': low_stock_count,
        'user_barangay': user_barangay,
        'debug_info': debug_info,  # Pass debug info to template
    }
    
    return render(request, 'HTML/vaccine_stocks.html', context)


#@login_required
#def add_stock(request):
    """Add medicine/vaccine stock - updated to include deworming and vitamin A"""
    if request.method == 'POST':
        name = request.POST.get('vaccine_name')
        
        try:
            quantity = int(request.POST.get('quantity'))
        except (ValueError, TypeError):
            messages.error(request, "Invalid quantity. Please enter a valid number.")
            return redirect('view_vaccine_stocks')

        # Define valid medicine/vaccine options
        valid_options = [
            'BCG',
            'Hepatitis B',
            'Pentavalent (DPT-Hep B HiB)',
            'Oral Polio Vaccine',
            'Inactivated Polio Vaccine',
            'Pneumococcal Conjugate Vaccine',
            'Measles Mumps and Rubella',
            'Deworming Tablets',
            'Vitamin A Capsules'
        ]

        if name not in valid_options:
            messages.error(request, "Invalid medicine/vaccine name: {name}")
            return redirect('view_vaccine_stocks')

        try:
            # Get user's barangay using the same logic as view_vaccine_stocks
            user_barangay = None
            
            if request.user.is_authenticated:
                try:
                    account = Account.objects.get(email=request.user.email)
                    user_barangay = account.barangay
                except Account.DoesNotExist:
                    try:
                        bhw = BHW.objects.get(email=request.user.email)
                        user_barangay = bhw.barangay
                    except BHW.DoesNotExist:
                        try:
                            bns = BNS.objects.get(email=request.user.email)
                            user_barangay = bns.barangay
                        except BNS.DoesNotExist:
                            try:
                                midwife = Midwife.objects.get(email=request.user.email)
                                user_barangay = midwife.barangay
                            except Midwife.DoesNotExist:
                                try:
                                    nurse = Nurse.objects.get(email=request.user.email)
                                    user_barangay = nurse.barangay
                                except Nurse.DoesNotExist:
                                    try:
                                        parent = Parent.objects.get(email=request.user.email)
                                        user_barangay = parent.barangay
                                    except Parent.DoesNotExist:
                                        pass

            print("Adding stock - User barangay: {user_barangay}")

            # Build query filters
            stock_filter = {'vaccine_name': name}
            if user_barangay:
                stock_filter['barangay'] = user_barangay

            print("Stock filter: {stock_filter}")

            # Try to find existing stock
            try:
                stock = VaccineStock.objects.get(**stock_filter)
                stock.total_stock += quantity
                stock.available_stock += quantity
                stock.save()
                
                messages.success(request, "Stock updated for {name}. Added {quantity} units.")
                print("Updated existing stock for {name} in barangay {user_barangay}")
                
            except VaccineStock.DoesNotExist:
                # Create new stock entry
                stock_data = {
                    'vaccine_name': name,
                    'total_stock': quantity,
                    'available_stock': quantity,
                    'barangay': user_barangay  # This will be None if no barangay
                }
                
                new_stock = VaccineStock.objects.create(**stock_data)
                messages.success(request, "New medicine/vaccine '{name}' added with {quantity} units.")
                print("Created new stock for {name} in barangay {user_barangay}")

        except Exception as e:
            messages.error(request, "An error occurred: {str(e)}")
            print("Error adding stock: {str(e)}")

    return redirect('view_vaccine_stocks')

# Helper function to get medicine categories for reporting
#def get_medicine_categories():
    """Returns categorized list of medicines/vaccines"""
    return {
        'vaccines': [
            'BCG',
            'Hepatitis B', 
            'Pentavalent (DPT-Hep B HiB)',
            'Oral Polio Vaccine',
            'Inactivated Polio Vaccine',
            'Pneumococcal Conjugate Vaccine',
            'Measles Mumps and Rubella'
        ],
        'medicines_supplements': [
            'Deworming Tablets',
            'Vitamin A Capsules'
        ]
    }

def email_endorsement(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    """Email endorsement view with proper barangay filtering"""
    
    # Initialize
    user_barangay = None
    account = None
    debug_info = []
    debug_info.append(f"Current user: {request.user}")
    debug_info.append(f"User email: {request.user.email}")

    if request.user.is_authenticated:
        # Try to find the user in various models
        try:
            account = Account.objects.select_related('barangay').get(email=request.user.email)
            user_barangay = account.barangay
            debug_info.append(f"✓ Found in Account: {account.email}, Barangay: {user_barangay}")
        except Account.DoesNotExist:
            debug_info.append("✗ Not found in Account model")
            # Try BHW
            try:
                bhw = BHW.objects.select_related('barangay').get(email=request.user.email)
                user_barangay = bhw.barangay
                account = type('MockAccount', (), {
                    'email': bhw.email,
                    'barangay': bhw.barangay,
                    'clean_full_name': bhw.full_name,
                    'clean_address': getattr(bhw, 'editable_address', None)
                })()
                debug_info.append(f"✓ Found in BHW: {bhw.email}, Barangay: {user_barangay}")
            except BHW.DoesNotExist:
                # Try BNS
                try:
                    bns = BNS.objects.select_related('barangay').get(email=request.user.email)
                    user_barangay = bns.barangay
                    account = type('MockAccount', (), {
                        'email': bns.email,
                        'barangay': bns.barangay,
                        'clean_full_name': bns.full_name,
                        'clean_address': getattr(bns, 'editable_address', None)
                    })()
                    debug_info.append(f"✓ Found in BNS: {bns.email}, Barangay: {user_barangay}")
                except BNS.DoesNotExist:
                    # Try Midwife
                    try:
                        midwife = Midwife.objects.select_related('barangay').get(email=request.user.email)
                        user_barangay = midwife.barangay
                        account = type('MockAccount', (), {
                            'email': midwife.email,
                            'barangay': midwife.barangay,
                            'clean_full_name': midwife.full_name,
                            'clean_address': getattr(midwife, 'editable_address', None)
                        })()
                        debug_info.append(f"✓ Found in Midwife: {midwife.email}, Barangay: {user_barangay}")
                    except Midwife.DoesNotExist:
                        # Try Nurse
                        try:
                            nurse = Nurse.objects.select_related('barangay').get(email=request.user.email)
                            user_barangay = nurse.barangay
                            account = type('MockAccount', (), {
                                'email': nurse.email,
                                'barangay': nurse.barangay,
                                'clean_full_name': nurse.full_name,
                                'clean_address': getattr(nurse, 'editable_address', None)
                            })()
                            debug_info.append(f"✓ Found in Nurse: {nurse.email}, Barangay: {user_barangay}")
                        except Nurse.DoesNotExist:
                            # Try Parent
                            try:
                                parent = Parent.objects.select_related('barangay').get(email=request.user.email)
                                user_barangay = parent.barangay
                                account = type('MockAccount', (), {
                                    'email': parent.email,
                                    'barangay': parent.barangay,
                                    'clean_full_name': parent.full_name,
                                    'clean_address': getattr(parent, 'editable_address', None)
                                })()
                                debug_info.append(f"✓ Found in Parent: {parent.email}, Barangay: {user_barangay}")
                            except Parent.DoesNotExist:
                                debug_info.append("✗ User not found in any model")

    debug_info.append(f"Final user_barangay: {user_barangay}")

    # Check for missing account or barangay
    if not account:
        messages.error(request, "User account not found. Please contact administrator.")
        return redirect('dashboard')
    
    if not user_barangay:
        messages.error(request, "No barangay assigned to your account. Please contact administrator.")
        return redirect('dashboard')

    # Filter parents from the same barangay
    parents = Parent.objects.filter(barangay=user_barangay).exclude(email__isnull=True)
    debug_info.append(f"Found {parents.count()} parents in barangay {user_barangay}")

    for parent in parents:
        debug_info.append(f"Parent: {parent.full_name} ({parent.email}) - Barangay: {parent.barangay}")

    # Print debug info
    for info in debug_info:
        print(info)

    if request.method == 'POST':
        from_email = request.POST.get('from_email')
        to_email = request.POST.get('to_email')
        subject = request.POST.get('subject')
        message = request.POST.get('message')

        # Validate recipient
        try:
            recipient_parent = Parent.objects.get(email=to_email, barangay=user_barangay)
            debug_info.append(f"Recipient validation passed: {recipient_parent.email} is in {user_barangay}")
        except Parent.DoesNotExist:
            messages.error(request, "Invalid recipient. You can only send emails to parents in your barangay.")
            return redirect('email_endorsement')

        try:
            # Prepare role badge class
            role_map = {
                'BHW': 'bhw',
                'BNS': 'bns',
                'Midwife': 'midwife',
                'Nurse': 'nurse'
            }
            user_role = getattr(account, 'role', 'Account')
            role_class = role_map.get(user_role, 'default')
            
            # Get current date
            from datetime import datetime
            current_date = datetime.now().strftime('%B %d, %Y')
            
            # Prepare context for HTML email
            email_context = {
                'full_name': recipient_parent.full_name,
                'email': recipient_parent.email,
                'sex': getattr(recipient_parent, 'sex', 'N/A'),
                'role': user_role,
                'role_class': role_class,
                'role_name': user_role,
                'barangay': user_barangay,
                'current_date': current_date,
                'message_body': message,
                'sender_name': getattr(account, 'clean_full_name', account.email)
            }
            
            # Generate HTML email
            html_message = render_endorsement_email_html(email_context)
            plain_message = render_endorsement_email_text(email_context)
            
            send_mail(
                subject=subject,
                message=plain_message,
                from_email=from_email,
                recipient_list=[to_email],
                html_message=html_message,
                fail_silently=False
            )
            messages.success(request, f"Endorsement email sent successfully to {to_email}.")
            return redirect('dashboard')
        except Exception as e:
            messages.error(request, f"Error sending email: {e}")
            return redirect('email_endorsement')

    return render(request, 'HTML/email_endorsement.html', {
        'from_email': account.email,
        'account': account,
        'parents': parents,
        'user_barangay': user_barangay,
        'clean_full_name': getattr(account, 'clean_full_name', None),
        'clean_address': getattr(account, 'clean_address', None),
        'debug_info': debug_info
    })


def render_endorsement_email_html(context):
    """Generate HTML endorsement email"""
    html = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>PPMS Endorsement</title>
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
            
            body {{
                font-family: 'Inter', Arial, sans-serif;
                background-color: #f9fafb;
                padding: 40px 20px;
                color: #334155;
                line-height: 1.6;
            }}
            
            .container {{
                max-width: 560px;
                margin: 0 auto;
                background: white;
                border-radius: 12px;
                overflow: hidden;
                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
            }}
            
            .header {{
                padding: 24px 32px;
                text-align: center;
                color: #111827;
                border-bottom: 4px solid #198754;
            }}
            .header h1 {{
                font-size: 24px;
                font-weight: 600;
                margin: 0;
                color: #111827;
            }}
            .header p {{
                font-size: 16px;
                margin: 4px 0 0 0;
                color: #6b7280;
            }}
            
            .content {{
                padding: 32px;
            }}
            
            .greeting {{
                font-size: 18px;
                margin-bottom: 16px;
                color: #1e293b;
                text-align: left;
                margin: 0 0 16px 0;
            }}
            
            .salutation {{
                font-size: 16px;
                margin-bottom: 16px;
                color: #64748b;
                text-align: left;
                margin: 0 0 16px 0;
            }}
            
            .message {{
                font-size: 16px;
                margin-bottom: 32px;
                color: #64748b;
                text-align: left;
                white-space: pre-wrap;
                margin: 0;
                padding: 0;
            }}
            
            .footer {{
                background: #f1f5f9;
                padding: 32px;
                text-align: center;
                border-top: 1px solid #e2e8f0;
            }}
            
            .footer h3 {{
                font-size: 18px;
                font-weight: 600;
                color: #1e293b;
                margin-bottom: 8px;
            }}
            
            .footer p {{
                font-size: 14px;
                color: #64748b;
                margin-bottom: 4px;
            }}
            
            .footer-divider {{
                margin: 24px 0;
                height: 1px;
                background: #e2e8f0;
            }}
            
            .footer-small {{
                font-size: 12px;
                color: #94a3b8;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>PPMS Cluster 4</h1>
                <p>Imus City Healthcare Management</p>
            </div>
            
            <div class="content">
                <div class="greeting">
                    Hello <strong>{context['full_name']}</strong>,
                </div>
                
                <div class="salutation">
                    
                </div>
                
                <div class="message">
{context['message_body'].strip()}
                </div>
            </div>
            
            <div class="footer">
                <h3>PPMS Cluster 4</h3>
                <p>Imus City Healthcare Management</p>
                <div class="footer-divider"></div>
                <p class="footer-small">
                    This is an automated message. Please do not reply.<br>
                    © 2025 PPMS Cluster 4. All rights reserved.
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    return html

def render_endorsement_email_text(context):
    """Generate plain text endorsement email"""
    text = f"""
PPMS Endorsement

Hello {context['full_name']},

{context['message_body'].strip()}

FROM: {context['role'].upper()}
Name: {context['sender_name']}
Barangay: {context['barangay'].name}
Date Sent: {context['current_date']}

PPMS Cluster 4
Imus City Healthcare Management

This is an automated message. Please do not reply.
© 2025 PPMS Cluster 4. All rights reserved.
    """
    return text.strip()



def generate_immunization_report(request):
    """Generate PDF report for immunization schedules and vaccination records (fixed total of 15 required vaccines)."""
    user = request.user
    try:
        account = Account.objects.select_related('barangay').get(email=user.email)
    except Account.DoesNotExist:
        return HttpResponse("Account not found.", status=404)

    user_role_lower = account.user_role.lower() if account.user_role else ''
    is_authorized = any(role in user_role_lower for role in [
        'bhw', 'health worker', 'bns', 'nutrition', 'nutritional', 'scholar', 'midwife', 'admin'
    ])
    if not is_authorized:
        messages.error(request, f"Your role ({account.user_role}) is not authorized to generate immunization reports.")
        return redirect('dashboard')

    barangay = account.barangay
    if not barangay:
        return HttpResponse("No barangay assigned to this account.", status=404)

    # Month filter
    month_str = request.GET.get("month")  # expected format: YYYY-MM
    month_filter = None
    month_display = "All Records"
    if month_str:
        try:
            month_filter = datetime.strptime(month_str, "%Y-%m")
            month_display = month_filter.strftime("%B %Y")
        except ValueError:
            month_filter = None
            month_display = "All Records"

    preschoolers = Preschooler.objects.filter(
        is_archived=False
    ).select_related('parent_id').prefetch_related('vaccination_schedules')

    today = date.today()

    # ✅ Updated summary categories
    vaccination_summary = {
        'Fully Vaccinated': 0,
        'Incomplete Vaccine': 0,
        'Not Vaccinated': 0,
        'Overdue': 0,
    }

    preschoolers_data = []

    # Unique vaccines (still used for report listing)
    scheduled_vaccines = VaccinationSchedule.objects.all().values_list(
        'vaccine_name', flat=True
    ).distinct().order_by('vaccine_name')
    required_vaccines = list(scheduled_vaccines) if scheduled_vaccines else [
        'BCG', 'Hepatitis B', 'DPT', 'OPV', 'MMR', 'Pneumococcal', 'Rotavirus'
    ]

    REQUIRED_VACCINE_COUNT = 15  # ✅ fixed total number of vaccines

    for p in preschoolers:
        schedules = [s for s in p.vaccination_schedules.all() if s.status == 'completed']
        if month_filter:
            schedules = [
                s for s in schedules
                if s.scheduled_date and
                   s.scheduled_date.year == month_filter.year and
                   s.scheduled_date.month == month_filter.month
            ]

        # Skip if no vaccination records at all
        if not schedules:
            vaccination_summary['Not Vaccinated'] += 1
            continue

        # Calculate age
        age_years = today.year - p.birth_date.year - (
            (today.month, today.day) < (p.birth_date.month, p.birth_date.day)
        )
        age_months = (today.year - p.birth_date.year) * 12 + today.month - p.birth_date.month

        # ✅ Vaccination logic based on 15 required vaccines
        total_completed = len(schedules)
        total_required = REQUIRED_VACCINE_COUNT

        if total_completed >= total_required:
            vaccination_status = "Fully Vaccinated"
            vaccination_summary['Fully Vaccinated'] += 1
        else:
            vaccination_status = "Incomplete Vaccine"
            vaccination_summary['Incomplete Vaccine'] += 1

        # Parent info
        parent_name = p.parent_id.full_name if p.parent_id else "N/A"
        address = getattr(p.parent_id, 'address', 'N/A') if p.parent_id else getattr(p, 'address', 'N/A') or "N/A"

        # Latest vaccination date
        last_completed = max(
            [s.completion_date or s.administered_date for s in schedules if s.completion_date or s.administered_date],
            default=None
        )
        last_vaccination_date = last_completed.strftime('%m/%d/%Y') if last_completed else "N/A"

        # ✅ Safe f-string for vaccine completion list
        completed_text = "; ".join([
            f"{s.vaccine_name} ({(v_date.strftime('%m/%d/%Y') if (v_date := (s.completion_date or s.administered_date or s.scheduled_date)) else 'N/A')})"
            for s in schedules
        ])


        preschoolers_data.append({
            'name': f"{p.first_name} {p.last_name}",
            'age': f"{age_years} years, {age_months % 12} months",
            'sex': p.sex,
            'vaccination_status': vaccination_status,
            'vaccines_received': f"{min(total_completed, total_required)}/{total_required}",
            'last_vaccination': last_vaccination_date,
            'vaccination_schedule': completed_text,
            'parent_name': parent_name,
            'address': address,
            'overdue_count': 0,  # no overdue tracking yet
        })

    context = {
        'account': account,
        'barangay': barangay,
        'preschoolers': preschoolers_data,
        'summary': vaccination_summary,
        'required_vaccines': required_vaccines,
        'month_filter': month_str or "All",
    }

    html_string = render_to_string('HTML/immunization_report.html', context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="Immunization-Report.pdf"'
    return response





@admin_required
def generate_admin_report(request):
    month = request.GET.get('month')  # Format: YYYY-MM
    
    if not month:
        return HttpResponse("Month parameter is required", status=400)
    
    # Parse month
    year, month_num = month.split('-')
    month_name = datetime.strptime(month, '%Y-%m').strftime('%B %Y')
    
    # Get the admin account
    account = request.user
    
    # Query all preschoolers across all barangays for the selected month
    preschoolers_qs = Preschooler.objects.filter(
        date_registered__year=year,
        date_registered__month=month_num
    ).select_related('barangay').prefetch_related('vaccination_schedules')
    
    # Barangay summary container
    barangay_summary = defaultdict(lambda: {"total": 0, "fully": 0, "partial": 0, "not": 0})
    
    for child in preschoolers_qs:
        barangay_name = child.barangay.name if child.barangay else "Unassigned"
        
        barangay_summary[barangay_name]["total"] += 1
        
        vaccination_status = child.get_vaccination_status() if hasattr(child, 'get_vaccination_status') else "Not Vaccinated"
        
        if vaccination_status == "Fully Vaccinated":
            barangay_summary[barangay_name]["fully"] += 1
        elif vaccination_status == "Partially Vaccinated":
            barangay_summary[barangay_name]["partial"] += 1
        else:
            barangay_summary[barangay_name]["not"] += 1
    
    context = {
        "account": account,
        "month_filter": month_name,
        "barangay_summary": dict(barangay_summary),
    }
    
    # Render template
    html_string = render_to_string("HTML/reportTemplate.html", context)
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="overall_barangay_immunization_report_{month}.pdf"'
    
    return response


def index(request):
    
    return HttpResponse('Welcome to the PPMA Web Application!')



import re

@admin_required
def addbarangay(request):
    if request.method == 'POST':
        name = request.POST.get('barangay-name', '').strip()
        phone_number = request.POST.get('phone-number', '').strip()
        hall_address = request.POST.get('hall-address', '').strip()

        # ✅ Check for empty barangay name
        if not name:
            messages.error(request, "Barangay name is required.")
            barangays = sorted(Barangay.objects.all(), key=lambda x: x.name.upper())
            return render(request, 'HTML/addbarangay.html', {'barangays': barangays})

        # ✅ Check if barangay name already exists
        if Barangay.objects.filter(name__iexact=name).exists():
            messages.error(request, f"A barangay named '{name}' already exists.")
            barangays = sorted(Barangay.objects.all(), key=lambda x: x.name.upper())
            return render(request, 'HTML/addbarangay.html', {'barangays': barangays})

        # ✅ Try saving the barangay
        try:
            Barangay.objects.create(
                name=name,
                phone_number=phone_number,
                hall_address=hall_address,
            )
            messages.success(request, f"Barangay {name} was added successfully!")
            return redirect('addbarangay')
        except Exception as e:
            messages.error(request, "Something went wrong while saving. Please try again.")
            barangays = sorted(Barangay.objects.all(), key=lambda x: x.name.upper())
            return render(request, 'HTML/addbarangay.html', {'barangays': barangays})
    
    # GET request - fetch and sort barangays (case-insensitive)
    barangays = sorted(Barangay.objects.all(), key=lambda x: x.name.upper())
    return render(request, 'HTML/addbarangay.html', {'barangays': barangays})



@admin_required
def Admin(request):
    # Import required modules at the top
    from django.db.models import Count, Q
    from datetime import timedelta
    from dateutil.relativedelta import relativedelta
    
    # Count health workers - include all health worker roles
    health_worker_roles = ['BHW', 'Barangay Nutritional Scholar', 'Midwife', 'Nurse']
    health_worker_count = Account.objects.filter(
        user_role__in=health_worker_roles,
        is_validated=True  # Only count validated health workers
    ).count() or 0

    # Get preschoolers and unvalidated accounts
    preschoolers = Preschooler.objects.all()
    accounts = Account.objects.filter(is_validated=False)

    # ✅ Total Fully Vaccinated Preschoolers
    # Define required vaccines with their required doses (from immunization schedule)
    required_vaccines = {
        'BCG Vaccine': 1,
        'Hepatitis B Vaccine': 1,
        'Pentavalent Vaccine': 3,
        'Oral Polio Vaccine': 3,
        'Inactivated Polio Vaccine': 2,
        'Pneumococcal Conjugate Vaccine': 3,
        'Measles, Mumps, and Rubella': 2,
    }
    # Total required doses: 1+1+3+3+2+3+2 = 15 doses
    
    total_vaccinated = 0
    
    # Get all active preschoolers
    active_preschoolers = Preschooler.objects.filter(is_archived=False)
    
    for preschooler in active_preschoolers:
        # Get all completed vaccinations for this preschooler
        completed_vaccinations = VaccinationSchedule.objects.filter(
            preschooler=preschooler,
            status='completed'
        ).values('vaccine_name').annotate(
            doses_completed=Count('id')
        )
        
        # Create a dictionary of completed vaccines
        completed_dict = {
            vac['vaccine_name']: vac['doses_completed'] 
            for vac in completed_vaccinations
        }
        
        # Check if preschooler has completed ALL required vaccines
        is_fully_vaccinated = True
        for vaccine_name, required_doses in required_vaccines.items():
            completed_doses = completed_dict.get(vaccine_name, 0)
            if completed_doses < required_doses:
                is_fully_vaccinated = False
                break
        
        if is_fully_vaccinated:
            total_vaccinated += 1

    # Build notifications
    notifications = []

    for acc in accounts:
        notifications.append({
            'type': 'account',
            'full_name': acc.full_name,
            'created_at': acc.created_at,
            'user_role': acc.user_role,
        })

    for child in preschoolers:
        notifications.append({
            'type': 'preschooler',
            'full_name': f"{child.first_name} {child.last_name}",
            'date_registered': child.date_registered,
            'bhw_image': getattr(child.bhw_id.account.profile_photo.image, 'url', None)
                          if child.bhw_id and child.bhw_id.account and hasattr(child.bhw_id.account, 'profile_photo') else None,
        })

    # Process timestamps
    for notif in notifications:
        timestamp = notif.get('created_at') or notif.get('date_registered')
        if timestamp:
            if isinstance(timestamp, date) and not isinstance(timestamp, datetime):
                timestamp = timezone.make_aware(datetime.combine(timestamp, timezone.now().time()))
            elif isinstance(timestamp, datetime) and timezone.is_naive(timestamp):
                timestamp = timezone.make_aware(timestamp)
        else:
            timestamp = timezone.now()
        notif['timestamp'] = timestamp

    notifications.sort(key=lambda x: x['timestamp'], reverse=True)
    latest_notifications = notifications[:15]
    latest_timestamp = latest_notifications[0]['timestamp'] if latest_notifications else None

    # Total preschoolers
    total_preschoolers = Preschooler.objects.filter(is_archived=False).count() or 0

    barangays = Barangay.objects.all()

    # Pie chart: nutritional status
    status_totals = {
        'Severely Wasted': 0,
        'Wasted': 0,
        'Normal': 0,
        'Risk of Overweight': 0,
        'Overweight': 0,
        'Obese': 0,
    }

    # Table summary by barangay
    summary = []
    today = date.today()
    
    for brgy in barangays:
        preschoolers_in_barangay = Preschooler.objects.filter(
            barangay=brgy,
            is_archived=False
        ).prefetch_related('bmi_set')

        nutritional_summary = {
            'severely_wasted': 0,
            'wasted': 0,
            'normal': 0,
            'risk_overweight': 0,
            'overweight': 0,
            'obese': 0,
        }

        preschooler_count = preschoolers_in_barangay.count()

        for p in preschoolers_in_barangay:
            latest_bmi = p.bmi_set.order_by('-date_recorded').first()
            if latest_bmi:
                try:
                    # Compute age in months
                    birth_date = p.birth_date
                    age_years = today.year - birth_date.year
                    age_months = today.month - birth_date.month
                    if today.day < birth_date.day:
                        age_months -= 1
                    if age_months < 0:
                        age_years -= 1
                        age_months += 12
                    total_age_months = age_years * 12 + age_months

                    # Compute BMI and classify
                    bmi_value = calculate_bmi(latest_bmi.weight, latest_bmi.height)
                    z = bmi_zscore(p.sex, total_age_months, bmi_value)
                    category = classify_bmi_for_age(z)

                    if category == "Severely Wasted":
                        nutritional_summary['severely_wasted'] += 1
                        status_totals['Severely Wasted'] += 1
                    elif category == "Wasted":
                        nutritional_summary['wasted'] += 1
                        status_totals['Wasted'] += 1
                    elif category == "Normal":
                        nutritional_summary['normal'] += 1
                        status_totals['Normal'] += 1
                    elif category == "Risk of Overweight":
                        nutritional_summary['risk_overweight'] += 1
                        status_totals['Risk of Overweight'] += 1
                    elif category == "Overweight":
                        nutritional_summary['overweight'] += 1
                        status_totals['Overweight'] += 1
                    elif category == "Obese":
                        nutritional_summary['obese'] += 1
                        status_totals['Obese'] += 1

                except Exception as e:
                    print(f"⚠️ BMI classification error for preschooler {p.id}: {e}")

        # Add barangay data even if it has 0 preschoolers
        summary.append({
            'barangay': brgy.name,
            'preschooler_count': preschooler_count,
            **nutritional_summary
        })

    # ✅ FIXED: Enhanced Vaccination Trend Data - Count Doses by Administered Date
    # Get filter parameters from request
    filter_month = request.GET.get('filter_month')
    
    # Determine center month
    if filter_month:
        try:
            center_date = datetime.strptime(filter_month, '%Y-%m').date()
        except ValueError:
            center_date = timezone.now().date().replace(day=1)  # Current month as fallback
    else:
        center_date = timezone.now().date().replace(day=1)  # Current month by default
    
    # Generate 11 months: center month ±5 months
    trend_labels = []
    monthly_registered_trend = []
    vaccinated_doses_trend = []
    all_months_data = []  # For JavaScript filtering
    
    # ✅ Count vaccine doses by their ADMINISTERED_DATE (when they were actually given)
    for i in range(-5, 6):  # -5 to +5 inclusive
        target_month = center_date + relativedelta(months=i)
        
        # Calculate next month for range queries
        next_month = target_month + relativedelta(months=1)
        
        # Month label
        month_label = target_month.strftime('%b %Y')
        trend_labels.append(month_label)
        
        # Monthly registrations for this specific month
        monthly_registered = Preschooler.objects.filter(
            date_registered__gte=target_month,
            date_registered__lt=next_month,
            is_archived=False
        ).count()
        monthly_registered_trend.append(monthly_registered)
        
        # ✅ Count vaccine doses completed in this month
        # Uses completion_date (datetime) to determine when vaccine was given
        completed_doses_count = VaccinationSchedule.objects.filter(
            status='completed',
            completion_date__isnull=False,  # Must have a completion date
            completion_date__date__gte=target_month,  # Extract date from datetime
            completion_date__date__lt=next_month,
            preschooler__is_archived=False
        ).count()
        
        vaccinated_doses_trend.append(completed_doses_count)
        
        # Store month data for JavaScript
        all_months_data.append({
            'month': target_month.strftime('%Y-%m'),
            'label': month_label,
            'registered': monthly_registered,
            'vaccinated': completed_doses_count
        })

    # Generate extended data for JavaScript (±12 months for smooth transitions)
    extended_months_data = []
    for i in range(-12, 13):  # -12 to +12 months
        target_month = center_date + relativedelta(months=i)  # ✅ Use center_date, not timezone.now()
        next_month = target_month + relativedelta(months=1)
        
        monthly_registered = Preschooler.objects.filter(
            date_registered__gte=target_month,
            date_registered__lt=next_month,
            is_archived=False
        ).count()
        
        # ✅ Count vaccine doses by completion_date in this month
        completed_doses_count = VaccinationSchedule.objects.filter(
            status='completed',
            completion_date__isnull=False,
            completion_date__date__gte=target_month,
            completion_date__date__lt=next_month,
            preschooler__is_archived=False
        ).count()
        
        extended_months_data.append({
            'month': target_month.strftime('%Y-%m'),
            'label': target_month.strftime('%b %Y'),
            'registered': monthly_registered,
            'vaccinated': completed_doses_count
        })

    vaccination_trend_data = {
        'labels': trend_labels,
        'registered': monthly_registered_trend,
        'vaccinated': vaccinated_doses_trend,
        'center_month': center_date.strftime('%Y-%m'),
        'all_months': extended_months_data  # For JavaScript filtering
    }

    # Prepare data for Barangay Bar Chart
    barangay_chart_data = {
        'barangays': [row['barangay'] for row in summary],
        'severely_wasted': [row['severely_wasted'] for row in summary],
        'wasted': [row['wasted'] for row in summary],
        'normal': [row['normal'] for row in summary],
        'risk_overweight': [row['risk_overweight'] for row in summary],
        'overweight': [row['overweight'] for row in summary],
        'obese': [row['obese'] for row in summary]
    }

    return render(request, 'HTML/Admindashboard.html', {
        'health_worker_count': health_worker_count,
        'notifications': latest_notifications,
        'latest_notif_timestamp': latest_timestamp.isoformat() if latest_timestamp else '',
        'total_preschoolers': total_preschoolers,
        'total_vaccinated': total_vaccinated,
        'barangay_summary': summary,
        'nutritional_data': {
            'labels': list(status_totals.keys()),
            'values': list(status_totals.values())
        },
        'vaccination_trend_data': vaccination_trend_data,
        'barangay_chart_data': barangay_chart_data,
        'current_filter_month': center_date.strftime('%Y-%m')  # For the date picker
    })


    

def archived(request):
    """Updated archived view with auto-archive check"""
    
    # Run auto-archive check here too
    auto_archived_count = auto_archive_aged_preschoolers()
    if auto_archived_count > 0:
        print(f"AUTO-ARCHIVED: {auto_archived_count} preschoolers in archived view")
    
    # Get the current user's account
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        account = Account.objects.select_related('profile_photo', 'barangay').get(email=request.user.email)
    except Account.DoesNotExist:
        messages.error(request, "Account not found. Please contact administrator.")
        return redirect('login')
    
    # Get user info for barangay filtering (if not admin)
    user_email = request.user.email
    raw_role = account.user_role.strip().lower() if account.user_role else ''
    
    if raw_role == 'admin':
        archived_preschoolers_qs = Preschooler.objects.filter(is_archived=True).select_related('barangay', 'parent_id')
    else:
        # Filter by user's barangay
        archived_preschoolers_qs = Preschooler.objects.filter(
            is_archived=True, 
            barangay=account.barangay
        ).select_related('barangay', 'parent_id')

    # Order by most recently archived
    archived_preschoolers_qs = archived_preschoolers_qs.order_by('-date_registered')

    # Paginate archived preschoolers
    paginator = Paginator(archived_preschoolers_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Convert to JSON with current age data
    archived_json = json.dumps([
        {
            "id": p.preschooler_id,
            "first_name": p.first_name,
            "last_name": p.last_name,
            "name": f"{p.first_name} {p.last_name}",
            "age": p.age_in_months if p.age_in_months else p.age,
            "age_display": f"{p.age_in_months} months" if p.age_in_months else f"{p.age} years",
            "barangay": p.barangay.name if p.barangay else "N/A",
            "gender": p.sex,
            "birthdate": str(p.birth_date),
            "parent_name": f"{p.parent_id.first_name} {p.parent_id.last_name}" if p.parent_id else "N/A",
            "archived_date": p.date_registered.strftime("%Y-%m-%d") if p.date_registered else "N/A",
        } for p in page_obj
    ])

    return render(request, 'HTML/archived.html', {
        'account': account,
        'archived_preschoolers_json': archived_json,
        'archived_page': page_obj,
        'notifications': [],
        'latest_notif_timestamp': None,
    })

def auto_archive_aged_preschoolers():
    """
    Simple auto-archive function - just sets is_archived=True for 60+ month olds
    """
    # Get all active preschoolers
    active_preschoolers = Preschooler.objects.filter(is_archived=False)
    archived_count = 0
    today = date.today()
    
    for preschooler in active_preschoolers:
        # Use the model's age_in_months property
        age_in_months = preschooler.age_in_months
        
        # Archive if 60+ months old
        if age_in_months and age_in_months >= 60:
            preschooler.is_archived = True
            preschooler.save()
            
            print("AUTO-ARCHIVED: {preschooler.first_name} {preschooler.last_name} - Age: {age_in_months} months")
            archived_count += 1
    
    return archived_count


def archived_details(request):

    if not request.user.is_authenticated:
        return redirect('login')
    
    return render(request, 'HTML/archived_details.html')

def dashboard(request):
    # ✅ Redirect to login if not authenticated
    if not request.user.is_authenticated:
        return redirect('login')

    # ✅ Get the current user's account (with profile photo)
    account = get_object_or_404(
        Account.objects.select_related('profile_photo'),
        email=request.user.email
    )

    # ✅ Active preschoolers (barangay-specific)
    preschoolers = Preschooler.objects.filter(
        is_archived=False,
        barangay=account.barangay
    ).prefetch_related('bmi_set')

    preschooler_count = preschoolers.count()

    # ✅ Archived preschoolers (barangay-specific)
    archived_preschooler_count = Preschooler.objects.filter(
        is_archived=True,
        barangay=account.barangay
    ).count()

    # ✅ Parents (barangay-specific)
    parent_accounts = Parent.objects.filter(
        barangay=account.barangay
    ).distinct().order_by('-created_at')

    parent_count = parent_accounts.count()
    print("DEBUG Dashboard ({account.barangay}): Parent count = {parent_count}")

    # ✅ Nutritional Summary via WHO BMI-for-age Z-scores
    nutritional_summary = {
        'severely_wasted': 0,
        'wasted': 0,
        'normal': 0,
        'risk_of_overweight': 0,
        'overweight': 0,
        'obese': 0,
    }

    preschoolers_with_bmi = 0
    today = date.today()

    for p in preschoolers:
        latest_bmi = p.bmi_set.order_by('-date_recorded').first()
        if latest_bmi:
            try:
                # --- Compute age in months ---
                birth_date = p.birth_date
                age_years = today.year - birth_date.year
                age_months = today.month - birth_date.month
                if today.day < birth_date.day:
                    age_months -= 1
                if age_months < 0:
                    age_years -= 1
                    age_months += 12
                total_age_months = age_years * 12 + age_months

                # --- Compute BMI and classify ---
                bmi_value = calculate_bmi(latest_bmi.weight, latest_bmi.height)
                z = bmi_zscore(p.sex, total_age_months, bmi_value)
                category = classify_bmi_for_age(z)

                preschoolers_with_bmi += 1
                if category == "Severely Wasted":
                    nutritional_summary['severely_wasted'] += 1
                elif category == "Wasted":
                    nutritional_summary['wasted'] += 1
                elif category == "Normal":
                    nutritional_summary['normal'] += 1
                elif category == "Risk of Overweight":
                    nutritional_summary['risk_of_overweight'] += 1
                elif category == "Overweight":
                    nutritional_summary['overweight'] += 1
                elif category == "Obese":
                    nutritional_summary['obese'] += 1

            except Exception as e:
                print("⚠️ BMI classification error for preschooler {p.id}: {e}")

    # ✅ Calculate percentages
    nutritional_percentages = {}
    if preschoolers_with_bmi > 0:
        for key, value in nutritional_summary.items():
            percentage = round((value / preschoolers_with_bmi) * 100, 1)
            nutritional_percentages[key] = percentage
    else:
        nutritional_percentages = {key: 0 for key in nutritional_summary.keys()}

    # ✅ Prepare pie chart data (matching WHO categories)
    pie_chart_data = {
        'labels': [
            'Severely Wasted',
            'Wasted',
            'Normal',
            'Risk of Overweight',
            'Overweight',
            'Obese'
        ],
        'values': [
            nutritional_summary['severely_wasted'],
            nutritional_summary['wasted'],
            nutritional_summary['normal'],
            nutritional_summary['risk_of_overweight'],
            nutritional_summary['overweight'],
            nutritional_summary['obese']
        ],
        'percentages': [
            nutritional_percentages['severely_wasted'],
            nutritional_percentages['wasted'],
            nutritional_percentages['normal'],
            nutritional_percentages['risk_of_overweight'],
            nutritional_percentages['overweight'],
            nutritional_percentages['obese']
        ],
        'colors': ['#e74c3c', '#f39c12', '#27ae60', '#f1c40f', '#e67e22', '#c0392b']
    }

    # ✅ Recent parent account notifications (barangay-specific)
    notifications = []
    seen_ids = set()

    for parent in parent_accounts:
        if parent.parent_id not in seen_ids:
            notifications.append({
                'type': 'account',
                'id': parent.parent_id,
                'full_name': parent.full_name,
                'user_role': 'Parent',
                'timestamp': parent.created_at,
            })
            seen_ids.add(parent.parent_id)

    notifications.sort(key=lambda x: x['timestamp'], reverse=True)
    latest_notif_timestamp = notifications[0]['timestamp'] if notifications else None

    # ✅ Fetch active announcements (global)
    try:
        announcements = Announcement.objects.filter(
            is_active=True
        ).order_by('-created_at')[:10]
    except Exception:
        announcements = []

    return render(request, 'HTML/dashboard.html', {
        'account': account,
        'full_name': account.full_name,
        'preschooler_count': preschooler_count,
        'archived_preschooler_count': archived_preschooler_count,
        'parent_count': parent_count,
        'nutritional_summary': nutritional_summary,
        'nutritional_percentages': nutritional_percentages,
        'preschoolers_with_bmi': preschoolers_with_bmi,
        'pie_chart_data': pie_chart_data,
        'notifications': notifications[:15],
        'latest_notif_timestamp': latest_notif_timestamp.isoformat() if latest_notif_timestamp else '',
        'announcements': announcements,
    })





@csrf_exempt
@login_required
def upload_preschooler_photo(request, preschooler_id):
    """Handle preschooler profile photo upload to Cloudinary"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)

    try:
        preschooler = get_object_or_404(Preschooler, pk=preschooler_id)
        
        # Verify parent ownership
        account = Account.objects.get(email=request.user.email)
        if preschooler.parent_id.email != account.email:
            return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)

        photo_file = request.FILES.get('profile_photo')
        if not photo_file:
            return JsonResponse({'status': 'error', 'message': 'No photo provided'}, status=400)

        # Validate file type
        if not photo_file.content_type.startswith('image/'):
            return JsonResponse({'status': 'error', 'message': 'Invalid file type'}, status=400)

        # Validate file size (max 10MB)
        if photo_file.size > 10 * 1024 * 1024:
            return JsonResponse({'status': 'error', 'message': 'File too large (max 10MB)'}, status=400)

        # ✅ Upload to Cloudinary
        upload_result = cloudinary.uploader.upload(
            photo_file,
            folder='preschooler_photos/',
            public_id=f'preschooler_{preschooler.preschooler_id}',
            overwrite=True,
            transformation=[
                {'width': 500, 'height': 500, 'crop': 'fill', 'gravity': 'face'},
                {'quality': 'auto:good'}
            ]
        )

        # Get Cloudinary URL and public ID
        cloudinary_url = upload_result.get('secure_url')
        cloudinary_public_id = upload_result.get('public_id')

        # Delete old image from Cloudinary if exists
        if preschooler.cloudinary_public_id:
            try:
                cloudinary.uploader.destroy(preschooler.cloudinary_public_id)
            except Exception as e:
                print(f"Error deleting old image: {e}")

        # Update preschooler record
        preschooler.profile_photo = cloudinary_url
        preschooler.cloudinary_public_id = cloudinary_public_id
        preschooler.save()

        return JsonResponse({
            'status': 'success',
            'message': 'Photo uploaded successfully',
            'new_photo_url': cloudinary_url
        })

    except Account.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Account not found'}, status=404)
    except Exception as e:
        print(f"Upload error: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
        
@csrf_exempt
def upload_cropped_photo(request):
    """Handle cropped photo upload to Cloudinary"""
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'unauthorized'}, status=403)

    if request.method == 'POST':
        try:
            image = request.FILES.get('cropped_image')
            account = Account.objects.get(email=request.user.email)

            if not image:
                return JsonResponse({'status': 'error', 'message': 'No image provided'}, status=400)

            # ✅ Upload to Cloudinary with optimizations
            upload_result = cloudinary.uploader.upload(
                image,
                folder='profile_photos/',
                public_id=f'profile_{account.account_id}',
                overwrite=True,
                transformation=[
                    {'width': 500, 'height': 500, 'crop': 'fill', 'gravity': 'face'},
                    {'quality': 'auto:good'}
                ]
            )

            # ✅ Get Cloudinary URL and public ID
            cloudinary_url = upload_result.get('secure_url')
            cloudinary_public_id = upload_result.get('public_id')

            # ✅ Update or create ProfilePhoto
            if hasattr(account, 'profile_photo'):
                # Delete old image from Cloudinary if exists
                if account.profile_photo.cloudinary_public_id:
                    try:
                        cloudinary.uploader.destroy(account.profile_photo.cloudinary_public_id)
                    except Exception as e:
                        print(f"Error deleting old image: {e}")
                
                account.profile_photo.image = cloudinary_url
                account.profile_photo.cloudinary_public_id = cloudinary_public_id
                account.profile_photo.save()
            else:
                ProfilePhoto.objects.create(
                    account=account,
                    image=cloudinary_url,
                    cloudinary_public_id=cloudinary_public_id
                )

            return JsonResponse({
                'status': 'success',
                'image_url': cloudinary_url,
                'message': 'Profile photo uploaded successfully'
            })

        except Account.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Account not found'}, status=404)
        except Exception as e:
            print(f"Upload error: {e}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


from django.views.decorators.csrf import csrf_exempt



@csrf_exempt
def login(request):
    # If already logged in, redirect based on role
    if request.user.is_authenticated:
        role = request.session.get('user_role', '').lower()
        if role == 'parent':
            return redirect('parent_dashboard')
        elif role == 'admin':
            return redirect('Admindashboard')
        else:
            return redirect('dashboard')

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        fcm_token = request.POST.get('fcm_token', None)

        # Hardcoded admin login
        if email.lower() == 'admin@gmail.com' and password == 'admin123':
            request.session['user_role'] = 'admin'
            request.session['full_name'] = 'Admin'
            request.session['email'] = 'admin@gmail.com'
            return redirect('Admindashboard')

        # Authenticate using email as username
        user = authenticate(request, username=email, password=password)

        if user is not None:
            try:
                from .models import Account, Parent, FCMToken
                account = Account.objects.get(email=email)

                if account.is_rejected:
                    messages.error(request, "Your account has been rejected by the admin.")
                    return render(request, 'HTML/login.html')

                # ADMIN LOGIN
                if account.user_role.lower() == 'admin':
                    # ✅ CHECK FOR FORCED PASSWORD CHANGE **BEFORE** LOGGING IN
                    if account.must_change_password:
                        # Store email in session but DON'T log them in yet
                        request.session['email'] = email
                        messages.warning(request, "You must change your password before continuing.")
                        return redirect('change_password_first')
                    
                    # If password doesn't need changing, proceed with normal login
                    auth_login(request, user)
                    account.last_activity = timezone.now()
                    account.save(update_fields=['last_activity'])

                    request.session['email'] = account.email
                    request.session['user_role'] = 'admin'
                    request.session['full_name'] = account.full_name or f"{account.first_name} {account.last_name}"
                    request.session['contact_number'] = account.contact_number or ''

                    return redirect('Admindashboard')

                # HEALTH WORKER LOGIN
                elif (account.user_role.lower() == 'healthworker' or 
                    account.user_role.lower() == 'bhw' or
                    account.user_role.lower() in ['bns', 'barangay nutritional scholar'] or
                    account.user_role.lower() == 'midwife' or
                    account.user_role.lower() == 'nurse'):

                    if not account.is_validated:
                        messages.error(request, "Your account is still pending admin validation.")
                        return render(request, 'HTML/login.html')

                    auth_login(request, user)
                    account.last_activity = timezone.now()
                    account.save(update_fields=['last_activity'])

                    request.session['email'] = account.email
                    request.session['user_role'] = account.user_role.lower()
                    request.session['full_name'] = account.full_name
                    request.session['contact_number'] = account.contact_number

                    if fcm_token:
                        account.fcm_token = fcm_token
                        account.save(update_fields=["fcm_token"])
                        FCMToken.objects.update_or_create(
                            token=fcm_token,
                            defaults={
                                'account': account,
                                'device_type': 'android',
                                'is_active': True,
                            }
                        )

                    return redirect('dashboard')

                # PARENT LOGIN
                elif account.user_role.lower() == 'parent':
                    try:
                        parent = Parent.objects.get(email=email)
                    except Parent.DoesNotExist:
                        messages.error(request, "Parent account not found.")
                        return render(request, 'HTML/login.html')

                    if parent.must_change_password:
                        request.session['email'] = email
                        return redirect('change_password_first')

                    auth_login(request, user)
                    account.last_activity = timezone.now()
                    account.save(update_fields=['last_activity'])

                    request.session['email'] = account.email
                    request.session['user_role'] = 'parent'
                    request.session['full_name'] = account.full_name
                    request.session['contact_number'] = account.contact_number

                    if fcm_token:
                        account.fcm_token = fcm_token
                        account.save(update_fields=["fcm_token"])
                        FCMToken.objects.update_or_create(
                            token=fcm_token,
                            defaults={
                                'account': account,
                                'device_type': 'android',
                                'is_active': True,
                            }
                        )

                    return redirect('parent_dashboard')

                else:
                    messages.warning(request, f"Unknown user role: {account.user_role}. Please contact support.")
                    return redirect('login')

            except Account.DoesNotExist:
                messages.error(request, "Account record not found. Please contact support.")
                return render(request, 'HTML/login.html')

        else:
            messages.error(request, "Invalid email or password.")

    # Fetch announcements
    try:
        from .models import Announcement
        announcements = Announcement.objects.filter(is_active=True).order_by('-created_at')[:5]
    except Exception as e:
        announcements = []

    return render(request, 'HTML/login.html', {'announcements': announcements})






def logout_view(request):
    if request.user.is_authenticated:
        try:
            account = Account.objects.get(email=request.user.email)
            account.last_activity = timezone.now()
            account.save(update_fields=['last_activity'])
        except Account.DoesNotExist:
            pass

    logout(request)
    return redirect('login')


from .models import BMI  # siguraduhin na naka-import

def parent_dashboard(request):
    if not request.user.is_authenticated:
        return redirect('login')

    account = get_object_or_404(Account.objects.select_related('profile_photo'), email=request.user.email)
    try:
        parent = Parent.objects.get(email=account.email)
        full_name = parent.full_name
    except Parent.DoesNotExist:
        full_name = account.full_name or "Unknown User"
    preschoolers_raw = Preschooler.objects.filter(parent_id__email=account.email)

    # Compute age per preschooler
    preschoolers = []
    today = date.today()

    for p in preschoolers_raw:
        birth_date = p.birth_date

        # Calculate years
        age_years = today.year - birth_date.year
        # Calculate months
        age_months = today.month - birth_date.month
        # Calculate days
        age_days = today.day - birth_date.day

        # Adjust if days are negative
        if age_days < 0:
            age_months -= 1
            if today.month == 1:
                last_month = 12
                last_year = today.year - 1
            else:
                last_month = today.month - 1
                last_year = today.year

            from calendar import monthrange
            days_in_last_month = monthrange(last_year, last_month)[1]
            age_days += days_in_last_month

        # Adjust if months are negative
        if age_months < 0:
            age_years -= 1
            age_months += 12

        # ✅ Convert total age in months (needed for WHO classification)
        total_age_months = age_years * 12 + age_months

        # ✅ Get latest BMI status using WHO BMI-for-age Z-scores
        latest_bmi = BMI.objects.filter(preschooler_id=p).order_by('-date_recorded').first()
        bmi_status = None
        if latest_bmi:
            try:
                bmi_value = calculate_bmi(latest_bmi.weight, latest_bmi.height)
                z = bmi_zscore(p.sex, total_age_months, bmi_value)
                bmi_status = classify_bmi_for_age(z)
            except Exception as e:
                print("Error calculating BMI/Z-score for preschooler {p.id}: {e}")
                bmi_status = p.nutritional_status  # fallback

        preschoolers.append({
            'data': p,
            'age_years': age_years,
            'age_months': age_months,
            'age_days': age_days,
            'bmi_status': bmi_status or p.nutritional_status  # fallback if no BMI record
        })

    if not request.session.get('first_login_shown', False):
        invalid_values = {"na", "n/a", "none", "null", "--"}
        
        # Clean name parts
        name_parts = [
            account.first_name,
            account.middle_name,
            account.user.last_name if account.user else "",
            account.suffix,
        ]
        
        clean_parts = [
            part for part in name_parts 
            if part and part.strip().lower() not in invalid_values
        ]
        
        clean_full_name = " ".join(clean_parts).strip()
        
        messages.success(request, f" Welcome, {clean_full_name}!")
        request.session['first_login_shown'] = True

    # ✅ FIXED: Filter upcoming schedules - only show 'scheduled' status
    upcoming_schedules = VaccinationSchedule.objects.filter(
        preschooler__in=preschoolers_raw,
        status='scheduled'
    ).order_by('scheduled_date')

    # ✅ Fetch active announcements for parents
    try:
        announcements = Announcement.objects.filter(
            is_active=True
        ).order_by('-created_at')[:10]
    except Exception as e:
        announcements = []

    return render(request, 'HTML/parent_dashboard.html', {
        'account': account,
        'full_name': "{account.first_name} {account.last_name}".strip(),
        'preschoolers': preschoolers,
        'upcoming_schedules': upcoming_schedules,
        'announcements': announcements,
        'today': today,
    })



import threading

def send_notifications_async(parent, account, preschooler, vaccine_name, dose_number, required_doses, immunization_date, next_schedule, schedule):
    """Background thread to handle both email + push notifications"""
    try:
        # === Email notification ===
        if parent.email:
            try:
                subject = f"[PPMS] Vaccination Scheduled for {preschooler.first_name}"
                
                # Prepare context for HTML email
                email_context = {
                    'parent_name': parent.full_name,
                    'child_name': f"{preschooler.first_name} {preschooler.last_name}",
                    'vaccine_name': vaccine_name,
                    'dose_number': dose_number,
                    'required_doses': required_doses,
                    'scheduled_date': immunization_date,
                    'next_schedule': next_schedule
                }
                
                # Generate HTML and plain text emails
                html_message = render_vaccination_schedule_email_html(email_context)
                plain_message = render_vaccination_schedule_email_text(email_context)

                send_mail(
                    subject,
                    plain_message,
                    settings.DEFAULT_FROM_EMAIL,
                    [parent.email],
                    html_message=html_message,
                    fail_silently=False
                )
                logger.info(f"[ASYNC] Email sent to {parent.email}")
            except Exception as email_error:
                logger.error(f"[ASYNC] Email failed for {parent.email}: {email_error}")

        # === Push notification ===
        if account and account.fcm_token:
            try:
                notification_title = f"Vaccination Scheduled for {preschooler.first_name}"
                notification_body = (
                    f"{vaccine_name} (Dose {dose_number}/{required_doses}) "
                    f"scheduled for {immunization_date}"
                )

                notification_data = {
                    "type": "vaccination_schedule",
                    "preschooler_id": str(preschooler.preschooler_id),
                    "preschooler_name": f"{preschooler.first_name} {preschooler.last_name}",
                    "vaccine_name": vaccine_name,
                    "dose_number": str(dose_number),
                    "total_doses": str(required_doses),
                    "scheduled_date": str(immunization_date),
                    "schedule_id": str(schedule.id)
                }

                logger.info(f"[ASYNC] Sending push to {parent.email}")
                PushNotificationService.send_push_notification(
                    token=account.fcm_token,
                    title=notification_title,
                    body=notification_body,
                    data=notification_data
                )
            except Exception as push_error:
                logger.error(f"[ASYNC] Push failed for {parent.email}: {push_error}")
        else:
            logger.warning(f"[ASYNC] No FCM token found for {parent.email}")

    except Exception as e:
        logger.error(f"[ASYNC] Notification error for {parent.email}: {e}")


def render_vaccination_schedule_email_html(context):
    """Generate HTML vaccination schedule email"""
    html = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>PPMS Vaccination Schedule</title>
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
            
            body {{
                font-family: 'Inter', Arial, sans-serif;
                background-color: #f9fafb;
                padding: 40px 20px;
                color: #334155;
                line-height: 1.6;
            }}
            
            .container {{
                max-width: 560px;
                margin: 0 auto;
                background: white;
                border-radius: 12px;
                overflow: hidden;
                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
            }}
            
            .header {{
                padding: 24px 32px;
                text-align: center;
                color: #111827;
                border-bottom: 4px solid #198754;
            }}
            .header h1 {{
                font-size: 24px;
                font-weight: 600;
                margin: 0;
                color: #111827;
            }}
            .header p {{
                font-size: 16px;
                margin: 4px 0 0 0;
                color: #6b7280;
            }}
            
            .content {{
                padding: 32px;
            }}
            
            .greeting {{
                font-size: 18px;
                margin: 0 0 16px 0;
                color: #1e293b;
                text-align: left;
            }}
            
            .salutation {{
                font-size: 16px;
                margin: 0 0 16px 0;
                color: #64748b;
                text-align: left;
            }}
            
            .message {{
                font-size: 16px;
                margin: 0 0 24px 0;
                color: #64748b;
                text-align: left;
                padding: 0;
            }}
            
            .schedule-details {{
                background-color: #f1f5f9;
                padding: 16px;
                border-radius: 8px;
                margin-bottom: 24px;
                border-left: 4px solid #198754;
            }}
            
            .detail-row {{
                display: flex;
                justify-content: space-between;
                padding: 8px 0;
                font-size: 14px;
                color: #334155;
            }}
            
            .detail-label {{
                font-weight: 600;
                color: #1e293b;
            }}
            
            .detail-value {{
                color: #64748b;
            }}
            
            .footer {{
                background: #f1f5f9;
                padding: 32px;
                text-align: center;
                border-top: 1px solid #e2e8f0;
            }}
            
            .footer h3 {{
                font-size: 18px;
                font-weight: 600;
                color: #1e293b;
                margin-bottom: 8px;
            }}
            
            .footer p {{
                font-size: 14px;
                color: #64748b;
                margin-bottom: 4px;
            }}
            
            .footer-divider {{
                margin: 24px 0;
                height: 1px;
                background: #e2e8f0;
            }}
            
            .footer-small {{
                font-size: 12px;
                color: #94a3b8;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>PPMS Cluster 4</h1>
                <p>Imus City Healthcare Management</p>
            </div>
            
            <div class="content">
                <div class="greeting">
                    Hello <strong>{context['parent_name']}</strong>,
                </div>
                
                <div class="salutation">
                    Dear Parent,
                </div>
                
                <div class="message">
                    A vaccination schedule has been created for your child <strong>{context['child_name']}</strong>.
                </div>
                
                <div class="schedule-details">
                    <div class="detail-row">
                        <span class="detail-label">Vaccine:</span>
                        <span class="detail-value">{context['vaccine_name']}</span>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Dose:</span>
                        <span class="detail-value">{context['dose_number']} of {context['required_doses']}</span>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Scheduled Date:</span>
                        <span class="detail-value">{context['scheduled_date']}</span>
                    </div>
                    {'<div class="detail-row"><span class="detail-label">Next Schedule:</span><span class="detail-value">' + context['next_schedule'] + '</span></div>' if context.get('next_schedule') else ''}
                </div>
                
                <div class="message">
                    Please mark your calendar and proceed to your barangay health center on the scheduled date. If you have any concerns or need to reschedule, please contact us immediately.
                </div>
            </div>
            
            <div class="footer">
                <h3>PPMS Cluster 4</h3>
                <p>Imus City Healthcare Management</p>
                <div class="footer-divider"></div>
                <p class="footer-small">
                    This is an automated message. Please do not reply.<br>
                    © 2025 PPMS Cluster 4. All rights reserved.
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    return html


def render_vaccination_schedule_email_text(context):
    """Generate plain text vaccination schedule email"""
    next_schedule_text = f"Next Schedule: {context['next_schedule']}\n" if context.get('next_schedule') else ""
    
    text = f"""
PPMS Vaccination Schedule

Hello {context['parent_name']},

Dear Parent,

A vaccination schedule has been created for your child {context['child_name']}.

VACCINATION DETAILS:
Vaccine: {context['vaccine_name']}
Dose: {context['dose_number']} of {context['required_doses']}
Scheduled Date: {context['scheduled_date']}
{next_schedule_text}
Please mark your calendar and proceed to your barangay health center on the scheduled date. If you have any concerns or need to reschedule, please contact us immediately.

PPMS Cluster 4
Imus City Healthcare Management

This is an automated message. Please do not reply.
© 2025 PPMS Cluster 4. All rights reserved.
    """
    return text.strip()


@login_required
def add_schedule(request, preschooler_id):
    """Add vaccination schedule with improved async notification handling"""
    logger.info(f"[DEBUG] Entered add_schedule view for preschooler {preschooler_id}")

    try:
        from .models import Preschooler, VaccinationSchedule, Account
        preschooler = get_object_or_404(Preschooler, pk=preschooler_id)
        logger.info(f"[DEBUG] Found preschooler: {preschooler.first_name} {preschooler.last_name}")
    except Exception as e:
        logger.error(f"[DEBUG] Error getting preschooler: {e}")
        messages.error(request, "Preschooler not found")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    if request.method != "POST":
        logger.warning("[DEBUG] Request method is not POST")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    # Extract POST values
    vaccine_name = request.POST.get("vaccine_name")
    required_doses = request.POST.get("required_doses", "1")
    vaccine_doses = request.POST.get("vaccine_doses", "1")
    immunization_date = request.POST.get("immunization_date")
    next_schedule = request.POST.get("next_vaccine_schedule")
    current_dose = request.POST.get("current_dose")

    if not vaccine_name or not immunization_date:
        logger.warning("[DEBUG] Missing required fields")
        messages.error(request, "Please fill in all required fields.")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    try:
        existing_completed = VaccinationSchedule.objects.filter(
            preschooler=preschooler,
            vaccine_name=vaccine_name,
            status='completed'
        ).count()

        existing_pending = VaccinationSchedule.objects.filter(
            preschooler=preschooler,
            vaccine_name=vaccine_name,
            status__in=['scheduled', 'rescheduled']
        ).count()

        dose_number = int(current_dose) if current_dose else existing_completed + existing_pending + 1

        schedule = VaccinationSchedule.objects.create(
            preschooler=preschooler,
            vaccine_name=vaccine_name,
            doses=dose_number,
            required_doses=int(required_doses) if required_doses else 1,
            scheduled_date=immunization_date,
            next_vaccine_schedule=next_schedule or None,
            status='scheduled',
            confirmed_by_parent=False
        )
        logger.info(f"[DEBUG] VaccinationSchedule saved: {schedule.id}")

        messages.success(
            request,
            f"Vaccination schedule for {vaccine_name} (Dose {dose_number}) added successfully!"
        )

        # === Fire off async notifications ===
        parents = preschooler.parents.all()
        for parent in parents:
            account = Account.objects.filter(email=parent.email).first()
            threading.Thread(
                target=send_notifications_async,
                args=(parent, account, preschooler, vaccine_name, dose_number, required_doses, immunization_date, next_schedule, schedule)
            ).start()

    except Exception as e:
        logger.error(f"[ERROR] Failed to save schedule or notify: {e}")
        messages.error(request, f"Error: {str(e)}")

    return redirect(request.META.get("HTTP_REFERER", "/"))


def send_nutrition_notifications_async(parents, preschooler, service_type, dose_number, total_doses, service_date, notes, schedule):
    """Background task: send email + push notifications for nutrition services"""
    from .models import Account
    from .services import PushNotificationService

    for parent in parents:
        try:
            # === Email Notification ===
            if parent.email:
                try:
                    subject = f"[PPMS] Nutrition Service Scheduled for {preschooler.first_name}"

                    notes_line = f"Notes: {notes}\n" if notes else ""

                    message = (
                        f"Dear {parent.full_name},\n\n"
                        f"A nutrition service appointment has been scheduled for your child, "
                        f"{preschooler.first_name} {preschooler.last_name}.\n\n"
                        f"Service Type: {service_type}\n"
                        f"Dose: {dose_number} of {total_doses}\n"
                        f"Scheduled Date: {service_date}\n"
                        f"{notes_line}"
                        "\nPlease bring your child on the scheduled date.\n"
                        "You can confirm completion on your dashboard.\n\n"
                        "Thank you,\nPPMS System"
                    )

                    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [parent.email], fail_silently=False)
                    logger.info(f"[ASYNC] Nutrition email sent to {parent.email}")

                except Exception as e:
                    logger.error(f"[ASYNC] Failed to send nutrition email to {parent.email}: {e}")

            # === Push Notification ===
            try:
                account = Account.objects.filter(email=parent.email).first()
                if account and account.fcm_token:
                    service_emoji = "🍎" if service_type == "Vitamin A" else "💊"
                    title = f"{service_emoji} Nutrition Service Scheduled for {preschooler.first_name}"
                    body = f"{service_type} (Dose {dose_number}/{total_doses}) scheduled for {service_date}"

                    data = {
                        "type": "nutrition_service_schedule",
                        "preschooler_id": str(preschooler.preschooler_id),
                        "preschooler_name": f"{preschooler.first_name} {preschooler.last_name}",
                        "service_type": service_type,
                        "dose_number": str(dose_number),
                        "total_doses": str(total_doses),
                        "scheduled_date": str(service_date),
                        "schedule_id": str(schedule.id),
                        "notes": notes or ""
                    }

                    PushNotificationService.send_push_notification(
                        token=account.fcm_token,
                        title=title,
                        body=body,
                        data=data
                    )
                    logger.info(f"[ASYNC] Nutrition push sent to {parent.email}")
                else:
                    logger.warning(f"[ASYNC] No FCM token for {parent.email}")

            except Exception as e:
                logger.error(f"[ASYNC] Failed to send nutrition push to {parent.email}: {e}")

        except Exception as e:
            logger.error(f"[ASYNC] Error handling parent {parent.email}: {e}")



@login_required
def schedule_nutrition_service(request, preschooler_id):
    """Schedule nutrition service with async push/email notifications"""
    from .models import Preschooler, NutritionService

    logger.info(f"[DEBUG] Entered schedule_nutrition_service view for preschooler {preschooler_id}")
    try:
        preschooler = get_object_or_404(Preschooler, pk=preschooler_id)
        logger.info(f"[DEBUG] Found preschooler: {preschooler.first_name} {preschooler.last_name}")
    except Exception as e:
        logger.error(f"[DEBUG] Error getting preschooler: {e}")
        messages.error(request, "Preschooler not found")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    if request.method != "POST":
        logger.warning("[DEBUG] Request method is not POST")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    # Extract POST values
    service_type = request.POST.get("service_type")
    service_date = request.POST.get("service_date")
    notes = request.POST.get("notes", "")

    if not service_type or not service_date:
        messages.error(request, "Please fill in all required fields.")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    try:
        # Calculate dose
        existing_completed = NutritionService.objects.filter(
            preschooler=preschooler,
            service_type=service_type,
            status='completed'
        ).count()
        existing_pending = NutritionService.objects.filter(
            preschooler=preschooler,
            service_type=service_type,
            status__in=['scheduled', 'rescheduled']
        ).count()
        dose_number = existing_completed + existing_pending + 1
        total_doses = 10

        # Save nutrition service schedule
        schedule = NutritionService.objects.create(
            preschooler=preschooler,
            service_type=service_type,
            scheduled_date=service_date,
            status='scheduled',
            notes=notes,
            confirmed_by_parent=False
        )
        logger.info(f"[DEBUG] NutritionService saved: {schedule.id}")

        messages.success(
            request,
            f"Nutrition service schedule for {service_type} (Dose {dose_number}) added successfully!"
        )

        # === Async notifications ===
        parents = preschooler.parents.all()
        if parents:
            threading.Thread(
                target=send_nutrition_notifications_async,
                args=(parents, preschooler, service_type, dose_number, total_doses, service_date, notes, schedule)
            ).start()
        else:
            logger.warning("No parents found for this preschooler")

    except Exception as e:
        logger.error(f"[ERROR] Failed to save nutrition schedule or send notifications: {e}")
        messages.error(request, f"Error: {str(e)}")

    return redirect(request.META.get("HTTP_REFERER", "/"))

@csrf_exempt
@login_required
def update_nutrition_status(request, schedule_id):
    """Update nutrition service status with enhanced notifications"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        import json
        from .models import NutritionService, Account

        data = json.loads(request.body)
        status = data.get('status')
        enhanced_notifications = data.get('enhanced_notifications', False)
        
        schedule = get_object_or_404(NutritionService, id=schedule_id)
        preschooler = schedule.preschooler
        
        logger.info(f"[DEBUG] Updating nutrition status for schedule {schedule_id} to {status}")
        
        if status == 'completed':
            from django.utils import timezone
            schedule.status = 'completed'
            schedule.completion_date = timezone.now()
            schedule.save()
            
            # Check if more doses are needed
            completed_doses = NutritionService.objects.filter(
                preschooler=preschooler,
                service_type=schedule.service_type,
                status='completed'
            ).count()
            
            total_doses = 10  # Standard
            needs_next_dose = completed_doses < total_doses
            fully_completed = completed_doses >= total_doses
            
            logger.info(f"[DEBUG] Completed doses: {completed_doses}/{total_doses}")
            
            # === Run notifications asynchronously ===
            if enhanced_notifications:
                def send_notifications():
                    parents = preschooler.parents.all()
                    for parent in parents:
                        # --- EMAIL ---
                        if parent.email:
                            try:
                                if fully_completed:
                                    subject = f"[PPMS] {schedule.service_type} Treatment Complete for {preschooler.first_name}"
                                    message = (
                                        f"Dear {parent.full_name},\n\n"
                                        f"Congratulations! Your child {preschooler.first_name} {preschooler.last_name} "
                                        f"has completed all {total_doses} doses of {schedule.service_type}.\n\n"
                                        f"Treatment completed on: {schedule.completion_date.strftime('%B %d, %Y at %I:%M %p')}\n\n"
                                        f"Thank you for ensuring your child received proper nutrition care.\n\n"
                                        f"PPMS System"
                                    )
                                else:
                                    subject = f"[PPMS] {schedule.service_type} Dose Completed for {preschooler.first_name}"
                                    message = (
                                        f"Dear {parent.full_name},\n\n"
                                        f"Your child {preschooler.first_name} {preschooler.last_name} "
                                        f"has received dose {completed_doses} of {total_doses} "
                                        f"for {schedule.service_type}.\n\n"
                                        f"Completed on: {schedule.completion_date.strftime('%B %d, %Y at %I:%M %p')}\n\n"
                                        f"Remaining doses needed: {total_doses - completed_doses}\n\n"
                                        f"Thank you,\nPPMS System"
                                    )

                                send_mail(
                                    subject, 
                                    message, 
                                    settings.DEFAULT_FROM_EMAIL,
                                    [parent.email], 
                                    fail_silently=False
                                )
                                logger.info(f"[DEBUG] Email sent to {parent.email}")
                            except Exception as email_error:
                                logger.error(f"[DEBUG] Email sending failed: {email_error}")
                        
                        # --- PUSH ---
                        try:
                            account = Account.objects.filter(email=parent.email).first()
                            if account and account.fcm_token:
                                service_emoji = "🍎" if schedule.service_type == "Vitamin A" else "💊"
                                
                                if fully_completed:
                                    notification_title = f"🏆 {schedule.service_type} Treatment Complete!"
                                    notification_body = f"{preschooler.first_name} completed all {total_doses} doses"
                                else:
                                    notification_title = f"{service_emoji} {schedule.service_type} Dose Complete"
                                    notification_body = f"Dose {completed_doses}/{total_doses} completed for {preschooler.first_name}"
                                
                                notification_data = {
                                    "type": "nutrition_service_completed",
                                    "preschooler_id": str(preschooler.preschooler_id),
                                    "preschooler_name": f"{preschooler.first_name} {preschooler.last_name}",
                                    "service_type": schedule.service_type,
                                    "completed_doses": str(completed_doses),
                                    "total_doses": str(total_doses),
                                    "fully_completed": str(fully_completed).lower(),
                                    "completion_date": schedule.completion_date.isoformat() if schedule.completion_date else ""
                                }
                                
                                push_result = PushNotificationService.send_push_notification(
                                    token=account.fcm_token,
                                    title=notification_title,
                                    body=notification_body,
                                    data=notification_data
                                )
                                logger.info(f"[DEBUG] Push result: {push_result}")
                        except Exception as push_error:
                            logger.error(f"[DEBUG] Push error: {push_error}")

                # 🔹 Run notifications in a background thread
                threading.Thread(target=send_notifications, daemon=True).start()
            
            # Return response immediately
            if fully_completed:
                return JsonResponse({
                    'success': True,
                    'message': f'All {schedule.service_type} doses completed successfully!',
                    'fully_completed': True,
                    'service_type': schedule.service_type,
                    'completed_doses': completed_doses,
                    'total_doses': total_doses
                })
            elif needs_next_dose:
                return JsonResponse({
                    'success': True,
                    'message': f'Dose {completed_doses} completed! {total_doses - completed_doses} doses remaining.',
                    'needs_next_dose': True,
                    'service_type': schedule.service_type,
                    'completed_doses': completed_doses,
                    'total_doses': total_doses
                })
            else:
                return JsonResponse({
                    'success': True,
                    'message': 'Nutrition service status updated successfully!',
                    'completed_doses': completed_doses,
                    'total_doses': total_doses
                })
        
        else:
            # Handle other status updates
            schedule.status = status
            schedule.save()
            return JsonResponse({
                'success': True,
                'message': f'Status updated to {status} successfully!'
            })
            
    except Exception as e:
        logger.error(f"[ERROR] Failed to update nutrition status: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Error updating status: {str(e)}'
        })


@login_required
def reschedule_nutrition_service(request, schedule_id):
    """Reschedule nutrition service with async push/email notification handling"""
    logger.info(f"[DEBUG] Entered reschedule_nutrition_service view for schedule {schedule_id}")

    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request method"})

    try:
        from .models import NutritionService, Account
        import json, threading
        
        # Parse JSON data
        data = json.loads(request.body)
        new_date = data.get('new_date')
        reschedule_reason = data.get('reschedule_reason', '')
        enhanced_notifications = data.get('enhanced_notifications', False)
        
        logger.info(f"[DEBUG] Reschedule data: new_date={new_date}, reason={reschedule_reason}, enhanced={enhanced_notifications}")

        # Get the nutrition service schedule
        schedule = get_object_or_404(NutritionService, pk=schedule_id)
        preschooler = schedule.preschooler
        old_date = schedule.scheduled_date
        
        # Validate required fields
        if not new_date or not reschedule_reason.strip():
            return JsonResponse({
                "success": False, 
                "message": "Please provide both a new date and reason for rescheduling."
            })

        # Update the schedule
        schedule.scheduled_date = new_date
        schedule.status = 'rescheduled'
        reschedule_info = f"Rescheduled from {old_date} to {new_date}. Reason: {reschedule_reason}"
        schedule.notes = f"{(schedule.notes + '; ') if schedule.notes else ''}{reschedule_info}"
        schedule.save()
        
        logger.info("[DEBUG] Schedule updated successfully")

        # === Run notifications asynchronously ===
        if enhanced_notifications:
            def send_notifications():
                parents = preschooler.parents.all()
                logger.info(f"[DEBUG] Found {parents.count()} parent(s) for reschedule notifications")
                
                for parent in parents:
                    logger.info(f"[DEBUG] Processing notifications for parent: {parent.full_name} ({parent.email})")

                    # --- EMAIL ---
                    if parent.email:
                        try:
                            subject = f"[PPMS] Nutrition Service Rescheduled for {preschooler.first_name}"
                            message = (
                                f"Dear {parent.full_name},\n\n"
                                f"The nutrition service appointment for your child, "
                                f"{preschooler.first_name} {preschooler.last_name}, has been rescheduled.\n\n"
                                f"Service Type: {schedule.service_type}\n"
                                f"Original Date: {old_date}\n"
                                f"New Date: {new_date}\n"
                                f"Reason: {reschedule_reason}\n"
                                f"\nPlease bring your child on the new scheduled date.\n"
                                f"You can view the updated schedule on your dashboard.\n\n"
                                f"Thank you for your understanding,\nPPMS System"
                            )
                            
                            send_mail(
                                subject, 
                                message, 
                                settings.DEFAULT_FROM_EMAIL,
                                [parent.email], 
                                fail_silently=False
                            )
                            logger.info(f"[DEBUG] Reschedule email sent to {parent.email}")
                        except Exception as email_error:
                            logger.error(f"[DEBUG] Reschedule email failed: {email_error}")

                    # --- PUSH ---
                    try:
                        account = Account.objects.filter(email=parent.email).first()
                        if account and account.fcm_token:
                            service_emoji = "🍎" if schedule.service_type == "Vitamin A" else "💊"
                            notification_title = f"{service_emoji} Nutrition Service Rescheduled"
                            notification_body = (
                                f"{schedule.service_type} for {preschooler.first_name} moved to {new_date}"
                            )
                            
                            notification_data = {
                                "type": "nutrition_service_reschedule",
                                "preschooler_id": str(preschooler.preschooler_id),
                                "preschooler_name": f"{preschooler.first_name} {preschooler.last_name}",
                                "service_type": schedule.service_type,
                                "old_date": str(old_date),
                                "new_date": str(new_date),
                                "reschedule_reason": reschedule_reason,
                                "schedule_id": str(schedule.id)
                            }
                            
                            push_result = PushNotificationService.send_push_notification(
                                token=account.fcm_token,
                                title=notification_title,
                                body=notification_body,
                                data=notification_data
                            )
                            logger.info(f"[DEBUG] Push notification result for {parent.email}: {push_result}")
                        else:
                            logger.warning(f"[DEBUG] No FCM token found for parent {parent.email}")
                    except Exception as push_error:
                        logger.error(f"[DEBUG] Push error for {parent.email}: {push_error}")

            # 🔹 Launch async thread for notifications
            threading.Thread(target=send_notifications, daemon=True).start()

        # Return response immediately
        return JsonResponse({
            "success": True,
            "message": f"{schedule.service_type} successfully rescheduled to {new_date}",
            "new_date": new_date,
            "reschedule_reason": reschedule_reason
        })

    except Exception as e:
        logger.error(f"[ERROR] Failed to reschedule nutrition service: {e}")
        return JsonResponse({
            "success": False,
            "message": f"Error rescheduling service: {str(e)}"
        })
@login_required
def add_nutrition_service(request, preschooler_id):
    """Add completed nutrition service with notifications"""
    logger.info(f"[DEBUG] Entered add_nutrition_service view for preschooler {preschooler_id}")

    try:
        from .models import Preschooler, NutritionHistory, Account
        preschooler = get_object_or_404(Preschooler, pk=preschooler_id)
        logger.info(f"[DEBUG] Found preschooler: {preschooler.first_name} {preschooler.last_name}")
    except Exception as e:
        logger.error(f"[DEBUG] Error getting preschooler: {e}")
        messages.error(request, "Preschooler not found")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    if request.method != "POST":
        logger.warning("[DEBUG] Request method is not POST")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    # Extract POST values
    service_type = request.POST.get("service_type")
    completion_date = request.POST.get("completion_date")
    notes = request.POST.get("notes", "")

    # Validate required fields
    if not service_type or not completion_date:
        logger.warning("[DEBUG] Missing required fields")
        messages.error(request, "Please fill in all required fields.")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    try:
        # Count existing completed services
        existing_count = NutritionHistory.objects.filter(
            preschooler=preschooler,
            service_type=service_type,
            status='completed'
        ).count()

        dose_number = existing_count + 1

        # Save nutrition history
        nutrition_history = NutritionHistory.objects.create(
            preschooler=preschooler,
            service_type=service_type,
            completion_date=completion_date,
            notes=notes,
            status='completed',
            dose_number=dose_number
        )
        logger.info(f"[DEBUG] NutritionHistory saved: {nutrition_history.id}")
        
        messages.success(
            request,
            f"Nutrition service {service_type} (Dose {dose_number}) added successfully!"
        )

        # === SEND COMPLETION NOTIFICATIONS ===
        parents = preschooler.parents.all()
        logger.info(f"[DEBUG] Found {parents.count()} parent(s) for completion notification")

        for parent in parents:
            # Prepare optional notes text
            notes_line = f"Notes: {notes}\n" if notes else ""

            # Send email
            if parent.email:
                try:
                    subject = f"[PPMS] Nutrition Service Completed for {preschooler.first_name}"
                    message = (
                        f"Dear {parent.full_name},\n\n"
                        f"A nutrition service has been completed for your child, "
                        f"{preschooler.first_name} {preschooler.last_name}.\n\n"
                        f"Service: {service_type}\n"
                        f"Dose: {dose_number}\n"
                        f"Completion Date: {completion_date}\n"
                        f"{notes_line}"
                        "\nThank you,\nPPMS System"
                    )

                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        [parent.email],
                        fail_silently=False
                    )
                    logger.info(f"[DEBUG] Completion email sent to parent {parent.email}")
                except Exception as email_error:
                    logger.error(f"[DEBUG] Completion email sending failed: {email_error}")

            # Send push notification
            try:
                account = Account.objects.filter(email=parent.email).first()
                if account and account.fcm_token:
                    nutrition_icon = "✅🍎" if service_type == "Vitamin A" else "✅💊"
                    notification_title = f"{nutrition_icon} Nutrition Service Completed"
                    notification_body = f"{service_type} completed for {preschooler.first_name}"

                    notification_data = {
                        "type": "nutrition_completed",
                        "preschooler_id": str(preschooler.preschooler_id),
                        "preschooler_name": f"{preschooler.first_name} {preschooler.last_name}",
                        "service_type": service_type,
                        "dose_number": str(dose_number),
                        "completion_date": str(completion_date),
                        "notes": notes or ""
                    }

                    push_result = PushNotificationService.send_push_notification(
                        token=account.fcm_token,
                        title=notification_title,
                        body=notification_body,
                        data=notification_data
                    )

                    if push_result.get("success"):
                        logger.info(f"[DEBUG] Completion push notification sent to {parent.email}")
                    else:
                        logger.error(f"[DEBUG] Completion push notification failed for {parent.email}")
                else:
                    logger.warning(f"[DEBUG] No FCM token found for {parent.email}")
                        
            except Exception as push_error:
                logger.error(f"[DEBUG] Error sending completion push notification: {push_error}")

    except Exception as e:
        logger.error(f"[ERROR] Failed to save nutrition history: {e}")
        messages.error(request, f"Error: {str(e)}")

    return redirect(request.META.get("HTTP_REFERER", "/"))

@require_POST #may binago ako dito
def confirm_schedule(request, schedule_id):
    if not request.user.is_authenticated:
        return redirect('login')

    schedule = get_object_or_404(VaccinationSchedule, id=schedule_id)

    if schedule.preschooler.parent_id.email != request.user.email:
        messages.error(request, "Unauthorized confirmation attempt.")
        return redirect('parent_dashboard')

    # ✅ Mark current schedule as confirmed
    schedule.confirmed_by_parent = True
    schedule.save()

    # 🚫 Prevent over-scheduling
    if (
        schedule.next_vaccine_schedule and 
        schedule.doses < schedule.required_doses
    ):
        # Double check that no future duplicate already exists
        existing = VaccinationSchedule.objects.filter(
            preschooler=schedule.preschooler,
            vaccine_name=schedule.vaccine_name,
            doses=schedule.doses + 1
        ).exists()

        if not existing:
            next_schedule = VaccinationSchedule.objects.create(
                preschooler=schedule.preschooler,
                vaccine_name=schedule.vaccine_name,
                doses=schedule.doses + 1,
                required_doses=schedule.required_doses,
                scheduled_date=schedule.next_vaccine_schedule,
                scheduled_by=schedule.scheduled_by,
                confirmed_by_parent=False
            )
            print("[DEBUG] ➕ Created next dose schedule:", next_schedule)
            messages.success(request, f"Dose {schedule.doses} confirmed. ✅ Next dose scheduled.")
        else:
            print("[DEBUG] ⛔ Skipped creating duplicate schedule")
    else:
        print("[DEBUG] ✅ Final dose confirmed. No more schedules.")
        messages.success(request, "Vaccination confirmed. ✅")

    return redirect('parent_dashboard')

def confirm_vaccine_schedule(request, schedule_id):
    if request.method == 'POST':
        schedule = get_object_or_404(VaccinationSchedule, id=schedule_id)
        schedule.confirmed_by_parent = True
        schedule.save()
        return JsonResponse({'status': 'success', 'message': 'Schedule confirmed!'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

def get_vaccine_status(preschooler, vaccine_name, total_doses):
    """Calculate vaccine status and completed doses"""
    # Count actual completed doses (each completed schedule = 1 dose)
    completed_doses = VaccinationSchedule.objects.filter(
        preschooler=preschooler,
        vaccine_name=vaccine_name,
        status='completed'
    ).count()
    
    # Get any pending schedule
    pending_schedule = VaccinationSchedule.objects.filter(
        preschooler=preschooler,
        vaccine_name=vaccine_name,
        status__in=['scheduled', 'rescheduled']
    ).first()
    
    # Determine status based on completed doses vs total required
    if completed_doses >= total_doses:
        # All doses completed
        return {
            'status': 'completed',
            'completed_doses': completed_doses,
            'total_doses': total_doses,
            'immunization_date': 'Completed',
            'action': 'Fully Vaccinated'
        }
    elif pending_schedule:
        # Has a pending schedule
        return {
            'status': pending_schedule.status,
            'completed_doses': completed_doses,
            'total_doses': total_doses,
            'immunization_date': pending_schedule.scheduled_date.strftime("%m/%d/%Y"),
            'action': 'scheduled',
            'schedule_id': pending_schedule.id
        }
    else:
        # No completed doses and no pending schedule
        return {
            'status': 'not_scheduled',
            'completed_doses': completed_doses,
            'total_doses': total_doses,
            'immunization_date': 'N/A',
            'action': 'needs_schedule'
        }

def get_vaccine_status_with_dose_tracking(preschooler, vaccine_name, total_doses):
    """
    Calculate vaccine status with proper dose tracking
    Assumes you have a 'current_dose' field in VaccinationSchedule model
    """
    # Get all schedules for this vaccine
    schedules = VaccinationSchedule.objects.filter(
        preschooler=preschooler,
        vaccine_name=vaccine_name
    ).order_by('scheduled_date')
    
    completed_doses = 0
    pending_schedule = None
    
    for schedule in schedules:
        if schedule.status == 'completed':
            # If you have a current_dose field, use it; otherwise count as 1 dose
            completed_doses += getattr(schedule, 'current_dose', 1)
        elif schedule.status in ['scheduled', 'rescheduled'] and not pending_schedule:
            pending_schedule = schedule
    
    if completed_doses >= total_doses:
        return {
            'status': 'completed',
            'completed_doses': min(completed_doses, total_doses),  # Cap at total_doses
            'total_doses': total_doses,
            'immunization_date': 'Completed',
            'action': 'Fully Vaccinated'
        }
    elif pending_schedule:
        return {
            'status': pending_schedule.status,
            'completed_doses': completed_doses,
            'total_doses': total_doses,
            'immunization_date': pending_schedule.scheduled_date.strftime("%m/%d/%Y"),
            'action': 'scheduled',
            'schedule_id': pending_schedule.id
        }
    else:
        return {
            'status': 'not_scheduled',
            'completed_doses': completed_doses,
            'total_doses': total_doses,
            'immunization_date': 'N/A',
            'action': 'needs_schedule'
        }

@login_required
def parents_mypreschooler(request, preschooler_id):
    if not request.user.is_authenticated:
        return redirect('login')

    preschooler = get_object_or_404(
        Preschooler.objects.prefetch_related(
            Prefetch('bmi_set', queryset=BMI.objects.order_by('-date_recorded'), to_attr='bmi_records')
        ),
        pk=preschooler_id
    )

    # --- Calculate age ---
    today = date.today()
    birth_date = preschooler.birth_date
    age_years = today.year - birth_date.year
    age_months = today.month - birth_date.month
    age_days = today.day - birth_date.day

    if age_days < 0:
        age_months -= 1
        if today.month == 1:
            last_month, last_year = 12, today.year - 1
        else:
            last_month, last_year = today.month - 1, today.year
        age_days += monthrange(last_year, last_month)[1]

    if age_months < 0:
        age_years -= 1
        age_months += 12

    total_age_months = age_years * 12 + age_months

    # --- Latest BMI ---
    latest_bmi = preschooler.bmi_records[0] if preschooler.bmi_records else None
    bmi_value = None

    # --- Nutrition & BMI classifications ---
    weight_for_age_status = "N/A"
    height_for_age_status = "N/A"
    weight_for_height_status = "N/A"
    nutritional_status = "N/A"

    if latest_bmi:
        sex = preschooler.sex.lower()
        try:
            # Compute BMI value (if not stored directly in DB)
            bmi_value = latest_bmi.bmi_value if hasattr(latest_bmi, 'bmi_value') else calculate_bmi(latest_bmi.weight, latest_bmi.height)

            if sex in ['female', 'girl', 'f']:
                weight_for_age_status = classify_weight_for_age(total_age_months, latest_bmi.weight, WEIGHT_REF_GIRLS)
                height_for_age_status = classify_height_for_age(total_age_months, latest_bmi.height, HEIGHT_REF_GIRLS)
                weight_for_height_status = classify_weight_for_height(latest_bmi.height, latest_bmi.weight, WFH_GIRLS)
            else:
                weight_for_age_status = classify_weight_for_age(total_age_months, latest_bmi.weight, WEIGHT_REF_BOYS)
                height_for_age_status = classify_height_for_age(total_age_months, latest_bmi.height, HEIGHT_REF_BOYS)
                weight_for_height_status = classify_weight_for_height(latest_bmi.height, latest_bmi.weight, WFH_BOYS)

            # Use z-score for nutritional status
            z = bmi_zscore(preschooler.sex, total_age_months, bmi_value)
            nutritional_status = classify_bmi_for_age(z)

        except Exception as e:
            print(f"⚠️ Error during BMI classification for preschooler {preschooler.id}: {e}")
            nutritional_status = preschooler.nutritional_status or "N/A"

    # --- Immunization history ---
    immunization_records = preschooler.vaccination_schedules.filter(
        confirmed_by_parent=True,
        status='completed'
    ).order_by('vaccine_name', 'scheduled_date')

    immunization_history = []
    vaccine_dose_counter = defaultdict(int)

    for record in immunization_records:
        vaccine_dose_counter[record.vaccine_name] += 1
        immunization_history.append({
            'vaccine_name': record.vaccine_name,
            'doses': f"{vaccine_dose_counter[record.vaccine_name]}/{record.required_doses}",
            'given_date': record.scheduled_date,
        })

    # --- Nutrition services ---
    nutrition_services = preschooler.nutrition_services.all().order_by('-completion_date')

    # --- Parent account ---
    account = get_object_or_404(Account, email=request.user.email)

    return render(request, 'HTML/parents_mypreschooler.html', {
        'preschooler': preschooler,
        'account': account,
        'age_years': age_years,
        'age_months': age_months,
        'age_days': age_days,
        'latest_bmi': latest_bmi,
        'bmi_value': bmi_value,
        'weight_for_age_status': weight_for_age_status,
        'height_for_age_status': height_for_age_status,
        'weight_for_height_status': weight_for_height_status,
        'nutritional_status': nutritional_status,
        'immunization_history': immunization_history,
        'nutrition_services': nutrition_services,
    })



@login_required
def add_vaccine(request, preschooler_id):
    if request.method == 'POST':
        preschooler = get_object_or_404(Preschooler, preschooler_id=preschooler_id)

        vaccine_name = request.POST.get('vaccine_name')
        doses = request.POST.get('required_doses')  # This will always be 1
        immunization_date = request.POST.get('immunization_date')

        try:
            # Parse the immunization date
            immunization_date_obj = datetime.strptime(immunization_date, '%Y-%m-%d').date()
            
            # Create completion_date from immunization_date (set time to noon)
            completion_datetime = datetime.combine(immunization_date_obj, datetime.min.time().replace(hour=12))
            completion_date = timezone.make_aware(completion_datetime, timezone.get_current_timezone())

            doses_int = int(doses)  # This will be 1

            # Create the vaccination record for 1 dose
            VaccinationSchedule.objects.create(
                preschooler=preschooler,
                vaccine_name=vaccine_name,
                required_doses=doses_int,  # Always 1
                scheduled_date=immunization_date_obj,
                status='completed',
                completion_date=completion_date,
                reschedule_reason=None
            )

            messages.success(request, f'1 dose of {vaccine_name} added to immunization history successfully!')

        except ValueError:
            messages.error(request, 'Invalid date format provided.')
        except Exception as e:
            messages.error(request, f'Error adding vaccine: {str(e)}')

    return redirect('preschooler_detail', preschooler_id=preschooler_id)


from datetime import date
from calendar import monthrange
from django.shortcuts import render, get_object_or_404
from .models import Preschooler

# -------------------------------
# WHO Growth Standards Reference
# -------------------------------

WEIGHT_REF_GIRLS = {
    0: {"-3SD": 2.0, "-2SD": 2.4, "median": 3.2, "+2SD": 4.2, "+3SD": 4.8},
    6: {"-3SD": 5.1, "-2SD": 5.7, "median": 7.3, "+2SD": 9.0, "+3SD": 9.8},
    12: {"-3SD": 6.0, "-2SD": 7.0, "median": 8.9, "+2SD": 10.8, "+3SD": 12.0},
    24: {"-3SD": 7.5, "-2SD": 8.7, "median": 11.5, "+2SD": 14.2, "+3SD": 15.8},
    36: {"-3SD": 8.8, "-2SD": 10.2, "median": 13.9, "+2SD": 17.0, "+3SD": 19.0},
    48: {"-3SD": 10.2, "-2SD": 11.8, "median": 16.0, "+2SD": 19.8, "+3SD": 22.2},
    60: {"-3SD": 11.5, "-2SD": 13.3, "median": 18.2, "+2SD": 22.7, "+3SD": 25.5},
}

HEIGHT_REF_GIRLS = {
    0: {"-3SD": 43.6, "-2SD": 45.4, "median": 49.1, "+2SD": 52.9, "+3SD": 54.7},
    6: {"-3SD": 58.1, "-2SD": 61.0, "median": 65.7, "+2SD": 70.5, "+3SD": 73.5},
    12: {"-3SD": 65.9, "-2SD": 68.9, "median": 74.0, "+2SD": 79.2, "+3SD": 82.3},
    24: {"-3SD": 76.0, "-2SD": 80.0, "median": 86.4, "+2SD": 92.9, "+3SD": 96.9},
    36: {"-3SD": 83.6, "-2SD": 88.0, "median": 95.1, "+2SD": 102.2, "+3SD": 106.6},
    48: {"-3SD": 90.7, "-2SD": 95.0, "median": 102.7, "+2SD": 110.6, "+3SD": 115.0},
    60: {"-3SD": 96.1, "-2SD": 100.7, "median": 109.4, "+2SD": 118.2, "+3SD": 122.8},
}

WEIGHT_REF_BOYS = {
    0: {"-3SD": 2.1, "-2SD": 2.5, "median": 3.3, "+2SD": 4.4, "+3SD": 5.0},
    6: {"-3SD": 5.4, "-2SD": 6.0, "median": 7.9, "+2SD": 9.7, "+3SD": 10.9},
    12: {"-3SD": 6.4, "-2SD": 7.5, "median": 9.6, "+2SD": 11.8, "+3SD": 13.3},
    24: {"-3SD": 8.6, "-2SD": 9.7, "median": 12.2, "+2SD": 15.3, "+3SD": 17.1},
    36: {"-3SD": 10.0, "-2SD": 11.3, "median": 14.3, "+2SD": 17.9, "+3SD": 20.0},
    48: {"-3SD": 11.5, "-2SD": 12.9, "median": 16.3, "+2SD": 20.6, "+3SD": 23.1},
    60: {"-3SD": 12.9, "-2SD": 14.5, "median": 18.7, "+2SD": 23.9, "+3SD": 26.8},
}

HEIGHT_REF_BOYS = {
    0: {"-3SD": 44.2, "-2SD": 46.1, "median": 49.9, "+2SD": 53.7, "+3SD": 55.6},
    6: {"-3SD": 61.0, "-2SD": 63.0, "median": 67.6, "+2SD": 72.2, "+3SD": 74.2},
    12: {"-3SD": 68.6, "-2SD": 71.0, "median": 75.7, "+2SD": 80.5, "+3SD": 82.9},
    24: {"-3SD": 78.0, "-2SD": 81.0, "median": 87.1, "+2SD": 93.3, "+3SD": 96.3},
    36: {"-3SD": 85.0, "-2SD": 89.0, "median": 96.0, "+2SD": 103.1, "+3SD": 106.9},
    48: {"-3SD": 91.9, "-2SD": 96.0, "median": 103.9, "+2SD": 111.9, "+3SD": 116.1},
    60: {"-3SD": 97.2, "-2SD": 101.2, "median": 111.2, "+2SD": 121.4, "+3SD": 125.9},
}

WFH_GIRLS = {
    45: {"-3SD": 1.8, "-2SD": 2.0, "median": 2.7, "+2SD": 3.7, "+3SD": 4.3},
    50: {"-3SD": 2.4, "-2SD": 2.6, "median": 3.5, "+2SD": 4.7, "+3SD": 5.4},
    55: {"-3SD": 3.0, "-2SD": 3.4, "median": 4.5, "+2SD": 6.0, "+3SD": 6.8},
    60: {"-3SD": 3.8, "-2SD": 4.3, "median": 5.7, "+2SD": 7.6, "+3SD": 8.6},
    65: {"-3SD": 5.0, "-2SD": 5.8, "median": 7.5, "+2SD": 9.5, "+3SD": 10.9},
    70: {"-3SD": 5.7, "-2SD": 6.5, "median": 8.5, "+2SD": 10.8, "+3SD": 12.4},
    75: {"-3SD": 6.3, "-2SD": 7.2, "median": 9.5, "+2SD": 12.1, "+3SD": 13.9},
    80: {"-3SD": 7.0, "-2SD": 8.0, "median": 10.6, "+2SD": 13.5, "+3SD": 15.6},
    85: {"-3SD": 7.9, "-2SD": 9.0, "median": 11.9, "+2SD": 15.1, "+3SD": 17.5},
    90: {"-3SD": 8.8, "-2SD": 10.1, "median": 13.1, "+2SD": 16.7, "+3SD": 19.4},
    95: {"-3SD": 10.0, "-2SD": 11.4, "median": 14.6, "+2SD": 18.6, "+3SD": 21.5},
    100: {"-3SD": 11.2, "-2SD": 12.8, "median": 16.0, "+2SD": 20.6, "+3SD": 23.9},
    105: {"-3SD": 12.7, "-2SD": 14.4, "median": 17.7, "+2SD": 22.8, "+3SD": 26.3},
    110: {"-3SD": 14.3, "-2SD": 16.4, "median": 19.2, "+2SD": 24.5, "+3SD": 28.2},
}


WFH_BOYS = {
    45: {"-3SD": 1.9, "-2SD": 2.1, "median": 2.8, "+2SD": 3.8, "+3SD": 4.4},
    50: {"-3SD": 2.5, "-2SD": 2.8, "median": 3.7, "+2SD": 5.0, "+3SD": 5.7},
    55: {"-3SD": 3.2, "-2SD": 3.6, "median": 4.8, "+2SD": 6.3, "+3SD": 7.1},
    60: {"-3SD": 4.1, "-2SD": 4.6, "median": 6.0, "+2SD": 7.9, "+3SD": 8.8},
    65: {"-3SD": 5.1, "-2SD": 5.9, "median": 7.6, "+2SD": 9.7, "+3SD": 11.1},
    70: {"-3SD": 5.9, "-2SD": 6.7, "median": 8.7, "+2SD": 11.0, "+3SD": 12.6},
    75: {"-3SD": 6.5, "-2SD": 7.5, "median": 9.8, "+2SD": 12.4, "+3SD": 14.2},
    80: {"-3SD": 7.3, "-2SD": 8.3, "median": 11.0, "+2SD": 13.9, "+3SD": 16.0},
    85: {"-3SD": 8.2, "-2SD": 9.3, "median": 12.3, "+2SD": 15.5, "+3SD": 17.8},
    90: {"-3SD": 9.1, "-2SD": 10.5, "median": 13.6, "+2SD": 17.2, "+3SD": 20.0},
    95: {"-3SD": 10.3, "-2SD": 11.8, "median": 15.1, "+2SD": 19.1, "+3SD": 22.1},
    100: {"-3SD": 11.6, "-2SD": 13.2, "median": 16.5, "+2SD": 21.0, "+3SD": 24.4},
    105: {"-3SD": 13.2, "-2SD": 14.9, "median": 18.2, "+2SD": 23.2, "+3SD": 27.0},
    110: {"-3SD": 14.8, "-2SD": 16.8, "median": 20.0, "+2SD": 25.6, "+3SD": 29.5},
}

# -------------------------------
# Helper Functions
# -------------------------------

def interpolate_value(x, ref_table):
    """Interpolate between reference points."""
    pts = sorted(ref_table.keys())
    if x <= pts[0]:
        return ref_table[pts[0]]
    if x >= pts[-1]:
        return ref_table[pts[-1]]
    for i in range(len(pts) - 1):
        p1, p2 = pts[i], pts[i + 1]
        if p1 <= x <= p2:
            r1, r2 = ref_table[p1], ref_table[p2]
            return {k: r1[k] + (r2[k] - r1[k]) * ((x - p1) / (p2 - p1)) for k in r1}

def classify_weight_for_age(age_months, weight, table):
    r = interpolate_value(age_months, table)
    if weight < r["-3SD"]:
        return "Severely Underweight"
    elif weight < r["-2SD"]:
        return "Underweight"
    elif weight > r["+2SD"]:
        return "Overweight"
    else:
        return "Normal"

def classify_height_for_age(age_months, height, table):
    r = interpolate_value(age_months, table)
    if height < r["-3SD"]:
        return "Severely stunted"
    elif height < r["-2SD"]:
        return "Stunted"
    elif height > r["+2SD"]:
        return "Tall"
    else:
        return "Normal"

def classify_weight_for_height(height, weight, table):
    r = interpolate_value(height, table)
    if weight < r["-3SD"]:
        return "Severely Wasted"
    elif weight < r["-2SD"]:
        return "Wasted"
    elif weight > r["+3SD"]:
        return "Obese"
    elif weight > r["+2SD"]:
        return "Overweight"
    else:
        return "Normal"
    
def archived_preschoolers(request):
    """Display list of archived preschoolers with search and pagination."""
    
    if not request.user.is_authenticated:
        return redirect('login')
    
    search_query = request.GET.get('q', '').strip()

    # Get all archived preschoolers
    archived = Preschooler.objects.filter(is_archived=True)

    # Filter by name if search query provided
    if search_query:
        archived = archived.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )

    # Order by most recently registered first
    archived = archived.order_by('-date_registered')

    # Pagination (10 rows per page)
    paginator = Paginator(archived, 10)
    page_number = request.GET.get('page')
    archived_page = paginator.get_page(page_number)

    return render(request, 'HTML/archived.html', {
        'archived_preschoolers': archived_page,
        'search_query': search_query
    })


def preschooler_detail(request, preschooler_id):
    """Detailed preschooler profile with BMI, temperature, vaccines, and nutrition."""

    # === Auto-archive preschoolers aged 60+ months ===
    auto_archived_count = auto_archive_aged_preschoolers()
    if auto_archived_count > 0:
        print(f"AUTO-ARCHIVED: {auto_archived_count} preschoolers during detail view")

    # === Get preschooler (non-archived only) ===
    preschooler = get_object_or_404(Preschooler, preschooler_id=preschooler_id)

    # === Check if archived ===
    is_archived = preschooler.is_archived

    # Auto-archive if age >= 60 months
    if preschooler.age_in_months and preschooler.age_in_months >= 60:
        preschooler.is_archived = True
        preschooler.save()
        is_archived = True
        messages.warning(
            request,
            f"{preschooler.first_name} {preschooler.last_name} has been automatically archived as they have exceeded 60 months."
        )
        if request.GET.get('from') == 'archived':
            return redirect('archived_preschoolers')
        return redirect('preschoolers')

    # === Calculate Age ===
    today = timezone.now().date()
    birth_date = preschooler.birth_date
    age_years = today.year - birth_date.year
    age_months = today.month - birth_date.month
    age_days = today.day - birth_date.day

    if age_days < 0:
        age_months -= 1
        if today.month == 1:
            last_month, last_year = 12, today.year - 1
        else:
            last_month, last_year = today.month - 1, today.year
        age_days += monthrange(last_year, last_month)[1]

    if age_months < 0:
        age_years -= 1
        age_months += 12

    total_age_months = age_years * 12 + age_months

    # === Compute Latest BMI-related classifications (for summary) ===
    bmi = preschooler.bmi_set.order_by('-date_recorded').first()

    weight_for_age_status = "N/A"
    height_for_age_status = "N/A"
    weight_for_height_status = "N/A"
    nutritional_status = "N/A"

    if bmi:
        sex = preschooler.sex.lower()
        try:
            if sex in ['female', 'girl', 'f']:
                weight_for_age_status = classify_weight_for_age(total_age_months, bmi.weight, WEIGHT_REF_GIRLS)
                height_for_age_status = classify_height_for_age(total_age_months, bmi.height, HEIGHT_REF_GIRLS)
                weight_for_height_status = classify_weight_for_height(bmi.height, bmi.weight, WFH_GIRLS)
            else:
                weight_for_age_status = classify_weight_for_age(total_age_months, bmi.weight, WEIGHT_REF_BOYS)
                height_for_age_status = classify_height_for_age(total_age_months, bmi.height, HEIGHT_REF_BOYS)
                weight_for_height_status = classify_weight_for_height(bmi.height, bmi.weight, WFH_BOYS)

            z = bmi_zscore(preschooler.sex, total_age_months, bmi.bmi_value)
            nutritional_status = classify_bmi_for_age(z)
        except Exception as e:
            print(f"Error during BMI classification: {e}")
            nutritional_status = "Error"

    # === Vaccine Scheduling ===
    standard_vaccines = [
        {'name': 'BCG Vaccine', 'total_doses': 1},
        {'name': 'Hepatitis B Vaccine', 'total_doses': 1},
        {'name': 'Pentavalent Vaccine', 'total_doses': 3},
        {'name': 'Oral Polio Vaccine', 'total_doses': 3},
        {'name': 'Inactivated Polio Vaccine', 'total_doses': 2},
        {'name': 'Pneumococcal Conjugate Vaccine', 'total_doses': 3},
        {'name': 'Measles, Mumps, and Rubella', 'total_doses': 2},
    ]

    vaccine_statuses = []
    total_vaccines = len(standard_vaccines)
    completed_vaccines = 0
    
    for vaccine in standard_vaccines:
        status = get_enhanced_vaccine_status(preschooler, vaccine['name'], vaccine['total_doses'])
        eligibility = get_vaccine_eligibility(preschooler, vaccine['name'])
        status['eligibility_info'] = eligibility
        
        # Check if this vaccine is fully completed
        if status.get('completed_doses', 0) >= vaccine['total_doses']:
            completed_vaccines += 1
        
        vaccine_statuses.append({
            'name': vaccine['name'],
            'total_doses': vaccine['total_doses'],
            **status
        })

    # Calculate vaccination completion status
    if completed_vaccines == total_vaccines:
        vaccination_completion_status = "Complete"
    else:
        vaccination_completion_status = "Incomplete"

    # === Nutrition Services ===
    standard_nutrition_services = [
        {'name': 'Vitamin A', 'total_doses': 10},
        {'name': 'Deworming', 'total_doses': 10},
    ]

    nutrition_statuses = []
    for service in standard_nutrition_services:
        status = get_enhanced_nutrition_status(preschooler, service['name'], service['total_doses'])
        nutrition_statuses.append({
            'name': service['name'],
            'total_doses': service['total_doses'],
            **status
        })

    # === Vaccination History ===
    from .models import VaccinationSchedule
    immunization_history = VaccinationSchedule.objects.filter(
        preschooler=preschooler,
        status='completed'
    ).order_by('vaccine_name', 'completion_date')

    pending_schedules = VaccinationSchedule.objects.filter(
        preschooler=preschooler,
        status__in=['scheduled', 'rescheduled', 'pending']
    ).exclude(status='completed').order_by('scheduled_date')

    # === Pair Each BMI with Temperature ===
    bmi_records = preschooler.bmi_set.all().order_by('-date_recorded')
    bmi_with_temps = []

    for record in bmi_records:
        # Safely handle date vs datetime
        if hasattr(record.date_recorded, "date"):
            date_value = record.date_recorded.date()
        else:
            date_value = record.date_recorded

        # Get the nearest temperature recorded within the same day
        temperature = Temperature.objects.filter(
            preschooler_id=preschooler,
            date_recorded=date_value
        ).order_by('-date_recorded').first()

        sex = preschooler.sex.lower()

        try:
            if sex in ['female', 'girl', 'f']:
                wfa = classify_weight_for_age(total_age_months, record.weight, WEIGHT_REF_GIRLS)
                hfa = classify_height_for_age(total_age_months, record.height, HEIGHT_REF_GIRLS)
                wfh = classify_weight_for_height(record.height, record.weight, WFH_GIRLS)
            else:
                wfa = classify_weight_for_age(total_age_months, record.weight, WEIGHT_REF_BOYS)
                hfa = classify_height_for_age(total_age_months, record.height, HEIGHT_REF_BOYS)
                wfh = classify_weight_for_height(record.height, record.weight, WFH_BOYS)

            z = bmi_zscore(preschooler.sex, total_age_months, record.bmi_value)
            nutrition_status = classify_bmi_for_age(z)
        except Exception as e:
            print(f"Error computing classification for record {record.bmi_id}: {e}")
            wfa = hfa = wfh = nutrition_status = "Error"

        bmi_with_temps.append({
            'date_recorded': record.date_recorded,
            'height': record.height,
            'weight': record.weight,
            'bmi_value': record.bmi_value,
            'nutritional_status': nutrition_status,
            'temperature': temperature.temperature_value if temperature else None,
            'weight_for_age_status': wfa,
            'height_for_age_status': hfa,
            'weight_for_height_status': wfh,
        })
        
    # === Nutrition Records ===
    try:
        nutrition_services = preschooler.nutrition_services.all().order_by('-completion_date')
    except:
        from .models import NutritionHistory
        nutrition_services = NutritionHistory.objects.filter(
            preschooler=preschooler
        ).order_by('-completion_date')

    # === Context ===
    context = {
        'preschooler': preschooler,
        'is_archived': is_archived,  
        'bmi': bmi,
        'bmi_with_temps': bmi_with_temps,
        'immunization_history': immunization_history,
        'pending_schedules': pending_schedules,
        'nutrition_services': nutrition_services,
        'nutrition_statuses': nutrition_statuses,
        'vaccine_statuses': vaccine_statuses,
        'vaccination_completion_status': vaccination_completion_status,
        'age_years': age_years,
        'age_months': age_months,
        'age_days': age_days,
        'total_age_months': total_age_months,
        'nutritional_status': nutritional_status,
        'weight_for_age_status': weight_for_age_status,
        'height_for_age_status': height_for_age_status,
        'weight_for_height_status': weight_for_height_status,
    }

    return render(request, 'HTML/preschooler_data.html', context)
    
def get_nutrition_eligibility(preschooler, service_type):
    """
    Enhanced nutrition eligibility that properly handles completed services
    and calculates next availability based on 6-month intervals.
    """ 
    today = timezone.now().date()
    birth_date = preschooler.birth_date

    # --- Age calculation in months ---
    age_years = today.year - birth_date.year
    age_months = today.month - birth_date.month
    age_days = today.day - birth_date.day

    if age_days < 0:
        age_months -= 1
        if today.month == 1:
            last_month, last_year = 12, today.year - 1
        else:
            last_month, last_year = today.month - 1, today.year
        age_days += monthrange(last_year, last_month)[1]

    if age_months < 0:
        age_years -= 1
        age_months += 12

    total_age_months = age_years * 12 + age_months

    # --- Nutrition service rules ---
    nutrition_schedule = {
        'Vitamin A': {
            'min_age_months': 6,
            'interval_months': 6,
            'max_age_months': 59,
            'eligible_ages': [6, 12, 18, 24, 30, 36, 42, 48, 54],
        },
        'Deworming': {
            'min_age_months': 6,
            'interval_months': 6,
            'max_age_months': 59,
            'eligible_ages': [6, 12, 18, 24, 30, 36, 42, 48, 54],
        },
    }

    if service_type not in nutrition_schedule:
        return {
            'can_schedule': False,
            'reason': 'Unknown service type',
            'next_eligible_age': None,
            'description': 'Service not recognized',
        }

    service_info = nutrition_schedule[service_type]

    # --- Too young ---
    if total_age_months < service_info['min_age_months']:
        return {
            'can_schedule': False,
            'reason': f'Too young. {service_type} starts at {service_info["min_age_months"]} months.',
            'next_eligible_age': service_info['min_age_months'],
            'description': f'Available at {service_info["min_age_months"]} months',
        }

    # --- Too old ---
    if total_age_months > service_info['max_age_months']:
        return {
            'can_schedule': False,
            'reason': f'Child has exceeded age limit for {service_type}.',
            'next_eligible_age': None,
            'description': 'No longer age-appropriate',
        }

    # --- Completed services ---
    try:
        completed_services = preschooler.nutrition_services.filter(
            service_type=service_type, status='completed'
        ).order_by('completion_date')
    except AttributeError:
        from .models import NutritionService
        completed_services = NutritionService.objects.filter(
            preschooler=preschooler, service_type=service_type, status='completed'
        ).order_by('service_date')

    # --- Pending schedules ---
    try:
        pending_schedules = preschooler.nutrition_schedules.filter(
            service_type=service_type, status__in=['scheduled', 'rescheduled']
        ).exists()
    except AttributeError:
        from .models import NutritionService
        pending_schedules = NutritionService.objects.filter(
            preschooler=preschooler, service_type=service_type,
            status__in=['scheduled', 'rescheduled']
        ).exists()

    if pending_schedules:
        return {
            'can_schedule': False,
            'reason': f'{service_type} already scheduled.',
            'next_eligible_age': None,
            'description': 'Service already scheduled',
        }

    # --- If previously completed ---
    if completed_services.exists():
        last_service = completed_services.last()
        last_service_date = getattr(last_service, 'completion_date', None) or getattr(last_service, 'service_date', None)

        # ✅ FIX: Ensure both are date objects
        if hasattr(last_service_date, 'date'):
            last_service_date = last_service_date.date()

        months_since_last = (today - last_service_date).days // 30
        print(f"DEBUG: {service_type} - Last service: {last_service_date}, Months since: {months_since_last}")

        if months_since_last < 6:
            months_to_wait = 6 - months_since_last
            next_eligible_date = last_service_date + timedelta(days=180)
            next_eligible_age_months = (
                (next_eligible_date.year - birth_date.year) * 12 +
                (next_eligible_date.month - birth_date.month)
            )

            return {
                'can_schedule': False,
                'reason': f'Too soon since last {service_type}. Wait {months_to_wait} more months.',
                'next_eligible_age': next_eligible_age_months,
                'description': f'Next dose available in {months_to_wait} months',
            }

        return {
            'can_schedule': True,
            'reason': f'Child is eligible for {service_type} (6+ months since last dose)',
            'next_eligible_age': None,
            'description': f'Ready for next dose (every 6 months)',
            'current_age_months': total_age_months,
            'last_service_months_ago': months_since_last,
        }

    # --- If no completed services yet ---
    eligible_ages = service_info['eligible_ages']
    next_eligible_age = None

    for age in eligible_ages:
        if total_age_months >= age:
            # Eligible for first or current dose
            for future_age in eligible_ages:
                if future_age > total_age_months:
                    next_eligible_age = future_age
                    break
            return {
                'can_schedule': True,
                'reason': f'Child is eligible for first {service_type} dose',
                'next_eligible_age': next_eligible_age,
                'description': f'First dose (starts at 6 months)',
                'current_age_months': total_age_months,
            }

    # --- Not yet eligible ---
    for age in eligible_ages:
        if age > total_age_months:
            next_eligible_age = age
            break

    return {
        'can_schedule': False,
        'reason': f'Not at eligible age for {service_type}.',
        'next_eligible_age': next_eligible_age,
        'description': f'First dose available at {next_eligible_age} months',
    }

def get_enhanced_nutrition_status(preschooler, service_type, total_doses):
    """
    Enhanced nutrition status that includes age-based eligibility - standalone version
    """
    # Get completed services count
    try:
        completed_services = preschooler.nutrition_services.filter(
            service_type=service_type,
            status='completed'
        )
        completed_count = completed_services.count()
        latest_service = completed_services.order_by('-completion_date').first() if completed_services.exists() else None
    except AttributeError:
        try:
            from .models import NutritionService
            completed_services = NutritionService.objects.filter(
                preschooler=preschooler,
                service_type=service_type,
                status='completed'
            )
            completed_count = completed_services.count()
            latest_service = completed_services.order_by('-service_date').first() if completed_services.exists() else None
        except:
            completed_count = 0
            latest_service = None
    
    # Get pending schedules
    try:
        pending_schedule = preschooler.nutrition_schedules.filter(
            service_type=service_type,
            status__in=['scheduled', 'rescheduled']
        ).first()
    except AttributeError:
        try:
            from .models import NutritionService
            pending_schedule = NutritionService.objects.filter(
                preschooler=preschooler,
                service_type=service_type,
                status__in=['scheduled', 'rescheduled']
            ).first()
        except:
            pending_schedule = None
    
    # Determine current status
    if pending_schedule:
        current_status = pending_schedule.status
        service_date = getattr(pending_schedule, 'service_date', None)
        schedule_id = pending_schedule.id
    elif completed_count > 0:
        current_status = 'completed'
        service_date = getattr(latest_service, 'completion_date', None) or getattr(latest_service, 'service_date', None)
        schedule_id = None
    else:
        current_status = 'pending'
        service_date = None
        schedule_id = None
    
    # Get eligibility information
    eligibility = get_nutrition_eligibility(preschooler, service_type)
    
    # Enhanced status logic
    enhanced_status = {
        'status': current_status,
        'completed_doses': completed_count,
        'total_doses': total_doses,
        'service_date': service_date,
        'schedule_id': schedule_id,
        'eligibility': eligibility,
        'can_schedule': eligibility['can_schedule'],
        'eligibility_reason': eligibility['reason'],
        'next_eligible_age': eligibility.get('next_eligible_age'),
        'age_description': eligibility['description']
    }
    
    # Debug logging
    print("DEBUG: {service_type} Status:")
    print("  - Completed doses: {completed_count}")
    print("  - Current status: {current_status}")
    print("  - Can schedule: {eligibility['can_schedule']}")
    print("  - Eligibility reason: {eligibility['reason']}")
    print("  - Age description: {eligibility['description']}")
    
    return enhanced_status

@login_required
def add_nutrition_service(request, preschooler_id):
    """Add completed nutrition service (Vitamin A or Deworming) for a preschooler"""
    
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('preschooler_detail', preschooler_id=preschooler_id)
    
    try:
        preschooler = get_object_or_404(Preschooler, preschooler_id=preschooler_id)
        
        # Get form data
        service_type = request.POST.get('service_type')
        completion_date = request.POST.get('completion_date')
        notes = request.POST.get('notes', '')
        source = request.POST.get('source', 'add_nutrition_modal')
        
        # Validate required fields
        if not service_type or not completion_date:
            messages.error(request, 'Service type and completion date are required.')
            return redirect('preschooler_detail', preschooler_id=preschooler_id)
        
        # Parse and validate the completion date
        try:
            # Handle both date and datetime formats
            if 'T' in completion_date:  # datetime-local format
                completion_date_obj = datetime.strptime(completion_date, '%Y-%m-%dT%H:%M').date()
            else:  # date format
                completion_date_obj = datetime.strptime(completion_date, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid date format.')
            return redirect('preschooler_detail', preschooler_id=preschooler_id)
        
        # Validate date is not in the future
        if completion_date_obj > timezone.now().date():
            messages.error(request, 'Completion date cannot be in the future.')
            return redirect('preschooler_detail', preschooler_id=preschooler_id)
        
        # Validate date is not before birth
        if completion_date_obj < preschooler.birth_date:
            messages.error(request, 'Completion date cannot be before the child was born.')
            return redirect('preschooler_detail', preschooler_id=preschooler_id)
        
        # Create nutrition service record as completed
        nutrition_service = NutritionService.objects.create(
            preschooler=preschooler,
            service_type=service_type,
            completion_date=completion_date_obj,
            status='completed',
            notes=notes if notes else f'Completed {service_type} service'
        )
        
        # Send notifications to parents
        try:
            parent = preschooler.parent_id
            if parent and parent.user:
                # Create in-app notification
                Notification.objects.create(
                    user=parent.user,
                    title=f"✅ {service_type} Service Completed",
                    message=f"{service_type} service has been completed for {preschooler.first_name} on {completion_date_obj.strftime('%B %d, %Y')}.",
                    notification_type='nutrition_completed',
                    is_read=False
                )
                
                # Send email notification if parent has email
                if parent.user.email:
                    service_emoji = "💊" if service_type == "Vitamin A" else "🪱"
                    send_mail(
                        subject=f"{service_emoji} {service_type} Service Completed - {preschooler.first_name}",
                        message=f"""
Dear Parent,

Good news! {service_type} service has been successfully completed for {preschooler.first_name} {preschooler.last_name}.

Service Details:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- Service Type: {service_type}
- Completion Date: {completion_date_obj.strftime('%B %d, %Y')}
- Child: {preschooler.first_name} {preschooler.last_name}
- Notes: {notes if notes else 'None'}

This service is an important part of your child's health and nutrition program.

Thank you for your cooperation in maintaining your child's health.

Best regards,
Preschooler Portal Management System
                        """,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[parent.user.email],
                        fail_silently=True,
                    )
                
                print(f"✅ Notifications sent to parent for completed {service_type} service")
        except Exception as e:
            print(f"⚠️ Could not send notification: {e}")
        
        messages.success(
            request,
            f'{service_type} service has been recorded as completed for {preschooler.first_name}. Parents have been notified.'
        )
        
    except Exception as e:
        print(f"Error adding nutrition service: {e}")
        messages.error(request, f'An error occurred: {str(e)}')
    
    return redirect('preschooler_detail', preschooler_id=preschooler_id)



add_completed_nutrition_service = add_nutrition_service



def send_completion_notifications_async(parents, preschooler, schedule, completed_count, required_doses):
    """Background task: send email + push notifications for vaccination completion"""
    from .models import Account
    from .services import PushNotificationService

    for parent in parents:
        try:
            # === Email Notification ===
            if parent.email:
                try:
                    if completed_count >= required_doses:
                        subject = f"[PPMS] {schedule.vaccine_name} Vaccination Complete for {preschooler.first_name}"
                        message = (
                            f"Dear {parent.full_name},\n\n"
                            f"Great news! Your child {preschooler.first_name} {preschooler.last_name} "
                            f"has completed all {required_doses} doses of {schedule.vaccine_name}.\n\n"
                            f"Vaccination completed on: {schedule.completion_date.strftime('%B %d, %Y at %I:%M %p')}\n"
                            f"Total doses completed: {completed_count}/{required_doses}\n\n"
                            "Your child is now fully protected against this disease.\n\n"
                            "Thank you for keeping your child's vaccinations up to date!\n\n"
                            "PPMS System"
                        )
                    else:
                        subject = f"[PPMS] {schedule.vaccine_name} Dose Completed for {preschooler.first_name}"
                        message = (
                            f"Dear {parent.full_name},\n\n"
                            f"Your child {preschooler.first_name} {preschooler.last_name} "
                            f"has received dose {completed_count} of {required_doses} for {schedule.vaccine_name}.\n\n"
                            f"Vaccination completed on: {schedule.completion_date.strftime('%B %d, %Y at %I:%M %p')}\n"
                            f"Progress: {completed_count}/{required_doses} doses completed\n"
                            f"Remaining doses: {required_doses - completed_count}\n\n"
                            "Please ensure your child receives the remaining doses.\n\n"
                            "Thank you,\nPPMS System"
                        )

                    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [parent.email], fail_silently=False)
                    logger.info("[ASYNC] Vaccination email sent to {parent.email}")
                except Exception as e:
                    logger.error("[ASYNC] Failed to send vaccination email to {parent.email}: {e}")

            # === Push Notification ===
            account = Account.objects.filter(email=parent.email).first()
            if account and account.fcm_token:
                try:
                    if completed_count >= required_doses:
                        title = f"🎉 {schedule.vaccine_name} Complete!"
                        body = f"{preschooler.first_name} completed all {required_doses} doses"
                    else:
                        title = f"💉 {schedule.vaccine_name} Dose Complete"
                        body = f"Dose {completed_count}/{required_doses} completed for {preschooler.first_name}"

                    data = {
                        "type": "vaccination_completed",
                        "preschooler_id": str(preschooler.preschooler_id),
                        "preschooler_name": "{preschooler.first_name} {preschooler.last_name}",
                        "vaccine_name": schedule.vaccine_name,
                        "completed_doses": str(completed_count),
                        "total_doses": str(required_doses),
                        "completion_date": schedule.completion_date.isoformat(),
                        "schedule_id": str(schedule.id),
                        "fully_completed": str(completed_count >= required_doses).lower(),
                        "needs_next_dose": str(completed_count < required_doses).lower()
                    }

                    PushNotificationService.send_push_notification(
                        token=account.fcm_token,
                        title=title,
                        body=body,
                        data=data
                    )
                    logger.info(f"[ASYNC] Vaccination push sent to {parent.email}")
                except Exception as e:
                    logger.error(f"[ASYNC] Failed to send push to {parent.email}: {e}")
            else:
                logger.warning(f"[ASYNC] No FCM token for {parent.email}")

        except Exception as e:
            logger.error(f"[ASYNC] Error handling parent {parent.email}: {e}")


@require_POST
@csrf_exempt
def update_schedule_status(request, schedule_id):
    """Update vaccination schedule status with async notifications"""
    try:
        from .models import VaccinationSchedule, PreschoolerActivityLog

        schedule = get_object_or_404(VaccinationSchedule, id=schedule_id)
        data = json.loads(request.body)
        new_status = data.get('status')

        if new_status not in ['scheduled', 'completed', 'rescheduled', 'missed']:
            return JsonResponse({'success': False, 'message': 'Invalid status'})

        old_status = schedule.status
        preschooler = schedule.preschooler

        # Prevent duplicate completion
        if old_status == 'completed' and new_status == 'completed':
            return JsonResponse({'success': False, 'message': 'Already completed'})

        # ✅ NEW CHECK: Prevent early completion
        if new_status == 'completed':
            immunization_date = schedule.scheduled_date
            if immunization_date and now().date() < immunization_date:
                return JsonResponse({
                    'success': False,
                    'message': f'You can only mark this as completed on or after {immunization_date.strftime("%B %d, %Y")}.'
                })

            schedule.status = new_status
            schedule.completion_date = timezone.now()
            schedule.administered_date = timezone.now().date()
            schedule.confirmed_by_parent = True
            schedule.save()

            current_dose = schedule.doses or 1
            required_doses = schedule.required_doses or 1

            completed_count = VaccinationSchedule.objects.filter(
                preschooler=schedule.preschooler,
                vaccine_name=schedule.vaccine_name,
                status='completed'
            ).count()

            PreschoolerActivityLog.objects.create(
                preschooler_name=f"{preschooler.first_name} {preschooler.last_name}",
                activity=f"Vaccination completed: {schedule.vaccine_name} (Dose {current_dose})",
                performed_by=request.user.username if request.user.is_authenticated else 'System',
                barangay=schedule.preschooler.barangay
            )

            # === Async notifications ===
            parents = preschooler.parents.all()
            threading.Thread(
                target=send_completion_notifications_async,
                args=(parents, preschooler, schedule, completed_count, required_doses)
            ).start()

            if completed_count < required_doses:
                return JsonResponse({
                    'success': True,
                    'message': f'Dose {completed_count} completed. {required_doses - completed_count} doses remaining.',
                    'needs_next_dose': True,
                    'completed_doses': completed_count,
                    'total_doses': required_doses,
                    'vaccine_name': schedule.vaccine_name
                })
            else:
                return JsonResponse({
                    'success': True,
                    'message': f'All {required_doses} doses completed for {schedule.vaccine_name}!',
                    'fully_completed': True
                })

        else:
            schedule.status = new_status
            schedule.save()
            status_messages = {
                'scheduled': 'Vaccination rescheduled successfully',
                'rescheduled': 'Vaccination marked as rescheduled',
                'missed': 'Vaccination marked as missed'
            }
            return JsonResponse({'success': True, 'message': status_messages.get(new_status, 'Status updated')})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON'})
    except Exception as e:
        logger.error(f"[ERROR] update_schedule_status: {e}")
        return JsonResponse({'success': False, 'message': f'Error updating status: {str(e)}'})



def send_reschedule_notifications_async(parent, account, preschooler, schedule, old_date, new_date, reschedule_reason):
    """Background thread for reschedule notifications"""
    try:
        # === Email notification ===
        if parent.email:
            try:
                subject = f"[PPMS] Vaccination Rescheduled for {preschooler.first_name}"
                
                # Prepare context for HTML email
                email_context = {
                    'parent_name': parent.full_name,
                    'child_name': f"{preschooler.first_name} {preschooler.last_name}",
                    'vaccine_name': schedule.vaccine_name,
                    'dose_number': schedule.doses,
                    'required_doses': schedule.required_doses,
                    'old_date': old_date.strftime('%B %d, %Y'),
                    'new_date': new_date.strftime('%B %d, %Y'),
                    'reschedule_reason': reschedule_reason
                }
                
                # Generate HTML and plain text emails
                html_message = render_reschedule_vaccination_email_html(email_context)
                plain_message = render_reschedule_vaccination_email_text(email_context)

                send_mail(
                    subject,
                    plain_message,
                    settings.DEFAULT_FROM_EMAIL,
                    [parent.email],
                    html_message=html_message,
                    fail_silently=False
                )
                logger.info(f"[ASYNC] Reschedule email sent to {parent.email}")
            except Exception as email_error:
                logger.error(f"[ASYNC] Reschedule email failed for {parent.email}: {email_error}")

        # === Push notification ===
        if account and account.fcm_token:
            try:
                notification_title = f"Vaccination Rescheduled - {preschooler.first_name}"
                notification_body = (
                    f"{schedule.vaccine_name} moved from {old_date.strftime('%b %d')} "
                    f"to {new_date.strftime('%b %d, %Y')}. "
                    f"Reason: {reschedule_reason}"
                )

                notification_data = {
                    "type": "vaccination_reschedule",
                    "preschooler_id": str(preschooler.preschooler_id),
                    "preschooler_name": f"{preschooler.first_name} {preschooler.last_name}",
                    "vaccine_name": schedule.vaccine_name,
                    "old_date": str(old_date),
                    "new_date": str(new_date),
                    "reason": reschedule_reason,
                    "schedule_id": str(schedule.id)
                }

                PushNotificationService.send_push_notification(
                    token=account.fcm_token,
                    title=notification_title,
                    body=notification_body,
                    data=notification_data
                )
                logger.info(f"[ASYNC] Reschedule push sent to {parent.email}")
            except Exception as push_error:
                logger.error(f"[ASYNC] Reschedule push failed for {parent.email}: {push_error}")
        else:
            logger.warning(f"[ASYNC] No FCM token for {parent.email}")

    except Exception as e:
        logger.error(f"[ASYNC] Error in reschedule notification for {parent.email}: {e}")


def render_reschedule_vaccination_email_html(context):
    """Generate HTML vaccination reschedule email"""
    html = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>PPMS Vaccination Reschedule</title>
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
            
            body {{
                font-family: 'Inter', Arial, sans-serif;
                background-color: #f9fafb;
                padding: 40px 20px;
                color: #334155;
                line-height: 1.6;
            }}
            
            .container {{
                max-width: 560px;
                margin: 0 auto;
                background: white;
                border-radius: 12px;
                overflow: hidden;
                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
            }}
            
            .header {{
                padding: 24px 32px;
                text-align: center;
                color: #111827;
                border-bottom: 4px solid #198754;
            }}
            .header h1 {{
                font-size: 24px;
                font-weight: 600;
                margin: 0;
                color: #111827;
            }}
            .header p {{
                font-size: 16px;
                margin: 4px 0 0 0;
                color: #6b7280;
            }}
            
            .content {{
                padding: 32px;
            }}
            
            .greeting {{
                font-size: 18px;
                margin: 0 0 16px 0;
                color: #1e293b;
                text-align: left;
            }}
            
            .salutation {{
                font-size: 16px;
                margin: 0 0 16px 0;
                color: #64748b;
                text-align: left;
            }}
            
            .message {{
                font-size: 16px;
                margin: 0 0 24px 0;
                color: #64748b;
                text-align: left;
                padding: 0;
            }}
            
            .schedule-details {{
                background-color: #f1f5f9;
                padding: 16px;
                border-radius: 8px;
                margin-bottom: 24px;
                border-left: 4px solid #198754;
            }}
            
            .detail-row {{
                display: flex;
                justify-content: space-between;
                padding: 8px 0;
                font-size: 14px;
                color: #334155;
                border-bottom: 1px solid #e2e8f0;
            }}
            
            .detail-row:last-child {{
                border-bottom: none;
            }}
            
            .detail-label {{
                font-weight: 600;
                color: #1e293b;
            }}
            
            .detail-value {{
                color: #64748b;
                text-align: right;
            }}
            
            .date-comparison {{
                margin-top: 16px;
                padding-top: 16px;
                border-top: 1px solid #e2e8f0;
            }}
            
            .old-date {{
                color: #ef4444;
                text-decoration: line-through;
            }}
            
            .new-date {{
                color: #198754;
                font-weight: 600;
            }}
            
            .reason-box {{
                background-color: #fef3c7;
                padding: 12px;
                border-radius: 8px;
                margin-bottom: 24px;
                border-left: 4px solid #f59e0b;
                font-size: 14px;
                color: #92400e;
            }}
            
            .footer {{
                background: #f1f5f9;
                padding: 32px;
                text-align: center;
                border-top: 1px solid #e2e8f0;
            }}
            
            .footer h3 {{
                font-size: 18px;
                font-weight: 600;
                color: #1e293b;
                margin-bottom: 8px;
            }}
            
            .footer p {{
                font-size: 14px;
                color: #64748b;
                margin-bottom: 4px;
            }}
            
            .footer-divider {{
                margin: 24px 0;
                height: 1px;
                background: #e2e8f0;
            }}
            
            .footer-small {{
                font-size: 12px;
                color: #94a3b8;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>PPMS Cluster 4</h1>
                <p>Imus City Healthcare Management</p>
            </div>
            
            <div class="content">
                <div class="greeting">
                    Hello <strong>{context['parent_name']}</strong>,
                </div>
                
                <div class="salutation">
                    Dear Parent,
                </div>
                
                <div class="message">
                    The vaccination schedule for your child <strong>{context['child_name']}</strong> has been rescheduled.
                </div>
                
                <div class="schedule-details">
                    <div class="detail-row">
                        <span class="detail-label">Vaccine:</span>
                        <span class="detail-value">{context['vaccine_name']}</span>
                    </div>
                    <div class="detail-row">
                        <span class="detail-label">Dose:</span>
                        <span class="detail-value">{context['dose_number']} of {context['required_doses']}</span>
                    </div>
                    <div class="date-comparison">
                        <div style="margin-bottom: 8px;">
                            <span class="detail-label">Previous Date:</span>
                            <span class="old-date">{context['old_date']}</span>
                        </div>
                        <div>
                            <span class="detail-label">New Date:</span>
                            <span class="new-date">{context['new_date']}</span>
                        </div>
                    </div>
                </div>
                
                {'<div class="reason-box"><strong>Reason for Rescheduling:</strong><br>' + context['reschedule_reason'] + '</div>' if context.get('reschedule_reason') else ''}
                
                <div class="message">
                    Please update your calendar with the new date and proceed to your barangay health center accordingly. If you have any questions or concerns, please contact us immediately.
                </div>
            </div>
            
            <div class="footer">
                <h3>PPMS Cluster 4</h3>
                <p>Imus City Healthcare Management</p>
                <div class="footer-divider"></div>
                <p class="footer-small">
                    This is an automated message. Please do not reply.<br>
                    © 2025 PPMS Cluster 4. All rights reserved.
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    return html


def render_reschedule_vaccination_email_text(context):
    """Generate plain text vaccination reschedule email"""
    reason_text = f"Reason for Rescheduling: {context['reschedule_reason']}\n\n" if context.get('reschedule_reason') else ""
    
    text = f"""
PPMS Vaccination Schedule Rescheduled

Hello {context['parent_name']},

Dear Parent,

The vaccination schedule for your child {context['child_name']} has been rescheduled.

VACCINATION DETAILS:
Vaccine: {context['vaccine_name']}
Dose: {context['dose_number']} of {context['required_doses']}

SCHEDULE CHANGE:
Previous Date: {context['old_date']}
New Date: {context['new_date']}

{reason_text}Please update your calendar with the new date and proceed to your barangay health center accordingly. If you have any questions or concerns, please contact us immediately.

PPMS Cluster 4
Imus City Healthcare Management

This is an automated message. Please do not reply.
© 2025 PPMS Cluster 4. All rights reserved.
    """
    return text.strip()


@require_POST
def reschedule_vaccination(request):
    """Handle vaccination rescheduling with async notifications"""
    schedule_id = request.POST.get('schedule_id')
    new_date = request.POST.get('new_date')
    reschedule_reason = request.POST.get('reschedule_reason', '')

    try:
        from .models import VaccinationSchedule, PreschoolerActivityLog, Account

        schedule = get_object_or_404(VaccinationSchedule, id=schedule_id)
        preschooler = schedule.preschooler

        # Parse and update
        new_schedule_date = datetime.strptime(new_date, '%Y-%m-%d').date()
        old_date = schedule.scheduled_date

        schedule.scheduled_date = new_schedule_date
        schedule.status = 'rescheduled'
        schedule.reschedule_reason = reschedule_reason
        schedule.save()

        logger.info(f"[DEBUG] Vaccination rescheduled: {schedule.vaccine_name} from {old_date} to {new_schedule_date}")

        PreschoolerActivityLog.objects.create(
            preschooler_name=f"{preschooler.first_name} {preschooler.last_name}",
            activity=f"Vaccination rescheduled: {schedule.vaccine_name} from {old_date} to {new_schedule_date}",
            performed_by=request.user.email if hasattr(request.user, 'email') else 'System',
            barangay=preschooler.barangay
        )

        # === Async notifications per parent ===
        parents = preschooler.parents.all()
        for parent in parents:
            account = Account.objects.filter(email=parent.email).first()
            threading.Thread(
                target=send_reschedule_notifications_async,
                args=(parent, account, preschooler, schedule, old_date, new_schedule_date, reschedule_reason)
            ).start()

        messages.success(
            request,
            f'Vaccination successfully rescheduled to {new_schedule_date.strftime("%B %d, %Y")}. Notifications are being sent.'
        )

    except Exception as e:
        logger.error(f"[ERROR] Failed to reschedule vaccination: {e}")
        messages.error(request, f'Error rescheduling vaccination: {str(e)}')
        return redirect(request.META.get('HTTP_REFERER', '/'))

    return redirect('preschooler_detail', preschooler_id=schedule.preschooler.preschooler_id)
    
def update_preschooler_photo(request, preschooler_id):
    preschooler = get_object_or_404(Preschooler, preschooler_id=preschooler_id)

    if request.method == 'POST' and 'profile_photo' in request.FILES:
        preschooler.profile_photo = request.FILES['profile_photo']
        preschooler.save()
        return JsonResponse({
            'success': True,
            'new_photo_url': preschooler.profile_photo.url
        })

    return JsonResponse({'success': False}, status=400)

from django.db.models import Prefetch
from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Prefetch
from .models import Preschooler, BMI, Temperature
from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Prefetch
from .models import Preschooler, BMI, Temperature

from datetime import date
from django.shortcuts import get_object_or_404, render
from django.db.models import Prefetch
from django.core.paginator import Paginator
import json

def auto_archive_aged_preschoolers():
    """
    Simple auto-archive function - just sets is_archived=True for 60+ month olds
    """
    # Get all active preschoolers
    active_preschoolers = Preschooler.objects.filter(is_archived=False)
    archived_count = 0
    today = date.today()
    
    for preschooler in active_preschoolers:
        # Use the model's age_in_months property
        age_in_months = preschooler.age_in_months
        
        # Archive if 60+ months old
        if age_in_months and age_in_months >= 60:
            preschooler.is_archived = True
            preschooler.save()
            
            print("AUTO-ARCHIVED: {preschooler.first_name} {preschooler.last_name} - Age: {age_in_months} months")
            archived_count += 1
    
    return archived_count



def preschoolers(request):

    if not request.user.is_authenticated:
        return redirect('login')
        
    user_email = request.session.get('email')
    raw_role = (request.session.get('user_role') or '').strip().lower()

   
    account = get_object_or_404(Account, email=user_email)

   
    auto_archived_count = auto_archive_aged_preschoolers()
    if auto_archived_count > 0:
        print(f"AUTO-ARCHIVED: {auto_archived_count} preschoolers aged out (60+ months)")

   
    if raw_role == 'admin':
        preschoolers_qs = Preschooler.objects.filter(is_archived=False)
        barangay_name = "All Barangays"
    else:
        preschoolers_qs = Preschooler.objects.filter(
            is_archived=False,
            barangay=account.barangay
        )
        barangay_name = account.barangay.name if account.barangay else "No Barangay"
      
    preschoolers_qs = preschoolers_qs.select_related('parent_id', 'barangay') \
        .prefetch_related(
            Prefetch('bmi_set', queryset=BMI.objects.order_by('-date_recorded'), to_attr='bmi_records'),
            Prefetch('temperature_set', queryset=Temperature.objects.order_by('-date_recorded'), to_attr='temp_records')
        )

    today = date.today()

   
    for p in preschoolers_qs:
        
        if not p.middle_name:
            p.middle_name = ''
        if not p.suffix:
            p.suffix = ''
            
        # Get latest BMI
        latest_bmi = None
        if hasattr(p, 'bmi_records') and p.bmi_records:
            latest_bmi = p.bmi_records[0]
        else:
            latest_bmi = p.bmi_set.order_by('-date_recorded').first()

        # Nutritional status (using z-score)
        if latest_bmi and latest_bmi.bmi_value:
            try:
                # Use the model's age_in_months property
                total_age_months = p.age_in_months

                # Compute z-score & classify
                z = bmi_zscore(p.sex, total_age_months, latest_bmi.bmi_value)
                p.nutritional_status = classify_bmi_for_age(z)
            except Exception as e:
                print(f"Error computing BMI classification for {p.first_name}: {e}")
                p.nutritional_status = "Error"
        else:
            p.nutritional_status = None

        # Delivery place color coding
        delivery_place = getattr(p, 'place_of_delivery', None)

        if delivery_place == 'Center to Center':
            p.delivery_class = 'delivery-center'
        elif delivery_place == 'Private/Lying-in':
            p.delivery_class = 'delivery-lying-in'
        elif delivery_place == 'Public Hospital':
            p.delivery_class = 'delivery-hospital'
        elif delivery_place == 'Others':
            p.delivery_class = 'delivery-others'
        else:
            p.delivery_class = 'delivery-na'

    # Convert to list for filtering
    preschoolers_qs = list(preschoolers_qs)

    # ✅ FILTER BY NUTRITIONAL STATUS (if provided in URL)
    filter_status = request.GET.get('status', 'All')
    if filter_status and filter_status != 'All':
        preschoolers_qs = [p for p in preschoolers_qs if p.nutritional_status == filter_status]

    # ✅ GLOBAL SEARCH - Search by preschooler name ACROSS ALL DATA
    search_query = request.GET.get('search', '').strip()
    is_searching = False
    
    if search_query:
        is_searching = True
        search_lower = search_query.lower()
        preschoolers_qs = [
            p for p in preschoolers_qs 
            if search_lower in f"{p.first_name} {p.last_name}".lower()
        ]

    # ✅ Pagination AFTER filtering and searching
    paginator = Paginator(preschoolers_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Determine user role for template
    if raw_role in ['bhw', 'bns', 'midwife', 'nurse']:
        template_user_role = 'health_worker'
    else:
        template_user_role = raw_role

    context = {
        'account': account, 
        'preschoolers': page_obj,
        'user_email': user_email,
        'user_role': template_user_role,
        'original_role': raw_role,
        'barangay_name': barangay_name,
        'filter_status': filter_status,  
        'search_query': search_query,     
        'is_searching': is_searching,     
    }

    return render(request, 'HTML/preschoolers.html', context)


@login_required
def profile(request):
    """User profile view with Cloudinary photo upload"""
    if not request.user.is_authenticated:
        return redirect('login')

    try:
        account = Account.objects.select_related('profile_photo', 'barangay').get(email=request.user.email)
    except Account.DoesNotExist:
        messages.error(request, "Account not found. Please contact administrator.")
        return redirect('login')

    # ✅ SIMPLIFIED ADDRESS BUILDING FUNCTION
    def build_complete_address(account):
        """Build complete address from Account model individual fields"""
        # Check if there's already an editable_address
        if hasattr(account, 'editable_address') and account.editable_address and account.editable_address.strip():
            if account.editable_address.strip().lower() not in ['n/a', 'none', 'no address provided']:
                return account.editable_address.strip()
        
        invalid_values = {'na', 'n/a', 'none', ''}
        address_parts = []

        # House number first
        if account.house_number and str(account.house_number).strip().lower() not in invalid_values:
            address_parts.append(f"{account.house_number}")

        # Then Block, Lot, Phase
        if account.block and str(account.block).strip().lower() not in invalid_values:
            address_parts.append(f"Block {account.block}")

        if account.lot and str(account.lot).strip().lower() not in invalid_values:
            address_parts.append(f"Lot {account.lot}")

        if account.phase and str(account.phase).strip().lower() not in invalid_values:
            address_parts.append(f"Phase {account.phase}")

        # Then Street
        if account.street and str(account.street).strip().lower() not in invalid_values:
            address_parts.append(account.street.strip())

        # Then Subdivision
        if account.subdivision and str(account.subdivision).strip().lower() not in invalid_values:
            address_parts.append(account.subdivision.strip())

        # City
        if account.city and str(account.city).strip().lower() not in invalid_values:
            address_parts.append(account.city.strip())

        # Province
        if account.province and str(account.province).strip().lower() not in invalid_values:
            address_parts.append(account.province.strip())

        # Build final address
        if address_parts:
            built_address = ", ".join(address_parts)
            if not account.editable_address:
                account.editable_address = built_address
                account.save()
            return built_address

        # ✅ FOR PARENTS, CHECK PARENT MODEL AS FALLBACK
        if account.user_role and account.user_role.lower() == 'parent':
            try:
                parent = Parent.objects.get(email=account.email)
                if parent.address and parent.address.strip():
                    parent_address = parent.address.strip()
                    # Also update the account's editable_address
                    if not account.editable_address:
                        account.editable_address = parent_address
                        account.save()
                    return parent_address
            except Parent.DoesNotExist:
                pass
        
        return "No address provided"

    # ✅ SET COMPLETE ADDRESS
    account.complete_address = build_complete_address(account)

    if request.method == 'POST':
        # ✅ HANDLE PHOTO UPLOAD WITH CLOUDINARY
        if 'photo' in request.FILES:
            photo_file = request.FILES['photo']
            try:
                # Upload to Cloudinary
                upload_result = cloudinary.uploader.upload(
                    photo_file,
                    folder='profile_photos/',
                    public_id=f'profile_{account.account_id}',
                    overwrite=True,
                    transformation=[
                        {'width': 500, 'height': 500, 'crop': 'fill', 'gravity': 'face'},
                        {'quality': 'auto:good'}
                    ]
                )

                cloudinary_url = upload_result.get('secure_url')
                cloudinary_public_id = upload_result.get('public_id')

                if hasattr(account, 'profile_photo') and account.profile_photo:
                    # Delete old image from Cloudinary
                    if account.profile_photo.cloudinary_public_id:
                        try:
                            cloudinary.uploader.destroy(account.profile_photo.cloudinary_public_id)
                        except Exception as e:
                            print(f"Error deleting old image: {e}")
                    
                    account.profile_photo.image = cloudinary_url
                    account.profile_photo.cloudinary_public_id = cloudinary_public_id
                    account.profile_photo.save()
                else:
                    ProfilePhoto.objects.create(
                        account=account,
                        image=cloudinary_url,
                        cloudinary_public_id=cloudinary_public_id
                    )
                
                messages.success(request, "Profile photo updated successfully.")
                return redirect('profile')
            except Exception as e:
                messages.error(request, f"Error uploading photo: {str(e)}")
                return redirect('profile')

        # Get form data
        full_name = request.POST.get('full_name', '').strip()
        address = request.POST.get('address', '').strip()
        contact = request.POST.get('contact_number', '').strip()
        birthdate = request.POST.get('birthdate')
        barangay_id = request.POST.get('barangay')

        # ✅ VALIDATE FULL NAME
        if full_name and full_name.lower() not in ['na', 'n/a', 'none', '']:
            # Split the full name into first, middle, and last names
            name_parts = full_name.split()
            
            if len(name_parts) >= 2:
                account.first_name = name_parts[0]
                account.last_name = name_parts[-1]
                
                # If there are middle names, join them
                if len(name_parts) > 2:
                    account.middle_name = ' '.join(name_parts[1:-1])
                else:
                    account.middle_name = ''
            elif len(name_parts) == 1:
                account.first_name = name_parts[0]
                account.last_name = ''
                account.middle_name = ''
            
            # Also update the full_name field
            account.full_name = full_name
            
            # ✅ If parent, also update Parent model
            if account.user_role and account.user_role.lower() == 'parent':
                try:
                    parent = Parent.objects.get(email=account.email)
                    parent.full_name = full_name
                    parent.save()
                except Parent.DoesNotExist:
                    pass

        # ✅ Validate contact number (must be exactly 11 digits)
        if contact:
            if not contact.isdigit():
                messages.error(request, "Contact number must contain only numbers.")
                return redirect('profile')
            
            if len(contact) != 11:
                messages.error(request, "Contact number must be exactly 11 digits.")
                return redirect('profile')
            
            account.contact_number = contact
            
            # ✅ If parent, also update Parent model
            if account.user_role and account.user_role.lower() == 'parent':
                try:
                    parent = Parent.objects.get(email=account.email)
                    parent.contact_number = contact
                    parent.save()
                except Parent.DoesNotExist:
                    pass

        # ✅ Update birthdate
        if birthdate:
            account.birthdate = birthdate or None

        # ✅ UPDATE ADDRESS - Store in both editable_address and complete_address
        if address:
            account.editable_address = address
            account.complete_address = address
            
            # If parent, update the parent model too
            if account.user_role and account.user_role.lower() == 'parent':
                try:
                    parent = Parent.objects.get(email=account.email)
                    parent.address = address
                    parent.save()
                except Parent.DoesNotExist:
                    pass

        # ✅ Handle barangay selection
        if barangay_id:
            try:
                barangay = Barangay.objects.get(id=barangay_id)
                if account.barangay != barangay:
                    old_barangay = account.barangay
                    account.barangay = barangay

                    if account.user_role.lower() == 'parent':
                        try:
                            parent = Parent.objects.get(email=account.email)
                            parent.barangay = barangay
                            parent.save()

                            # Update preschoolers
                            Preschooler.objects.filter(parent_id=parent).update(barangay=barangay)

                            # Log transfer
                            if old_barangay:  
                                ParentActivityLog.objects.create(
                                    parent=parent,
                                    activity=f"Transferred to {barangay.name}",
                                    barangay=old_barangay,
                                    timestamp=timezone.now()
                                )
                                ParentActivityLog.objects.create(
                                    parent=parent,
                                    activity=f"Recently transferred from {old_barangay.name}",
                                    barangay=barangay,
                                    timestamp=timezone.now()
                                )

                                for p in Preschooler.objects.filter(parent_id=parent):
                                    PreschoolerActivityLog.objects.create(
                                        preschooler_name=f"{p.first_name} {p.last_name}",
                                        performed_by=parent.full_name,
                                        activity=f"Transferred to {barangay.name}",
                                        barangay=old_barangay,
                                        timestamp=timezone.now()
                                    )
                                    PreschoolerActivityLog.objects.create(
                                        preschooler_name=f"{p.first_name} {p.last_name}",
                                        performed_by=parent.full_name,
                                        activity=f"Recently transferred from {old_barangay.name}",
                                        barangay=barangay,
                                        timestamp=timezone.now()
                                    )
                        except Parent.DoesNotExist:
                            pass
            except Barangay.DoesNotExist:
                messages.error(request, "Selected barangay does not exist.")
                return redirect('profile')

        # ✅ SAVE THE ACCOUNT
        account.save()
        
        # Refresh the complete address after saving
        account.complete_address = build_complete_address(account)

        messages.success(request, "Profile updated successfully.")
        return redirect('profile')

    # Get dashboard URL based on role
    role = account.user_role.lower() if account.user_role else ''
    if role == 'parent':
        dashboard_url = reverse('parent_dashboard')
    elif role == 'admin':
        dashboard_url = reverse('Admindashboard') 
    else:
        dashboard_url = reverse('dashboard')

    barangays = Barangay.objects.all()

    return render(request, 'HTML/profile.html', {
        'account': account,
        'dashboard_url': dashboard_url,
        'barangays': barangays,
    })
    
def registered_parents(request):
    # ✅ Redirect if not authenticated
    if not request.user.is_authenticated:
        return redirect('login')

    # ✅ Get the current user's account
    account = get_object_or_404(Account, email=request.user.email)
    raw_role = (account.user_role or '').strip().lower()

    print("=== REGISTERED PARENTS VIEW DEBUG ===")
    print(f"User: {account.full_name} ({account.user_role})")

    # ✅ Query parents
    if raw_role == 'admin':
        parents_qs = Parent.objects.all().order_by('-created_at')
        barangay_name = "All Barangays"
    else:
        parents_qs = Parent.objects.filter(
            barangay=account.barangay
        ).order_by('-created_at')
        barangay_name = account.barangay.name if account.barangay else "No Barangay"
        print(f"Showing parents for barangay: {barangay_name}")

    # ✅ GLOBAL SEARCH - Search by parent name ACROSS ALL DATA
    search_query = request.GET.get('search', '').strip()
    is_searching = False
    
    if search_query:
        is_searching = True
        search_lower = search_query.lower()
        # Convert to list and filter by full name
        parents_qs = [
            p for p in parents_qs 
            if search_lower in f"{p.first_name} {p.last_name}".lower()
        ]
        print(f"Search query: '{search_query}' - Found {len(parents_qs)} parents")
    else:
        # Convert to list for consistency
        parents_qs = list(parents_qs)

    # ✅ Pagination AFTER filtering and searching
    paginator = Paginator(parents_qs, 10)  # 10 parents per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ✅ Compute children + age
    today = date.today()
    for parent in page_obj:
        preschoolers = Preschooler.objects.filter(parent_id=parent)
        for child in preschoolers:
            years = today.year - child.birth_date.year
            months = today.month - child.birth_date.month
            if today.day < child.birth_date.day:
                months -= 1
            if months < 0:
                years -= 1
                months += 12
            child.age_text = f"{years} year(s) and {months} month(s)"
        parent.children = preschoolers

    context = {
        'account': account,
        'parents': page_obj,
        'barangay_name': barangay_name,
        'has_parents': len(parents_qs) > 0,  
        'search_query': search_query,  
        'is_searching': is_searching,  
    }

    print(f"Parents count: {len(parents_qs)}")
    print("=== END REGISTERED PARENTS DEBUG ===")

    return render(request, 'HTML/registered_parent.html', context)
    
def register(request):
    if request.method == 'POST':
        first_name   = request.POST.get("firstName")
        middle_name  = request.POST.get("middleName")
        suffix       = request.POST.get("suffix")
        last_name    = request.POST.get("lastName")
        email        = request.POST.get("email")
        contact      = request.POST.get("contact")
        password     = request.POST.get("password")
        confirm      = request.POST.get("confirm")
        birthdate    = request.POST.get("birthdate")
        sex          = request.POST.get("sex")
        house_number = request.POST.get("house_number")   
        block        = request.POST.get("block")
        lot          = request.POST.get("lot")
        phase        = request.POST.get("phase")
        street       = request.POST.get("street")
        subdivision  = request.POST.get("subdivision")  # ✅ FIXED
        city         = request.POST.get("city")         # ✅ FIXED
        province     = request.POST.get("province")
        barangay_id  = request.POST.get("barangay_id")
        role         = request.POST.get("role")

        # --- VALIDATIONS ---
        if not all([first_name, last_name, email, contact, password, confirm, birthdate, sex,
                    house_number, block, lot, phase, street, subdivision, city, province, barangay_id, role]):
            messages.error(request, "Please fill out all required fields.")
            return render(request, 'HTML/register.html', {'barangays': Barangay.objects.all()})

        if password != confirm:
            messages.error(request, "Passwords do not match.")
            return render(request, 'HTML/register.html', {'barangays': Barangay.objects.all()})

        # ✅ FIXED: Check both tables but handle cleanup
        user_exists = User.objects.filter(username=email).exists()
        account_exists = Account.objects.filter(email=email).exists()
        
        if user_exists or account_exists:
            # If User exists but Account doesn't (orphaned User), delete the User
            if user_exists and not account_exists:
                try:
                    User.objects.filter(username=email).delete()
                    print(f"[DEBUG] 🗑️ Deleted orphaned User for email: {email}")
                except Exception as e:
                    print(f"[DEBUG] ❌ Error deleting orphaned User: {e}")
                    messages.error(request, "An error occurred. Please try again.")
                    return render(request, 'HTML/register.html', {'barangays': Barangay.objects.all()})
            else:
                # Both exist or Account exists
                messages.error(request, "This email is already registered.")
                return render(request, 'HTML/register.html', {'barangays': Barangay.objects.all()})

        # Convert birthdate string to date object
        try:
            birthdate_obj = datetime.strptime(birthdate, '%Y-%m-%d').date()
        except ValueError as e:
            print(f"[DEBUG] ❌ Birthdate conversion error: {e}")
            messages.error(request, "Invalid birthdate format. Please try again.")
            return render(request, 'HTML/register.html', {'barangays': Barangay.objects.all()})

        try:
            # Step 1: Create Django User
            print("[DEBUG] Creating Django User...")
            user = User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )

            # Step 2: Get Barangay
            barangay = Barangay.objects.get(id=int(barangay_id))

            # Step 3: Create Account with sex included
            print("[DEBUG] Creating Account with all info...")
            account = Account.objects.create(
                first_name=first_name,
                middle_name=middle_name,
                suffix=suffix,
                last_name=last_name,
                email=email,
                contact_number=contact,
                house_number=house_number,
                block=block,
                lot=lot,
                phase=phase,
                street=street,
                subdivision=subdivision,
                city=city,
                province=province,
                birthdate=birthdate_obj,
                sex=sex,
                password=make_password(password),
                user_role=role,
                is_validated=False,
                is_rejected=False,
                barangay=barangay
            )

        except Barangay.DoesNotExist:
            # Cleanup: Delete created User if Account creation fails
            if 'user' in locals():
                user.delete()
            messages.error(request, "Invalid barangay selected.")
            return render(request, 'HTML/register.html', {'barangays': Barangay.objects.all()})
        except Exception as e:
            # Cleanup: Delete created User if Account creation fails
            if 'user' in locals():
                user.delete()
            print(f"[DEBUG] ❌ Registration error: {e}")
            messages.error(request, f"Registration failed: {str(e)}")
            return render(request, 'HTML/register.html', {'barangays': Barangay.objects.all()})

        # Prepare email data
        full_name_parts = [first_name]
        if middle_name:
            full_name_parts.append(middle_name)
        full_name_parts.append(last_name)
        if suffix:
            full_name_parts.append(suffix)
        full_name = " ".join(full_name_parts)
        
        # Role-specific badge classes
        role_classes = {
            'BHW': 'bhw',
            'BNS': 'bns', 
            'Midwife': 'midwife',
            'Nurse': 'nurse'
        }
        role_class = role_classes.get(role, 'bhw')
        
        # Role display names
        role_display = {
            'BHW': 'BHW (Barangay Health Worker)',
            'BNS': 'BNS (Barangay Nutrition Scholar)',
            'Midwife': 'Midwife',
            'Nurse': 'Nurse'
        }
        role_name = role_display.get(role, role)
        
        # Get current date
        current_date = datetime.now().strftime('%B %d, %Y')
        
        subject = f'PPMS Registration Confirmation - {role}'
        
        html_message = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PPMS Registration Confirmation</title>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
        
        body {{
            font-family: 'Inter', Arial, sans-serif;
            background-color: #f9fafb;
            padding: 40px 20px;
            color: #334155;
            line-height: 1.6;
        }}
        
        .container {{
            max-width: 560px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        }}
        
        .header {{
            padding: 24px 32px;
            text-align: center;
            color: #111827;
            border-bottom: 4px solid #198754;
        }}
        .header h1 {{
            font-size: 24px;
            font-weight: 600;
            margin: 0;
            color: #111827;
        }}
        .header p {{
            font-size: 16px;
            margin: 4px 0 0 0;
            color: #6b7280;
        }}
        
        .content {{
            padding: 32px;
        }}
        
        .greeting {{
            font-size: 18px;
            margin-bottom: 24px;
            color: #1e293b;
        }}
        
        .message {{
            font-size: 16px;
            margin-bottom: 32px;
            color: #64748b;
        }}
        
        .status {{
            background: #fef3c7;
            border: 1px solid #f59e0b;
            border-radius: 8px;
            padding: 20px;
            margin-bottom: 32px;
            text-align: center;
        }}
        
        .status-icon {{
            font-size: 32px;
            margin-bottom: 12px;
        }}
        
        .status h3 {{
            font-size: 18px;
            font-weight: 600;
            color: #92400e;
            margin-bottom: 8px;
        }}
        
        .status p {{
            font-size: 14px;
            color: #92400e;
        }}
        
        .details {{
            background: #ecfdf5;
            border: 1px solid #10b981;
            padding: 24px;
            border-radius: 6px;
            margin: 24px 0;
        }}
        
        .details h4 {{
            margin-bottom: 16px;
            font-size: 16px;
            font-weight: 600;
            color: #065f46;
        }}
        
        .detail-item {{
            margin-bottom: 12px;
            line-height: 1.6;
        }}
        
        .detail-item:last-child {{
            margin-bottom: 0;
        }}
        
        .detail-label {{
            font-weight: bold;
            color: #065f46;
            display: inline;
        }}
        
        .detail-value {{
            color: #065f46;
            display: inline;
            margin-left: 4px;
        }}
        
        .role-badge {{
            display: inline-block;
            background: #6366f1;
            color: white;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 500;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        
        .role-badge.bhw {{ background: #10b981; }}
        .role-badge.bns {{ background: #3b82f6; }}
        .role-badge.midwife {{ background: #ec4899; }}
        .role-badge.nurse {{ background: #8b5cf6; }}
        
        .footer {{
            background: #f1f5f9;
            padding: 32px;
            text-align: center;
            border-top: 1px solid #e2e8f0;
        }}
        
        .footer h3 {{
            font-size: 18px;
            font-weight: 600;
            color: #1e293b;
            margin-bottom: 8px;
        }}
        
        .footer p {{
            font-size: 14px;
            color: #64748b;
            margin-bottom: 4px;
        }}
        
        .footer-divider {{
            margin: 24px 0;
            height: 1px;
            background: #e2e8f0;
        }}
        
        .footer-small {{
            font-size: 12px;
            color: #94a3b8;
        }}
        
        @media (max-width: 600px) {{
            .content {{ 
                padding: 28px 20px; 
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>PPMS Cluster 4</h1>
            <p>Imus City Healthcare Management</p>
        </div>
        
        <div class="content">
            <div class="greeting">
                Hello <strong>{full_name}</strong>,
            </div>
            
            <div class="message">
                Thank you for registering with PPMS Cluster 4. We've received your application to join our healthcare team as a <span class="role-badge {role_class}">{role}</span>.
            </div>
            
            <div class="status">
                <div class="status-icon">⏳</div>
                <h3>Pending Approval</h3>
                <p>Your account is under review by our admin team</p>
            </div>
            
            <div class="details">
                <h4>Registration Summary</h4>
                <div class="detail-item">
                    <span class="detail-label">Full Name:</span>
                    <span class="detail-value">{full_name}</span>
                </div>
                <div class="detail-item">
                    <span class="detail-label">Sex:</span>
                    <span class="detail-value">{sex}</span>
                </div>
                <div class="detail-item">
                    <span class="detail-label">Email:</span>
                    <span class="detail-value">{email}</span>
                </div>
                <div class="detail-item">
                    <span class="detail-label">Role:</span>
                    <span class="detail-value">{role_name}</span>
                </div>
                <div class="detail-item">
                    <span class="detail-label">Barangay:</span>
                    <span class="detail-value">{barangay.name}</span>
                </div>
                <div class="detail-item">
                    <span class="detail-label">Date Submitted:</span>
                    <span class="detail-value">{current_date}</span>
                </div>
            </div>
        </div>
        
        <div class="footer">
            <h3>PPMS Cluster 4</h3>
            <p>Imus City Healthcare Management</p>
            <div class="footer-divider"></div>
            <p class="footer-small">
                This is an automated message. Please do not reply.<br>
                © 2025 PPMS Cluster 4. All rights reserved.
            </p>
        </div>
    </div>
</body>
</html>
"""

        plain_message = f"""
PPMS Registration Confirmation

Hello {full_name},

Thank you for registering with PPMS Cluster 4. We've received your application to join our healthcare team as a {role}.

STATUS: Pending Approval
Your account is under review by our admin team.

Registration Summary:
- Full Name: {full_name}
- Sex: {sex}
- Email: {email}
- Role: {role_name}
- Barangay: {barangay.name}
- Date Submitted: {current_date}

PPMS Cluster 4
Imus City Healthcare Management

This is an automated message. Please do not reply.
© 2025 PPMS Cluster 4. All rights reserved.
"""

        # ========== Background email sending ==========
        def send_confirmation_email():
            try:
                send_mail(
                    subject,
                    plain_message,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    html_message=html_message,
                    fail_silently=False,
                )
                print("[DEBUG] ✅ Email sent successfully")
            except Exception as e:
                print(f"[EMAIL ERROR]: {e}")

        # Start email sending in background thread
        threading.Thread(target=send_confirmation_email).start()

        # Success message and redirect
        messages.success(request, "Registration successful! Please check your email for confirmation.")
        return redirect('login')

    # GET request - show registration form
    return render(request, 'HTML/register.html', {'barangays': Barangay.objects.all()})


def register_preschooler(request):
    """Register preschooler with proper barangay filtering - only allows registration in user's barangay"""
    
    if not request.user.is_authenticated:
        return redirect('login')
    
        # AUTO-ARCHIVE CHECK - Run before showing registration
    auto_archived_count = auto_archive_aged_preschoolers()
    if auto_archived_count > 0:
        print(f"AUTO-ARCHIVED: {auto_archived_count} preschoolers during registration view")

    # Get user's barangay using consistent logic
    user_barangay = get_user_barangay(request.user)
    current_user_info = None
    
   
    
    if request.user.is_authenticated:
        # Try each model to find the user and get their info
        try:
            # Try Account model first
            account = Account.objects.select_related('barangay').get(email=request.user.email)
            current_user_info = {
                'model': 'Account',
                'name': account.full_name,
                'role': account.user_role,
                'object': account
            }
         
        except Account.DoesNotExist:
            try:
                # Try BHW model
                bhw = BHW.objects.select_related('barangay').get(email=request.user.email)
                current_user_info = {
                    'model': 'BHW',
                    'name': bhw.full_name,
                    'role': 'BHW',
                    'object': bhw
                }
          
            except BHW.DoesNotExist:
                try:
                    # Try BNS model
                    bns = BNS.objects.select_related('barangay').get(email=request.user.email)
                    current_user_info = {
                        'model': 'BNS',
                        'name': bns.full_name,
                        'role': 'BNS',
                        'object': bns
                    }
                  
                except BNS.DoesNotExist:
                    try:
                        # Try Midwife model
                        midwife = Midwife.objects.select_related('barangay').get(email=request.user.email)
                        current_user_info = {
                            'model': 'Midwife',
                            'name': midwife.full_name,
                            'role': 'Midwife',
                            'object': midwife
                        }
            
                    except Midwife.DoesNotExist:
                        try:
                            # Try Nurse model
                            nurse = Nurse.objects.select_related('barangay').get(email=request.user.email)
                            current_user_info = {
                                'model': 'Nurse',
                                'name': nurse.full_name,
                                'role': 'Nurse',
                                'object': nurse
                            }
                          
                        except Nurse.DoesNotExist:
                            print("DEBUG: User not found in any authorized user model")

    # Validate that user exists and has proper authorization
    if not current_user_info:
        messages.error(request, "You are not authorized to register preschoolers. Please contact the administrator.")
        return redirect('dashboard')

    # Validate that user has a barangay assigned
    if not user_barangay:
        messages.error(request, f"No barangay assigned to your {current_user_info['role']} account. Please contact the administrator to assign a barangay before registering preschoolers.")
        return redirect('dashboard')

    # Validate user role permissions (only for Account model)
    if current_user_info['model'] == 'Account':
        user_role_lower = current_user_info['role'].lower()
        is_authorized = (
            'bhw' in user_role_lower or 
            'health worker' in user_role_lower or
            'bns' in user_role_lower or 
            'nutritional' in user_role_lower or 
            'nutrition' in user_role_lower or
            'scholar' in user_role_lower or
            'midwife' in user_role_lower or
            'admin' in user_role_lower
        )
        
        if not is_authorized:
            messages.error(request, f"Your role '{current_user_info['role']}' is not authorized to register preschoolers. Only BHW, BNS, Midwife, or Admin roles can register preschoolers.")
            return redirect('dashboard')

   

    # Get parents from the SAME barangay only - no cross-barangay registration
    parents_qs = Parent.objects.filter(barangay=user_barangay).order_by('-created_at')
  

    # Debug: Show which parents were found
    for parent in parents_qs[:5]:  # Show first 5 for debugging
        print(f"DEBUG: Parent: {parent.full_name} ({parent.email}) - Barangay: {parent.barangay}")

    # Pagination
    paginator = Paginator(parents_qs, 10)  # Show 10 parents per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'HTML/register_preschooler.html', {
        'account': current_user_info['object'],
        'parents': page_obj,
        'user_barangay': user_barangay,
        'current_user_info': current_user_info,  # For debugging if needed
    })

@csrf_exempt
def register_preschooler_entry(request):
    """Create preschooler entry with barangay validation and WHO BMI integration"""
    
    if request.method == 'POST':
        try:
            # Required fields
            parent_id = request.POST.get('parent_id')
            first_name = request.POST.get('first_name', '').strip()
            middle_name = request.POST.get('middle_name', '').strip() or None  # ✅ save as None if blank
            last_name = request.POST.get('last_name', '').strip()
            suffix = request.POST.get('suffix', '').strip() or None            # ✅ save as None if blank
            birthdate = request.POST.get('birthdate')
            gender = request.POST.get('gender')

            # Validate required fields
            if not all([parent_id, first_name, last_name, birthdate, gender]):
                return JsonResponse({'status': 'error', 'message': 'All required fields must be filled.'})

            # Convert gender to WHO format (M/F)
            if gender.lower() in ['male', 'boy', 'm']:
                sex = 'M'
            elif gender.lower() in ['female', 'girl', 'f']:
                sex = 'F'
            else:
                return JsonResponse({'status': 'error', 'message': 'Invalid gender value. Must be Male or Female.'})

            # Get the registering user's barangay for validation
            user_barangay = get_user_barangay(request.user)
            if not user_barangay:
                return JsonResponse({'status': 'error', 'message': 'No barangay assigned to your account. Please contact administrator.'})

            # Get parent object and validate it's in the same barangay
            try:
                parent = Parent.objects.get(pk=parent_id, barangay=user_barangay)
            except Parent.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Parent not found in your barangay. You can only register preschoolers for parents in your assigned barangay.'})

            # Parse and validate birthdate
            birth_date = datetime.strptime(birthdate, '%Y-%m-%d').date()
            today = date.today()
            age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

            if age < 0 or age > 6:
                return JsonResponse({'status': 'error', 'message': 'Invalid age for preschooler registration. Age must be between 0-6 years.'})

            # Create preschooler (model.save() will build full_name automatically)
            preschooler = Preschooler.objects.create(
                first_name=first_name,
                middle_name=middle_name,
                last_name=last_name,
                suffix=suffix,
                sex=sex,
                birth_date=birth_date,
                age=age,
                address=parent.address,
                parent_id=parent,
                barangay=user_barangay,
                place_of_birth=request.POST.get('place_of_birth') or None,
                birth_weight=request.POST.get('birth_weight') or None,
                birth_height=request.POST.get('birth_length') or None,
                time_of_birth=request.POST.get('time_of_birth') or None,
                type_of_birth=request.POST.get('type_of_birth') or None,
                place_of_delivery=request.POST.get('place_of_delivery') or None,
            )

            parent.registered_preschoolers.add(preschooler)

           
            return JsonResponse({
                'status': 'success', 
                'message': f'Preschooler {preschooler.full_name} registered successfully in {user_barangay.name}!',
                'data': {
                    'preschooler_id': preschooler.preschooler_id,
                    'name': preschooler.full_name,
                    'sex': sex,
                    'age_months': preschooler.age_in_months,
                    'barangay': str(user_barangay)
                }
            })

        except ValueError as e:
            return JsonResponse({'status': 'error', 'message': f'Invalid date format: {str(e)}'})
        except Exception as e:
            print(f"DEBUG: Error in preschooler registration: {e}")
            return JsonResponse({'status': 'error', 'message': f'Registration failed: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})

    
# Additional helper function to get preschoolers by barangay
def get_preschoolers_by_barangay(request):
    """Get preschoolers filtered by user's barangay"""
    user_barangay = get_user_barangay(request.user)
    
    if not user_barangay:
        return Preschooler.objects.none()  # Return empty queryset
    
    return Preschooler.objects.filter(barangay=user_barangay).select_related('parent_id', 'barangay')

def registered_bhw(request):
    if not request.user.is_authenticated:
        return redirect('login')
    # Use same filter pattern as validate function
    bhw_list = Account.objects.filter(
        Q(user_role__iexact='healthworker') | Q(user_role__iexact='BHW'),
        is_validated=True
    )

    # Debug: Print what we found
   
    for bhw in bhw_list:
        print(f"- {bhw.full_name} (role: '{bhw.user_role}', validated: {bhw.is_validated})")

    for bhw in bhw_list:
        bhw.bhw_data = BHW.objects.filter(email=bhw.email).first()

        if bhw.last_activity:
            if timezone.now() - bhw.last_activity <= timedelta(minutes=1):
                bhw.last_activity_display = "🟢 Online"
            else:
                time_diff = timesince(bhw.last_activity, timezone.now())
                bhw.last_activity_display = f"{time_diff} ago"
        else:
            bhw.last_activity_display = "No activity"

    # Get all barangays for the filter dropdown
    barangays = Barangay.objects.all().order_by('name')

    paginator = Paginator(bhw_list, 10)  # 10 BHWs per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'HTML/registered_bhw.html', {
        'bhws': page_obj,
        'barangays': barangays,  # Added this for the filter dropdown
        'total_bhw_count': bhw_list.count()  # Para makita ninyo ang total
    })


def registered_bns(request):

    if not request.user.is_authenticated:
        return redirect('login')
    # Use the same filter pattern as validate function
    bns_list = Account.objects.filter(
        Q(user_role__iexact='bns') | 
        Q(user_role__iexact='BNS') |
        Q(user_role__iexact='Barangay Nutritional Scholar'),
        is_validated=True
    )

    # Debug: Print what we found
 
    for bns in bns_list:
        print(f"- {bns.full_name} (role: '{bns.user_role}', validated: {bns.is_validated})")

    for bns in bns_list:
        if bns.last_activity:
            if timezone.now() - bns.last_activity <= timedelta(minutes=1):
                bns.last_activity_display = "🟢 Online"
            else:
                time_diff = timesince(bns.last_activity, timezone.now())
                bns.last_activity_display = f"{time_diff} ago"
        else:
            bns.last_activity_display = "No activity"

    paginator = Paginator(bns_list, 10)  # 10 BNS per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'HTML/registered_bns.html', {
        'bnss': page_obj,
        'total_bns_count': bns_list.count()  # Para makita ninyo ang total
    })


@admin_required
def registered_preschoolers(request):
    preschoolers_qs = (
        Preschooler.objects.filter(is_archived=False)
        .select_related('parent_id', 'barangay')
        .prefetch_related(
            Prefetch('bmi_set', queryset=BMI.objects.order_by('-date_recorded'), to_attr='bmi_records'),
            Prefetch('temperature_set', queryset=Temperature.objects.order_by('-date_recorded'), to_attr='temp_records')
        )
        .order_by('first_name', 'last_name')
    )

    today = date.today()

    # --- Compute nutritional status and delivery class ---
    for p in preschoolers_qs:
        # Calculate nutritional status
        latest_bmi = p.bmi_records[0] if hasattr(p, 'bmi_records') and p.bmi_records else None

        if latest_bmi:
            try:
                birth_date = p.birth_date
                age_years = today.year - birth_date.year
                age_months = today.month - birth_date.month
                if today.day < birth_date.day:
                    age_months -= 1
                if age_months < 0:
                    age_years -= 1
                    age_months += 12
                total_age_months = age_years * 12 + age_months

                bmi_value = calculate_bmi(latest_bmi.weight, latest_bmi.height)
                z = bmi_zscore(p.sex, total_age_months, bmi_value)
                p.nutritional_status = classify_bmi_for_age(z)
            except Exception as e:
                print(f"⚠️ BMI classification error for preschooler {p.id}: {e}")
                p.nutritional_status = "N/A"
        else:
            p.nutritional_status = "N/A"

        # Set delivery class for row coloring
        delivery_place = getattr(p, 'place_of_delivery', None)
        if delivery_place == 'Center to Center':
            p.delivery_class = 'delivery-center'
        elif delivery_place == 'Private/Lying-in':
            p.delivery_class = 'delivery-lying-in'
        elif delivery_place == 'Public Hospital':
            p.delivery_class = 'delivery-hospital'
        elif delivery_place == 'Others':
            p.delivery_class = 'delivery-others'
        else:
            p.delivery_class = 'delivery-na'

    # Convert to list for filtering
    preschoolers_qs = list(preschoolers_qs)

    # ✅ FILTER BY NUTRITIONAL STATUS (if provided)
    filter_status = request.GET.get('status', 'All')
    if filter_status and filter_status != 'All':
        preschoolers_qs = [p for p in preschoolers_qs if p.nutritional_status == filter_status]

    # ✅ GLOBAL SEARCH - Search by preschooler name ONLY
    search_query = request.GET.get('search', '').strip()
    is_searching = False
    
    if search_query:
        is_searching = True
        search_lower = search_query.lower()
        preschoolers_qs = [
            p for p in preschoolers_qs 
            if search_lower in f"{p.first_name} {p.last_name}".lower()
        ]

    # ✅ Pagination after filtering and sorting
    paginator = Paginator(preschoolers_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    barangays = Barangay.objects.all()

    return render(request, 'HTML/registered_preschoolers.html', {
        'preschoolers': page_obj,
        'barangays': barangays,
        'filter_status': filter_status,
        'search_query': search_query,
        'is_searching': is_searching,
    })

def reportTemplate(request):
    if not request.user.is_authenticated:
        return redirect('login')
    # Check if this is a PDF generation request (from admin dashboard)
    if request.GET.get('generate_pdf') == 'true':
        # Get all barangays and their data for admin report
        barangays = Barangay.objects.all()
        
        # Overall summary across all barangays
        overall_summary = {
            'Severely Wasted': 0,
            'Wasted': 0,
            'Normal': 0,
            'Overweight': 0,
            'Obese': 0,
        }
        
        barangay_details = []
        total_preschoolers = 0
        
        for barangay in barangays:
            preschoolers = Preschooler.objects.filter(
                barangay=barangay,
                is_archived=False
            ).prefetch_related(
                Prefetch('bmi_set', queryset=BMI.objects.order_by('-date_recorded'), to_attr='bmi_records')
            )
            
            # Barangay-specific summary
            barangay_summary = {
                'severely_underweight': 0,
                'underweight': 0,
                'normal': 0,
                'overweight': 0,
                'obese': 0,
            }
            
            barangay_count = preschoolers.count()
            total_preschoolers += barangay_count
            
            for p in preschoolers:
                latest_bmi = p.bmi_records[0] if hasattr(p, 'bmi_records') and p.bmi_records else None
                
                if latest_bmi and latest_bmi.bmi_value:
                    bmi = latest_bmi.bmi_value
                    if bmi < 13:
                        barangay_summary['severely_wasted'] += 1
                        overall_summary['severely_wasted'] += 1
                    elif 13 <= bmi < 14.9:
                        barangay_summary['wasted'] += 1
                        overall_summary['Wasted'] += 1
                    elif 14.9 <= bmi <= 17.5:
                        barangay_summary['normal'] += 1
                        overall_summary['Normal'] += 1
                    elif 17.6 <= bmi <= 18.9:
                        barangay_summary['overweight'] += 1
                        overall_summary['Overweight'] += 1
                    else:
                        barangay_summary['obese'] += 1
                        overall_summary['Obese'] += 1
            
            # Calculate at-risk percentage (severely underweight + underweight + obese)
            at_risk_count = barangay_summary['severely_wasted'] + barangay_summary['underweight'] + barangay_summary['obese']
            at_risk_percentage = (at_risk_count / barangay_count * 100) if barangay_count > 0 else 0
            
            barangay_details.append({
                'name': barangay.name,
                'total_preschoolers': barangay_count,
                'severely_wasted': barangay_summary['severely_wasted'],
                'underweight': barangay_summary['underweight'],
                'normal': barangay_summary['normal'],
                'overweight': barangay_summary['overweight'],
                'obese': barangay_summary['obese'],
                'at_risk_percentage': round(at_risk_percentage, 1)
            })
        
        # Find highest count barangay
        highest_barangay = max(barangay_details, key=lambda x: x['total_preschoolers']) if barangay_details else None
        
        # Calculate total at-risk children
        total_at_risk = overall_summary['Severely Wasted'] + overall_summary['Underweight'] + overall_summary['Obese']
        
        # Get current account - handle both authenticated and anonymous users
        account = None
        if request.user.is_authenticated:
            try:
                account = Account.objects.get(email=request.user.email)
            except Account.DoesNotExist:
                # Create a default account info for display
                account = type('obj', (object,), {
                    'full_name': 'System Administrator',
                    'email': 'admin@system.local'
                })()
        else:
            # Create a default account info for anonymous users
            account = type('obj', (object,), {
                'full_name': 'System Administrator',
                'email': 'admin@system.local'
            })()
        
        # Render HTML for PDF
        html_string = render_to_string('HTML/reportTemplate.html', {
            'account': account,
            'barangay_details': barangay_details,
            'overall_summary': overall_summary,
            'total_barangays': barangays.count(),
            'total_preschoolers': total_preschoolers,
            'highest_barangay': highest_barangay,
            'total_at_risk': total_at_risk,
            'is_admin_report': True,
        })
        
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf = html.write_pdf()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="Overall-Barangay-Summary-Report.pd"'
        return response
    
    # Default behavior - just render the template for preview
    return render(request, 'HTML/reportTemplate.html')


def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        
        # Validate email
        if not email:
            messages.error(request, 'Email address is required.')
            return render(request, 'HTML/forgot_password.html')
        
        try:
            validate_email(email)
        except ValidationError:
            messages.error(request, 'Please enter a valid email address.')
            return render(request, 'HTML/forgot_password.html')
        
        # Check if user exists
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            messages.error(request, 'No account found with this email address.')
            return render(request, 'HTML/forgot_password.html')
        
        # Delete existing OTPs
        PasswordResetOTP.objects.filter(user=user, is_used=False).delete()
        
        # Create new OTP
        otp_instance = PasswordResetOTP.objects.create(user=user)
        
        # Compose email
        subject = '🔐 Password Reset OTP - PPMS Cluster 4'

        text_message = f"""
        Hello {user.first_name or user.username},

        You requested a password reset. Your OTP code is: {otp_instance.otp_code}

        This code will expire in 10 minutes.

        If you didn't request this, please ignore this email.
        """

        html_message = f"""
        <!DOCTYPE html>
        <html>
        <head>
          <meta charset="UTF-8">
          <style>
            body {{
              font-family: Arial, sans-serif;
              background-color: #f9f9f9;
              padding: 20px;
              color: #333;
            }}
            .container {{
              background-color: #fff;
              padding: 20px;
              border-radius: 10px;
              max-width: 600px;
              margin: auto;
              box-shadow: 0 4px 10px rgba(0,0,0,0.1);
            }}
            .header {{
              background-color: #007bff;
              padding: 10px 20px;
              border-radius: 10px 10px 0 0;
              color: white;
              text-align: center;
            }}
            .otp {{
              font-size: 28px;
              font-weight: bold;
              color: #007bff;
              text-align: center;
              margin: 30px 0;
            }}
            .footer {{
              font-size: 12px;
              text-align: center;
              color: #777;
              margin-top: 30px;
            }}
          </style>
        </head>
        <body>
          <div class="container">
            <div class="header">
              <h2>PPMS Cluster 4 – Password Reset</h2>
            </div>

            <p>Hello <strong>{user.first_name or user.username}</strong>,</p>

            <p>You requested to reset your password. Please use the following OTP:</p>

            <div class="otp">{otp_instance.otp_code}</div>

            <p>This OTP is valid for <strong>10 minutes</strong>.</p>

            <p>If you did not make this request, you can safely ignore this email.</p>

            <div class="footer">
              &copy; 2025 PPMS Cluster 4 Imus City
            </div>
          </div>
        </body>
        </html>
        """

        try:
            email_msg = EmailMultiAlternatives(
                subject=subject,
                body=text_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email]
            )
            email_msg.attach_alternative(html_message, "text/html")
            email_msg.send()

            messages.success(request, 'OTP sent to your email address.')
            return redirect('verify_otp', user_id=user.id)
        except Exception as e:
            print(f"[ERROR] Email send failed: {e}")
            messages.error(request, 'Failed to send email. Please try again.')

    return render(request, 'HTML/forgot_password.html')

@admin_required 
def admin_registered_parents(request):
    # Ensure user is authenticated and is admin
    user_email = request.session.get('email')
    user_role = request.session.get('user_role', '').lower()

    if user_role != 'admin':
        return render(request, 'unauthorized.html')

    # Fetch all parents
    parents_qs = Parent.objects.select_related('barangay').order_by('-created_at')

    # Handle search query
    search_query = request.GET.get('search', '').strip()
    if search_query:
        parents_qs = parents_qs.filter(
            Q(full_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(contact_number__icontains=search_query) |
            Q(barangay__name__icontains=search_query)
        )

    paginator = Paginator(parents_qs, 10)  # Show 20 parents per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'HTML/admin_registeredparents.html', {
        'parents': page_obj,
        'user_email': user_email,
        'user_role': user_role,
        'search_query': search_query,
    })

def verify_otp(request, user_id):
    user = get_object_or_404(User, id=user_id)

    # ✅ If resend is requested, generate a new OTP
    if request.method == 'GET' and request.GET.get('resend') == '1':
        # Mark any previous OTPs as used
        PasswordResetOTP.objects.filter(user=user, is_used=False).update(is_used=True)

        # Generate and save new OTP
        new_otp = get_random_string(length=6, allowed_chars='0123456789')
        PasswordResetOTP.objects.create(user=user, otp_code=new_otp)

        # (Optional) Send the OTP to user's email here
        # send_mail(...)

        messages.success(request, 'A new OTP has been sent to your email.')

    # Existing POST logic
    if request.method == 'POST':
        otp_code = request.POST.get('otp_code', '').strip()

        if not otp_code:
            messages.error(request, 'OTP code is required.')
            return render(request, 'HTML/verify_otp.html', {'user': user})

        if len(otp_code) != 6 or not otp_code.isdigit():
            messages.error(request, 'OTP must be exactly 6 digits.')
            return render(request, 'HTML/verify_otp.html', {'user': user})

        try:
            otp_instance = PasswordResetOTP.objects.get(
                user=user,
                otp_code=otp_code,
                is_used=False
            )

            if otp_instance.is_expired():
                messages.error(request, 'OTP has expired. Please request a new one.')
                return redirect('verify_otp', user_id=user.id)

            otp_instance.is_used = True
            otp_instance.save()

            messages.success(request, 'OTP verified successfully.')
            return redirect('reset_password', user_id=user.id)

        except PasswordResetOTP.DoesNotExist:
            messages.error(request, 'Invalid OTP. Please try again.')

    return render(request, 'HTML/verify_otp.html', {'user': user})

def reset_password(request, user_id):
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Check if there's a recent used OTP for this user
        # Comment out OTP check if PasswordResetOTP model is not imported yet
        try:
            recent_otp = PasswordResetOTP.objects.filter(
                user=user,
                is_used=True,
                created_at__gte=timezone.now() - timezone.timedelta(minutes=15)
            ).first()
            
            if not recent_otp:
                messages.error(request, 'Session expired. Please start the process again.')
                return redirect('forgot_password')
        except NameError:
            # If PasswordResetOTP is not imported, skip this check for now
            pass
        
        if request.method == 'POST':
            password1 = request.POST.get('password1', '')
            password2 = request.POST.get('password2', '')
            
            # Validate passwords
            if not password1 or not password2:
                messages.error(request, 'Both password fields are required.')
                return render(request, 'HTML/reset_password.html', {'user': user})
            
            if password1 != password2:
                messages.error(request, 'Passwords do not match.')
                return render(request, 'HTML/reset_password.html', {'user': user})
            
            if len(password1) < 8:
                messages.error(request, 'Password must be at least 8 characters long.')
                return render(request, 'HTML/reset_password.html', {'user': user})
            
            # Selective password validation (WITHOUT CommonPasswordValidator)
            try:
                validators = [
                    MinimumLengthValidator(min_length=8),
                    NumericPasswordValidator(),
                    UserAttributeSimilarityValidator(),
                ]
                for validator in validators:
                    validator.validate(password1, user)
            except ValidationError as e:
                for error in e.messages:
                    messages.error(request, error)
                return render(request, 'HTML/reset_password.html', {'user': user})
            
            # Set new password
            user.set_password(password1)
            user.save()
            
            # Optional: Delete the used OTP to prevent reuse
            try:
                if 'recent_otp' in locals() and recent_otp:
                    recent_otp.delete()
            except:
                pass
            
            messages.success(request, 'Password reset successfully. You can now login with your new password.')
            return redirect('login')
        
        return render(request, 'HTML/reset_password.html', {'user': user})
    
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        return redirect('forgot_password')
    
def remove_bns(request, account_id):
    if request.method == 'POST':
        try:
            bns = get_object_or_404(Account, pk=account_id)
            
            # Safety check for BNS role
            bns_role_keywords = ['bns', 'nutritional', 'scholar']
            if not any(keyword in bns.user_role.lower() for keyword in bns_role_keywords):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': 'Not a BNS worker'})
                messages.error(request, f"Cannot remove {bns.full_name}: Not a BNS worker.")
                return redirect('healthcare_workers')
            
            name = bns.full_name
            email = bns.email
            
            # Get current date
            current_date = datetime.now().strftime('%B %d, %Y')
            
            # Prepare email
            subject = 'PPMS Cluster 4 – Account Removal Notification'
            
            html_message = f"""
            <!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>PPMS Account Removal Notification</title>
                <style>
                    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
                    
                    body {{
                        font-family: 'Inter', Arial, sans-serif;
                        background-color: #f9fafb;
                        padding: 40px 20px;
                        color: #334155;
                        line-height: 1.6;
                        margin: 0;
                    }}
                    
                    .container {{
                        max-width: 560px;
                        margin: 0 auto;
                        background: white;
                        border-radius: 12px;
                        overflow: hidden;
                        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
                    }}
                    
                    .header {{
                        padding: 24px 32px;
                        text-align: center;
                        color: #111827;
                        border-bottom: 4px solid #dc3545;
                    }}
                    .header h1 {{
                        font-size: 24px;
                        font-weight: 600;
                        margin: 0;
                        color: #111827;
                    }}
                    .header p {{
                        font-size: 16px;
                        margin: 4px 0 0 0;
                        color: #6b7280;
                    }}
                    
                    .content {{
                        padding: 32px;
                    }}
                    
                    .greeting {{
                        font-size: 18px;
                        margin-bottom: 24px;
                        color: #1e293b;
                    }}
                    
                    .message {{
                        font-size: 16px;
                        margin-bottom: 24px;
                        color: #64748b;
                    }}
                    
                    .details {{
                        background: #fef2f2;
                        border: 1px solid #ef4444;
                        padding: 24px;
                        border-radius: 6px;
                        margin: 24px 0;
                    }}
                    
                    .details h4 {{
                        margin-bottom: 16px;
                        font-size: 16px;
                        font-weight: 600;
                        color: #7f1d1d;
                    }}
                    
                    .detail-item {{
                        margin-bottom: 12px;
                        line-height: 1.6;
                    }}
                    
                    .detail-item:last-child {{
                        margin-bottom: 0;
                    }}
                    
                    .detail-label {{
                        font-weight: bold;
                        color: #7f1d1d;
                        display: inline;
                    }}
                    
                    .detail-value {{
                        color: #7f1d1d;
                        display: inline;
                        margin-left: 4px;
                    }}
                    
                    .notice {{
                        background: #fef3c7;
                        border: 1px solid #f59e0b;
                        border-radius: 8px;
                        padding: 20px;
                        margin-bottom: 32px;
                        text-align: center;
                    }}
                    
                    .notice h3 {{
                        font-size: 18px;
                        font-weight: 600;
                        color: #92400e;
                        margin-bottom: 8px;
                    }}
                    
                    .notice p {{
                        font-size: 14px;
                        color: #92400e;
                        margin: 0;
                    }}
                    
                    .footer {{
                        background: #f1f5f9;
                        padding: 32px;
                        text-align: center;
                        border-top: 1px solid #e2e8f0;
                    }}
                    
                    .footer h3 {{
                        font-size: 18px;
                        font-weight: 600;
                        color: #1e293b;
                        margin-bottom: 8px;
                    }}
                    
                    .footer p {{
                        font-size: 14px;
                        color: #64748b;
                        margin-bottom: 4px;
                    }}
                    
                    .footer-divider {{
                        margin: 24px 0;
                        height: 1px;
                        background: #e2e8f0;
                    }}
                    
                    .footer-small {{
                        font-size: 12px;
                        color: #94a3b8;
                    }}
                    
                    @media (max-width: 600px) {{
                        .content {{ 
                            padding: 28px 20px; 
                        }}
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>Account Removed</h1>
                        <p>PPMS Cluster 4 - Imus City Healthcare Management</p>
                    </div>
                    
                    <div class="content">
                        <div class="greeting">
                            Hello <strong>{name}</strong>,
                        </div>
                        
                        <div class="message">
                            We would like to inform you that your BNS account has been removed from the PPMS Cluster 4 system.
                        </div>
                        
                        <div class="notice">
                            <h3>Important Notice</h3>
                            <p>If you believe this was a mistake or have any questions, please contact the system administrator.</p>
                        </div>
                        
                        <div class="details">
                            <h4>Account Information</h4>
                            <div class="detail-item">
                                <span class="detail-label">Full Name:</span>
                                <span class="detail-value">{name}</span>
                            </div>
                            <div class="detail-item">
                                <span class="detail-label">Email:</span>
                                <span class="detail-value">{email}</span>
                            </div>
                            <div class="detail-item">
                                <span class="detail-label">Date Removed:</span>
                                <span class="detail-value">{current_date}</span>
                            </div>
                        </div>
                    </div>
                    
                    <div class="footer">
                        <h3>PPMS Cluster 4</h3>
                        <p>Imus City Healthcare Management</p>
                        <div class="footer-divider"></div>
                        <p class="footer-small">
                            This is an automated message. Please do not reply.<br>
                            © 2025 PPMS Cluster 4. All rights reserved.
                        </p>
                    </div>
                </div>
            </body>
            </html>
            """

            plain_message = f"""
PPMS Account Removal Notification

Hello {name},

We would like to inform you that your BNS account has been removed from the PPMS Cluster 4 system.

IMPORTANT NOTICE:
If you believe this was a mistake or have any questions, please contact the system administrator.

Account Information:
- Full Name: {name}
- Email: {email}
- Date Removed: {current_date}

PPMS Cluster 4
Imus City Healthcare Management

This is an automated message. Please do not reply.
© 2025 PPMS Cluster 4. All rights reserved.
            """

            # ========== Background email sending ==========
            # Capture all data before threading
            email_subject = subject
            email_plain = plain_message
            email_html = html_message
            recipient_email = email
            from_email = settings.DEFAULT_FROM_EMAIL

            def send_removal_email():
                try:
                    send_mail(
                        subject=email_subject,
                        message=email_plain,
                        from_email=from_email,
                        recipient_list=[recipient_email],
                        html_message=email_html,
                        fail_silently=False
                    )
                    print(f"[DEBUG] ✅ Removal email sent successfully to {recipient_email}")
                except Exception as e:
                    print(f"[EMAIL ERROR]: {e}")

            # Start email sending in background thread
            import threading
            threading.Thread(target=send_removal_email).start()

            # Delete account
            bns.delete()
            
            # Return JSON for AJAX requests (happens immediately)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'name': name, 'type': 'BNS'})
            
            # Fallback for regular form submission
            messages.success(request, f"{name} has been successfully removed and notification email is being sent.")
            
        except Account.DoesNotExist:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Worker no longer exists'})
            messages.error(request, "The worker you're trying to remove no longer exists.")
        except Exception as e:
            print(f"[ERROR] Failed to remove BNS: {e}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'An error occurred while removing the worker'})
            messages.error(request, "An error occurred while removing the BNS.")

    return redirect('healthcare_workers')

def remove_bhw(request, account_id):
    if request.method == 'POST':
        try:
            bhw = get_object_or_404(Account, pk=account_id)
            
            # Safety check for BHW role
            bhw_role_keywords = ['bhw', 'healthworker', 'health worker']
            if not any(keyword in bhw.user_role.lower() for keyword in bhw_role_keywords):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': 'Not a BHW worker'})
                messages.error(request, f"Cannot remove {bhw.full_name}: Not a BHW worker.")
                return redirect('healthcare_workers')
            
            name = bhw.full_name
            email = bhw.email
            
            # Get current date
            current_date = datetime.now().strftime('%B %d, %Y')
            
            # Prepare email
            subject = 'PPMS Cluster 4 – Account Removal Notification'
            
            html_message = f"""
            <!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>PPMS Account Removal Notification</title>
                <style>
                    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
                    
                    body {{
                        font-family: 'Inter', Arial, sans-serif;
                        background-color: #f9fafb;
                        padding: 40px 20px;
                        color: #334155;
                        line-height: 1.6;
                        margin: 0;
                    }}
                    
                    .container {{
                        max-width: 560px;
                        margin: 0 auto;
                        background: white;
                        border-radius: 12px;
                        overflow: hidden;
                        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
                    }}
                    
                    .header {{
                        padding: 24px 32px;
                        text-align: center;
                        color: #111827;
                        border-bottom: 4px solid #dc3545;
                    }}
                    .header h1 {{
                        font-size: 24px;
                        font-weight: 600;
                        margin: 0;
                        color: #111827;
                    }}
                    .header p {{
                        font-size: 16px;
                        margin: 4px 0 0 0;
                        color: #6b7280;
                    }}
                    
                    .content {{
                        padding: 32px;
                    }}
                    
                    .greeting {{
                        font-size: 18px;
                        margin-bottom: 24px;
                        color: #1e293b;
                    }}
                    
                    .message {{
                        font-size: 16px;
                        margin-bottom: 24px;
                        color: #64748b;
                    }}
                    
                    .details {{
                        background: #fef2f2;
                        border: 1px solid #ef4444;
                        padding: 24px;
                        border-radius: 6px;
                        margin: 24px 0;
                    }}
                    
                    .details h4 {{
                        margin-bottom: 16px;
                        font-size: 16px;
                        font-weight: 600;
                        color: #7f1d1d;
                    }}
                    
                    .detail-item {{
                        margin-bottom: 12px;
                        line-height: 1.6;
                    }}
                    
                    .detail-item:last-child {{
                        margin-bottom: 0;
                    }}
                    
                    .detail-label {{
                        font-weight: bold;
                        color: #7f1d1d;
                        display: inline;
                    }}
                    
                    .detail-value {{
                        color: #7f1d1d;
                        display: inline;
                        margin-left: 4px;
                    }}
                    
                    .notice {{
                        background: #fef3c7;
                        border: 1px solid #f59e0b;
                        border-radius: 8px;
                        padding: 20px;
                        margin-bottom: 32px;
                        text-align: center;
                    }}
                    
                    .notice h3 {{
                        font-size: 18px;
                        font-weight: 600;
                        color: #92400e;
                        margin-bottom: 8px;
                    }}
                    
                    .notice p {{
                        font-size: 14px;
                        color: #92400e;
                        margin: 0;
                    }}
                    
                    .footer {{
                        background: #f1f5f9;
                        padding: 32px;
                        text-align: center;
                        border-top: 1px solid #e2e8f0;
                    }}
                    
                    .footer h3 {{
                        font-size: 18px;
                        font-weight: 600;
                        color: #1e293b;
                        margin-bottom: 8px;
                    }}
                    
                    .footer p {{
                        font-size: 14px;
                        color: #64748b;
                        margin-bottom: 4px;
                    }}
                    
                    .footer-divider {{
                        margin: 24px 0;
                        height: 1px;
                        background: #e2e8f0;
                    }}
                    
                    .footer-small {{
                        font-size: 12px;
                        color: #94a3b8;
                    }}
                    
                    @media (max-width: 600px) {{
                        .content {{ 
                            padding: 28px 20px; 
                        }}
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>Account Removed</h1>
                        <p>PPMS Cluster 4 - Imus City Healthcare Management</p>
                    </div>
                    
                    <div class="content">
                        <div class="greeting">
                            Hello <strong>{name}</strong>,
                        </div>
                        
                        <div class="message">
                            We would like to inform you that your BHW account has been removed from the PPMS Cluster 4 system.
                        </div>
                        
                        <div class="notice">
                            <h3>Important Notice</h3>
                            <p>If you believe this was a mistake or have any questions, please contact the system administrator.</p>
                        </div>
                        
                        <div class="details">
                            <h4>Account Information</h4>
                            <div class="detail-item">
                                <span class="detail-label">Full Name:</span>
                                <span class="detail-value">{name}</span>
                            </div>
                            <div class="detail-item">
                                <span class="detail-label">Email:</span>
                                <span class="detail-value">{email}</span>
                            </div>
                            <div class="detail-item">
                                <span class="detail-label">Date Removed:</span>
                                <span class="detail-value">{current_date}</span>
                            </div>
                        </div>
                    </div>
                    
                    <div class="footer">
                        <h3>PPMS Cluster 4</h3>
                        <p>Imus City Healthcare Management</p>
                        <div class="footer-divider"></div>
                        <p class="footer-small">
                            This is an automated message. Please do not reply.<br>
                            © 2025 PPMS Cluster 4. All rights reserved.
                        </p>
                    </div>
                </div>
            </body>
            </html>
            """

            plain_message = f"""
PPMS Account Removal Notification

Hello {name},

We would like to inform you that your BHW account has been removed from the PPMS Cluster 4 system.

IMPORTANT NOTICE:
If you believe this was a mistake or have any questions, please contact the system administrator.

Account Information:
- Full Name: {name}
- Email: {email}
- Date Removed: {current_date}

PPMS Cluster 4
Imus City Healthcare Management

This is an automated message. Please do not reply.
© 2025 PPMS Cluster 4. All rights reserved.
            """

            # ========== Background email sending ==========
            # Capture all data before threading
            email_subject = subject
            email_plain = plain_message
            email_html = html_message
            recipient_email = email
            from_email = settings.DEFAULT_FROM_EMAIL

            def send_removal_email():
                try:
                    send_mail(
                        subject=email_subject,
                        message=email_plain,
                        from_email=from_email,
                        recipient_list=[recipient_email],
                        html_message=email_html,
                        fail_silently=False
                    )
                    print(f"[DEBUG] ✅ Removal email sent successfully to {recipient_email}")
                except Exception as e:
                    print(f"[EMAIL ERROR]: {e}")

            # Start email sending in background thread
            import threading
            threading.Thread(target=send_removal_email).start()

            # Delete account
            bhw.delete()
            
            # Return JSON for AJAX requests (happens immediately)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'name': name, 'type': 'BHW'})
            
            # Fallback for regular form submission
            messages.success(request, f"{name} has been successfully removed and notification email is being sent.")
            
        except Account.DoesNotExist:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Worker no longer exists'})
            messages.error(request, "The worker you're trying to remove no longer exists.")
        except Exception as e:
            print(f"[ERROR] Failed to remove BHW: {e}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'An error occurred while removing the worker'})
            messages.error(request, "An error occurred while removing the BHW.")

    return redirect('healthcare_workers')

def registered_midwife(request):

    if not request.user.is_authenticated:
        return redirect('login')
    midwife_list = Account.objects.filter(user_role='midwife', is_validated=True)

    for midwife in midwife_list:
        # Assuming you have a Midwife model similar to BHW
        midwife.midwife_data = Midwife.objects.filter(email=midwife.email).first()

        if midwife.last_activity:
            if timezone.now() - midwife.last_activity <= timedelta(minutes=1):
                midwife.last_activity_display = "🟢 Online"
            else:
                time_diff = timesince(midwife.last_activity, timezone.now())
                midwife.last_activity_display = f"{time_diff} ago"
        else:
            midwife.last_activity_display = "No activity"

    paginator = Paginator(midwife_list, 10)  # 10 midwives per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'HTML/registered_midwife.html', {'midwives': page_obj})

def remove_midwife(request, account_id):
    if request.method == 'POST':
        try:
            midwife = get_object_or_404(Account, pk=account_id)
            
            # Safety check for Midwife role
            if 'midwife' not in midwife.user_role.lower():
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': 'Not a Midwife worker'})
                messages.error(request, f"Cannot remove {midwife.full_name}: Not a Midwife worker.")
                return redirect('healthcare_workers')
            
            name = midwife.full_name
            email = midwife.email
            
            # Get current date
            current_date = datetime.now().strftime('%B %d, %Y')
            
            # Prepare email
            subject = 'PPMS Cluster 4 – Account Removal Notification'
            
            html_message = f"""
            <!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>PPMS Account Removal Notification</title>
                <style>
                    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
                    
                    body {{
                        font-family: 'Inter', Arial, sans-serif;
                        background-color: #f9fafb;
                        padding: 40px 20px;
                        color: #334155;
                        line-height: 1.6;
                        margin: 0;
                    }}
                    
                    .container {{
                        max-width: 560px;
                        margin: 0 auto;
                        background: white;
                        border-radius: 12px;
                        overflow: hidden;
                        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
                    }}
                    
                    .header {{
                        padding: 24px 32px;
                        text-align: center;
                        color: #111827;
                        border-bottom: 4px solid #dc3545;
                    }}
                    .header h1 {{
                        font-size: 24px;
                        font-weight: 600;
                        margin: 0;
                        color: #111827;
                    }}
                    .header p {{
                        font-size: 16px;
                        margin: 4px 0 0 0;
                        color: #6b7280;
                    }}
                    
                    .content {{
                        padding: 32px;
                    }}
                    
                    .greeting {{
                        font-size: 18px;
                        margin-bottom: 24px;
                        color: #1e293b;
                    }}
                    
                    .message {{
                        font-size: 16px;
                        margin-bottom: 24px;
                        color: #64748b;
                    }}
                    
                    .details {{
                        background: #fef2f2;
                        border: 1px solid #ef4444;
                        padding: 24px;
                        border-radius: 6px;
                        margin: 24px 0;
                    }}
                    
                    .details h4 {{
                        margin-bottom: 16px;
                        font-size: 16px;
                        font-weight: 600;
                        color: #7f1d1d;
                    }}
                    
                    .detail-item {{
                        margin-bottom: 12px;
                        line-height: 1.6;
                    }}
                    
                    .detail-item:last-child {{
                        margin-bottom: 0;
                    }}
                    
                    .detail-label {{
                        font-weight: bold;
                        color: #7f1d1d;
                        display: inline;
                    }}
                    
                    .detail-value {{
                        color: #7f1d1d;
                        display: inline;
                        margin-left: 4px;
                    }}
                    
                    .notice {{
                        background: #fef3c7;
                        border: 1px solid #f59e0b;
                        border-radius: 8px;
                        padding: 20px;
                        margin-bottom: 32px;
                        text-align: center;
                    }}
                    
                    .notice h3 {{
                        font-size: 18px;
                        font-weight: 600;
                        color: #92400e;
                        margin-bottom: 8px;
                    }}
                    
                    .notice p {{
                        font-size: 14px;
                        color: #92400e;
                        margin: 0;
                    }}
                    
                    .footer {{
                        background: #f1f5f9;
                        padding: 32px;
                        text-align: center;
                        border-top: 1px solid #e2e8f0;
                    }}
                    
                    .footer h3 {{
                        font-size: 18px;
                        font-weight: 600;
                        color: #1e293b;
                        margin-bottom: 8px;
                    }}
                    
                    .footer p {{
                        font-size: 14px;
                        color: #64748b;
                        margin-bottom: 4px;
                    }}
                    
                    .footer-divider {{
                        margin: 24px 0;
                        height: 1px;
                        background: #e2e8f0;
                    }}
                    
                    .footer-small {{
                        font-size: 12px;
                        color: #94a3b8;
                    }}
                    
                    @media (max-width: 600px) {{
                        .content {{ 
                            padding: 28px 20px; 
                        }}
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>Account Removed</h1>
                        <p>PPMS Cluster 4 - Imus City Healthcare Management</p>
                    </div>
                    
                    <div class="content">
                        <div class="greeting">
                            Hello <strong>{name}</strong>,
                        </div>
                        
                        <div class="message">
                            We would like to inform you that your Midwife account has been removed from the PPMS Cluster 4 system.
                        </div>
                        
                        <div class="notice">
                            <h3>Important Notice</h3>
                            <p>If you believe this was a mistake or have any questions, please contact the system administrator.</p>
                        </div>
                        
                        <div class="details">
                            <h4>Account Information</h4>
                            <div class="detail-item">
                                <span class="detail-label">Full Name:</span>
                                <span class="detail-value">{name}</span>
                            </div>
                            <div class="detail-item">
                                <span class="detail-label">Email:</span>
                                <span class="detail-value">{email}</span>
                            </div>
                            <div class="detail-item">
                                <span class="detail-label">Date Removed:</span>
                                <span class="detail-value">{current_date}</span>
                            </div>
                        </div>
                    </div>
                    
                    <div class="footer">
                        <h3>PPMS Cluster 4</h3>
                        <p>Imus City Healthcare Management</p>
                        <div class="footer-divider"></div>
                        <p class="footer-small">
                            This is an automated message. Please do not reply.<br>
                            © 2025 PPMS Cluster 4. All rights reserved.
                        </p>
                    </div>
                </div>
            </body>
            </html>
            """

            plain_message = f"""
PPMS Account Removal Notification

Hello {name},

We would like to inform you that your Midwife account has been removed from the PPMS Cluster 4 system.

IMPORTANT NOTICE:
If you believe this was a mistake or have any questions, please contact the system administrator.

Account Information:
- Full Name: {name}
- Email: {email}
- Date Removed: {current_date}

PPMS Cluster 4
Imus City Healthcare Management

This is an automated message. Please do not reply.
© 2025 PPMS Cluster 4. All rights reserved.
            """

            # ========== Background email sending ==========
            # Capture all data before threading
            email_subject = subject
            email_plain = plain_message
            email_html = html_message
            recipient_email = email
            from_email = settings.DEFAULT_FROM_EMAIL

            def send_removal_email():
                try:
                    send_mail(
                        subject=email_subject,
                        message=email_plain,
                        from_email=from_email,
                        recipient_list=[recipient_email],
                        html_message=email_html,
                        fail_silently=False
                    )
                    print(f"[DEBUG] ✅ Removal email sent successfully to {recipient_email}")
                except Exception as e:
                    print(f"[EMAIL ERROR]: {e}")

            # Start email sending in background thread
            import threading
            threading.Thread(target=send_removal_email).start()

            # Delete account
            midwife.delete()
            
            # Return JSON for AJAX requests (happens immediately)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'name': name, 'type': 'Midwife'})
            
            # Fallback for regular form submission
            messages.success(request, f"{name} has been successfully removed and notification email is being sent.")
            
        except Account.DoesNotExist:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Worker no longer exists'})
            messages.error(request, "The worker you're trying to remove no longer exists.")
        except Exception as e:
            print(f"[ERROR] Failed to remove Midwife: {e}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'An error occurred while removing the worker'})
            messages.error(request, "An error occurred while removing the midwife.")

    return redirect('healthcare_workers')




    
@admin_required
def validate(request):
    """Display BHW, BNS, Midwife, and Nurse accounts - focusing on pending validation by default"""
    # Get filter parameters
    role_filter = request.GET.get('role', 'all')
    status_filter = request.GET.get('status', 'pending')  # Default to pending only
    
    # Base queryset - BHW, BNS, Midwife, and Nurse accounts
    accounts = Account.objects.filter(
        Q(user_role__iexact='BHW') | 
        Q(user_role__iexact='healthworker') |
        Q(user_role__iexact='BNS') | 
        Q(user_role__iexact='bns') | 
        Q(user_role__iexact='Barangay Nutritional Scholar') |
        Q(user_role__iexact='Midwife') |
        Q(user_role__iexact='Nurse')  # Added Nurse here
    )
    
    # Apply status filter
    if status_filter == 'pending':
        accounts = accounts.filter(is_validated=False, is_rejected=False)
    elif status_filter == 'validated':
        accounts = accounts.filter(is_validated=True)
    elif status_filter == 'rejected':
        accounts = accounts.filter(is_rejected=True)
    # If status_filter == 'all', show all accounts
    
    # Apply role filter if specified
    if role_filter and role_filter != 'all':
        if role_filter.lower() == 'bns':
            accounts = accounts.filter(
                Q(user_role__iexact='bns') | 
                Q(user_role__iexact='BNS') |
                Q(user_role__iexact='Barangay Nutritional Scholar')
            )
        elif role_filter.lower() == 'healthworker':
            accounts = accounts.filter(
                Q(user_role__iexact='healthworker') |
                Q(user_role__iexact='BHW')
            )
        elif role_filter.lower() == 'midwife':
            accounts = accounts.filter(user_role__iexact='Midwife')
        elif role_filter.lower() == 'nurse':  # Added nurse filter
            accounts = accounts.filter(user_role__iexact='Nurse')
        else:
            accounts = accounts.filter(user_role=role_filter)
    
    accounts = accounts.order_by('-created_at')
    
    # Paginate results
    paginator = Paginator(accounts, 10)  # Show 10 accounts per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Count statistics for display
    base_filter = (
        Q(user_role__iexact='BHW') | Q(user_role__iexact='healthworker') | 
        Q(user_role__iexact='BNS') | Q(user_role__iexact='bns') | 
        Q(user_role__iexact='Barangay Nutritional Scholar') | 
        Q(user_role__iexact='Midwife') |
        Q(user_role__iexact='Nurse')  # Added Nurse here
    )
    
    total_pending = Account.objects.filter(
        base_filter,
        is_validated=False,
        is_rejected=False
    ).count()
    
    bhw_pending = Account.objects.filter(
        Q(user_role__iexact='healthworker') | Q(user_role__iexact='BHW'),
        is_validated=False,
        is_rejected=False
    ).count()
    
    bns_pending = Account.objects.filter(
        Q(user_role__iexact='bns') | Q(user_role__iexact='BNS') | Q(user_role__iexact='Barangay Nutritional Scholar'),
        is_validated=False,
        is_rejected=False
    ).count()
    
    midwife_pending = Account.objects.filter(
        user_role__iexact='Midwife',
        is_validated=False,
        is_rejected=False
    ).count()
    
    nurse_pending = Account.objects.filter(  # Added nurse pending count
        user_role__iexact='Nurse',
        is_validated=False,
        is_rejected=False
    ).count()
    
    total_validated = Account.objects.filter(
        base_filter,
        is_validated=True
    ).count()
    
    total_rejected = Account.objects.filter(
        base_filter,
        is_rejected=True
    ).count()
    
    context = {
        'accounts': page_obj,
        'page_obj': page_obj,
        'current_filter': role_filter,
        'current_status': status_filter,
        'total_pending': total_pending,
        'bhw_pending': bhw_pending,
        'bns_pending': bns_pending,
        'midwife_pending': midwife_pending,
        'nurse_pending': nurse_pending,  # Pass nurse count to template
        'total_validated': total_validated,
        'total_rejected': total_rejected,
    }
    
    return render(request, 'HTML/validate.html', context)

@csrf_exempt
def validate_account(request, account_id):
    if request.method == 'POST':
        account = get_object_or_404(Account, pk=account_id)
        account.is_validated = True
        account.is_rejected = False
        account.save()

        # ========== Background email sending ==========
        if account.email:
            from datetime import datetime
            
            # Capture all necessary data before threading
            full_name = account.full_name
            user_role = account.user_role
            barangay_name = account.barangay.name if account.barangay else "N/A"
            email = account.email
            current_date = datetime.now().strftime('%B %d, %Y')

            def send_validation_email():
                try:
                    subject = 'Account Validated - PPMS Cluster 4'

                    html_message = f"""
                    <!DOCTYPE html>
                    <html lang="en">
                    <head>
                        <meta charset="UTF-8">
                        <style>
                            body {{
                                font-family: Arial, sans-serif;
                                background-color: #f9f9f9;
                                padding: 20px;
                                color: #333;
                            }}
                            .container {{
                                background-color: #fff;
                                padding: 0;
                                border-radius: 10px;
                                max-width: 600px;
                                margin: auto;
                                box-shadow: 0 4px 10px rgba(0,0,0,0.1);
                                overflow: hidden;
                            }}
                            .header {{
                                padding: 20px;
                                text-align: center;
                                font-size: 22px;
                                font-weight: bold;
                                color: #111827;
                            }}
                            .divider {{
                                height: 4px;
                                background-color: #198754; /* ✅ Green divider */
                            }}
                            .content {{
                                padding: 32px;
                            }}
                            .content p {{
                                margin-bottom: 16px;
                                font-size: 15px;
                            }}
                            .details {{
                                background: #ecfdf5;
                                border: 1px solid #10b981;
                                padding: 16px;
                                border-radius: 6px;
                                margin: 24px 0;
                            }}
                            .details h4 {{
                                margin-bottom: 12px;
                                font-size: 16px;
                                font-weight: 600;
                                color: #065f46;
                            }}
                            .footer {{
                                text-align: center;
                                font-size: 12px;
                                color: #777;
                                margin-top: 30px;
                                border-top: 1px solid #e2e8f0;
                                padding: 12px;
                            }}
                        </style>
                    </head>
                    <body>
                        <div class="container">
                            <div class="header">
                                Account Validated
                            </div>
                            <div class="divider"></div>
                            <div class="content">
                                <p>Hello <strong>{full_name}</strong>,</p>
                                <p>Your PPMS account has been successfully 
                                <strong>validated</strong>. You can now log in using your registered email address.</p>

                                <div class="details">
                                    <h4>Account Details</h4>
                                    <p><strong>Role:</strong> {user_role}</p>
                                    <p><strong>Barangay:</strong> {barangay_name}</p>
                                    <p><strong>Email:</strong> {email}</p>
                                    <p><strong>Date Validated:</strong> {current_date}</p>
                                </div>

                                <p>Thank you for being part of the Preschooler Profiling and Monitoring System (PPMS).</p>
                            </div>
                            <div class="footer">
                                &copy; 2025 PPMS Cluster 4 - Imus City Healthcare Management<br>
                                This is an automated message. Please do not reply.
                            </div>
                        </div>
                    </body>
                    </html>
                    """

                    plain_message = f"""
Hello {full_name},

Your PPMS account has been successfully validated. 
You can now log in using your registered email address.

Account Details:
- Role: {user_role}
- Barangay: {barangay_name}
- Email: {email}
- Date Validated: {current_date}

Thank you for being part of the Preschooler Profiling and Monitoring System (PPMS).
                    """

                    send_mail(
                        subject,
                        plain_message,
                        settings.DEFAULT_FROM_EMAIL,
                        [email],
                        html_message=html_message,
                        fail_silently=False,
                    )
                    print(f"[DEBUG] ✅ Validation email sent successfully to {email}")
                except Exception as e:
                    print(f"[EMAIL ERROR]: {e}")

            # Start email sending in background thread
            import threading
            threading.Thread(target=send_validation_email).start()

        # Success message and redirect (happens immediately without waiting for email)
        messages.success(
            request,
            f"{account.full_name} ({account.user_role}) has been validated and will be notified via email."
        )
        return redirect('validate')
    
@csrf_exempt
def reject_account(request, account_id):
    account = get_object_or_404(Account, pk=account_id)

    account.is_validated = False
    account.is_rejected = True
    account.save()

    if account.email:
        from datetime import datetime
        current_date = datetime.now().strftime('%B %d, %Y')

        subject = 'PPMS Account Rejected'

        plain_message = f"""
Hello {account.full_name},

We regret to inform you that your registration to the PPMS Cluster 4 Imus City platform has been rejected.

If you believe this was a mistake, please contact the system administrator.

Date Rejected: {current_date}

Thank you,
PPMS Admin
        """

        html_message = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    background-color: #f9fafb;
                    padding: 20px;
                    color: #334155;
                    line-height: 1.6;
                }}
                .container {{
                    max-width: 600px;
                    margin: 0 auto;
                    background: #ffffff;
                    border-radius: 8px;
                    padding: 0;
                    box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                    overflow: hidden;
                }}
                .header {{
                    padding: 20px;
                    text-align: center;
                    font-size: 22px;
                    font-weight: bold;
                    color: #111827;
                }}
                .divider {{
                    height: 4px;
                    background-color: #dc3545; /* 🔴 Red divider for rejected */
                }}
                .content {{
                    padding: 32px;
                }}
                .content p {{
                    margin-bottom: 16px;
                    font-size: 16px;
                }}
                .details {{
                    background: #fef2f2;
                    border: 1px solid #f87171;
                    padding: 16px;
                    border-radius: 6px;
                    margin: 24px 0;
                }}
                .details h4 {{
                    margin-bottom: 12px;
                    font-size: 16px;
                    font-weight: 600;
                    color: #991b1b;
                }}
                .footer {{
                    text-align: center;
                    font-size: 13px;
                    color: #94a3b8;
                    margin-top: 32px;
                    border-top: 1px solid #e2e8f0;
                    padding: 16px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    Account Rejected
                </div>
                <div class="divider"></div>
                <div class="content">
                    <p>Hello <strong>{account.full_name}</strong>,</p>
                    <p>We regret to inform you that your registration to the PPMS Cluster 4 Imus City platform has been 
                    <strong>rejected</strong>.</p>

                    <div class="details">
                        <h4>Rejection Details</h4>
                        <p><strong>Email:</strong> {account.email}</p>
                        <p><strong>Date Rejected:</strong> {current_date}</p>
                    </div>

                    <p>If you believe this was a mistake, please contact the system administrator.</p>
                </div>
                <div class="footer">
                    © 2025 PPMS Cluster 4 - Imus City Healthcare Management<br>
                    This is an automated message. Please do not reply.
                </div>
            </div>
        </body>
        </html>
        """

        try:
            send_mail(
                subject,
                plain_message,
                settings.DEFAULT_FROM_EMAIL,
                [account.email],
                html_message=html_message,
                fail_silently=True,
            )
        except Exception as e:
            print("[EMAIL ERROR]", e)

    messages.success(request, f"{account.full_name} has been rejected and notified.")
    return redirect('validate')

def registered_nurse(request):
    if not request.user.is_authenticated:
        return redirect('login')
    nurse_list = Account.objects.filter(user_role='nurse', is_validated=True)
    for nurse in nurse_list:
        # Assuming you have a Nurse model similar to Midwife
        nurse.nurse_data = Nurse.objects.filter(email=nurse.email).first()
        if nurse.last_activity:
            if timezone.now() - nurse.last_activity <= timedelta(minutes=1):
                nurse.last_activity_display = "🟢 Online"
            else:
                time_diff = timesince(nurse.last_activity, timezone.now())
                nurse.last_activity_display = "{time_diff} ago"
        else:
            nurse.last_activity_display = "No activity"
    paginator = Paginator(nurse_list, 10)  # 10 nurses per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'HTML/registered_nurse.html', {'nurses': page_obj})


def remove_nurse(request, account_id):
    if request.method == 'POST':
        try:
            nurse = get_object_or_404(Account, pk=account_id)
            
            # Safety check for Nurse role
            if 'nurse' not in nurse.user_role.lower():
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': 'Not a Nurse worker'})
                messages.error(request, f"Cannot remove {nurse.full_name}: Not a Nurse worker.")
                return redirect('healthcare_workers')
            
            name = nurse.full_name
            email = nurse.email
            
            # Get current date
            current_date = datetime.now().strftime('%B %d, %Y')
            
            # Prepare email
            subject = 'PPMS Cluster 4 – Account Removal Notification'
            
            html_message = f"""
            <!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>PPMS Account Removal Notification</title>
                <style>
                    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
                    
                    body {{
                        font-family: 'Inter', Arial, sans-serif;
                        background-color: #f9fafb;
                        padding: 40px 20px;
                        color: #334155;
                        line-height: 1.6;
                        margin: 0;
                    }}
                    
                    .container {{
                        max-width: 560px;
                        margin: 0 auto;
                        background: white;
                        border-radius: 12px;
                        overflow: hidden;
                        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
                    }}
                    
                    .header {{
                        padding: 24px 32px;
                        text-align: center;
                        color: #111827;
                        border-bottom: 4px solid #dc3545;
                    }}
                    .header h1 {{
                        font-size: 24px;
                        font-weight: 600;
                        margin: 0;
                        color: #111827;
                    }}
                    .header p {{
                        font-size: 16px;
                        margin: 4px 0 0 0;
                        color: #6b7280;
                    }}
                    
                    .content {{
                        padding: 32px;
                    }}
                    
                    .greeting {{
                        font-size: 18px;
                        margin-bottom: 24px;
                        color: #1e293b;
                    }}
                    
                    .message {{
                        font-size: 16px;
                        margin-bottom: 24px;
                        color: #64748b;
                    }}
                    
                    .details {{
                        background: #fef2f2;
                        border: 1px solid #ef4444;
                        padding: 24px;
                        border-radius: 6px;
                        margin: 24px 0;
                    }}
                    
                    .details h4 {{
                        margin-bottom: 16px;
                        font-size: 16px;
                        font-weight: 600;
                        color: #7f1d1d;
                    }}
                    
                    .detail-item {{
                        margin-bottom: 12px;
                        line-height: 1.6;
                    }}
                    
                    .detail-item:last-child {{
                        margin-bottom: 0;
                    }}
                    
                    .detail-label {{
                        font-weight: bold;
                        color: #7f1d1d;
                        display: inline;
                    }}
                    
                    .detail-value {{
                        color: #7f1d1d;
                        display: inline;
                        margin-left: 4px;
                    }}
                    
                    .notice {{
                        background: #fef3c7;
                        border: 1px solid #f59e0b;
                        border-radius: 8px;
                        padding: 20px;
                        margin-bottom: 32px;
                        text-align: center;
                    }}
                    
                    .notice h3 {{
                        font-size: 18px;
                        font-weight: 600;
                        color: #92400e;
                        margin-bottom: 8px;
                    }}
                    
                    .notice p {{
                        font-size: 14px;
                        color: #92400e;
                        margin: 0;
                    }}
                    
                    .footer {{
                        background: #f1f5f9;
                        padding: 32px;
                        text-align: center;
                        border-top: 1px solid #e2e8f0;
                    }}
                    
                    .footer h3 {{
                        font-size: 18px;
                        font-weight: 600;
                        color: #1e293b;
                        margin-bottom: 8px;
                    }}
                    
                    .footer p {{
                        font-size: 14px;
                        color: #64748b;
                        margin-bottom: 4px;
                    }}
                    
                    .footer-divider {{
                        margin: 24px 0;
                        height: 1px;
                        background: #e2e8f0;
                    }}
                    
                    .footer-small {{
                        font-size: 12px;
                        color: #94a3b8;
                    }}
                    
                    @media (max-width: 600px) {{
                        .content {{ 
                            padding: 28px 20px; 
                        }}
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>Account Removed</h1>
                        <p>PPMS Cluster 4 - Imus City Healthcare Management</p>
                    </div>
                    
                    <div class="content">
                        <div class="greeting">
                            Hello <strong>{name}</strong>,
                        </div>
                        
                        <div class="message">
                            We would like to inform you that your Nurse account has been removed from the PPMS Cluster 4 system.
                        </div>
                        
                        <div class="notice">
                            <h3>Important Notice</h3>
                            <p>If you believe this was a mistake or have any questions, please contact the system administrator.</p>
                        </div>
                        
                        <div class="details">
                            <h4>Account Information</h4>
                            <div class="detail-item">
                                <span class="detail-label">Full Name:</span>
                                <span class="detail-value">{name}</span>
                            </div>
                            <div class="detail-item">
                                <span class="detail-label">Email:</span>
                                <span class="detail-value">{email}</span>
                            </div>
                            <div class="detail-item">
                                <span class="detail-label">Date Removed:</span>
                                <span class="detail-value">{current_date}</span>
                            </div>
                        </div>
                    </div>
                    
                    <div class="footer">
                        <h3>PPMS Cluster 4</h3>
                        <p>Imus City Healthcare Management</p>
                        <div class="footer-divider"></div>
                        <p class="footer-small">
                            This is an automated message. Please do not reply.<br>
                            © 2025 PPMS Cluster 4. All rights reserved.
                        </p>
                    </div>
                </div>
            </body>
            </html>
            """

            plain_message = f"""
PPMS Account Removal Notification

Hello {name},

We would like to inform you that your Nurse account has been removed from the PPMS Cluster 4 system.

IMPORTANT NOTICE:
If you believe this was a mistake or have any questions, please contact the system administrator.

Account Information:
- Full Name: {name}
- Email: {email}
- Date Removed: {current_date}

PPMS Cluster 4
Imus City Healthcare Management

This is an automated message. Please do not reply.
© 2025 PPMS Cluster 4. All rights reserved.
            """

            # ========== Background email sending ==========
            # Capture all data before threading
            email_subject = subject
            email_plain = plain_message
            email_html = html_message
            recipient_email = email
            from_email = settings.DEFAULT_FROM_EMAIL

            def send_removal_email():
                try:
                    send_mail(
                        subject=email_subject,
                        message=email_plain,
                        from_email=from_email,
                        recipient_list=[recipient_email],
                        html_message=email_html,
                        fail_silently=False
                    )
                    print(f"[DEBUG] ✅ Removal email sent successfully to {recipient_email}")
                except Exception as e:
                    print(f"[EMAIL ERROR]: {e}")

            # Start email sending in background thread
            import threading
            threading.Thread(target=send_removal_email).start()

            # Delete account
            nurse.delete()
            
            # Return JSON for AJAX requests (happens immediately)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'name': name, 'type': 'Nurse'})
            
            # Fallback for regular form submission
            messages.success(request, f"{name} has been successfully removed and notification email is being sent.")
            
        except Account.DoesNotExist:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Worker no longer exists'})
            messages.error(request, "The worker you're trying to remove no longer exists.")
        except Exception as e:
            print(f"[ERROR] Failed to remove Nurse: {e}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'An error occurred while removing the worker'})
            messages.error(request, "An error occurred while removing the nurse.")

    return redirect('healthcare_workers')

def fix_existing_bns_records():
    """
    Run this once to fix any existing BNS records that might have wrong role names
    You can call this from Django shell or create a management command
    """
    try:
        # Find accounts that might be BNS but have wrong role
        bns_accounts = Account.objects.filter(
            Q(user_role__icontains='BNS') | 
            Q(user_role__icontains='bns') |
            Q(user_role='BNS')
        )
        
        count = 0
        for account in bns_accounts:
            if account.user_role != 'Barangay Nutritional Scholar':
                account.user_role = 'Barangay Nutritional Scholar'
                account.save()
                count += 1
                
    
        return count
        
    except Exception as e:
        print(f"Error fixing BNS records: {e}")
        return 0
    


@login_required #dinagadag
def add_vaccination_schedule(request, preschooler_id):
    preschooler = get_object_or_404(Preschooler, pk=preschooler_id)

    if request.method == "POST":
        vaccine_name = request.POST.get("vaccine_name")
        doses = request.POST.get("vaccine_doses")
        required_doses = request.POST.get("required_doses")
        scheduled_date = request.POST.get("immunization_date")
        next_schedule = request.POST.get("next_vaccine_schedule")

        # Assuming BHW is logged in via request.user.account
        bhw = get_object_or_404(BHW, email=request.user.email)

        VaccinationSchedule.objects.create(
            preschooler=preschooler,
            vaccine_name=vaccine_name,
            doses=doses,
            required_doses=required_doses,
            scheduled_date=scheduled_date,
            next_vaccine_schedule=next_schedule,
            administered_by=bhw  # Optional, can leave null initially
        )

        messages.success(request, "Vaccination schedule has been saved successfully.")
        return redirect("preschooler_details", preschooler_id=preschooler_id)

    return redirect("preschooler_details", preschooler_id=preschooler_id)


from .who_lms import WHO_BMI_LMS
from .models import classify_bmi_for_age, calculate_bmi,bmi_zscore

@csrf_exempt
def submit_bmi(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.method == 'POST':
        preschooler_id = request.POST.get('preschooler_id')
        weight = request.POST.get('weight')
        height_cm = request.POST.get('height')
        temperature = request.POST.get('temperature')

        preschooler = get_object_or_404(Preschooler, pk=preschooler_id)

        if not weight or not height_cm or not temperature:
            return JsonResponse({
                'status': 'error',
                'message': 'All fields are required'
            })

        try:
            # calculate BMI
            weight = float(weight)
            height_m = float(height_cm) / 100
            bmi_value = weight / (height_m ** 2)

            # save BMI
            today = date.today()
            BMI.objects.update_or_create(
                preschooler_id=preschooler,
                date_recorded=today,
                defaults={'weight': weight, 'height': height_cm, 'bmi_value': bmi_value}
            )

            # save Temperature
            Temperature.objects.update_or_create(
                preschooler_id=preschooler,
                date_recorded=today,
                defaults={'temperature_value': temperature}
            )

            return JsonResponse({
                'status': 'success',
                'message': f'BMI & Temperature successfully recorded for {preschooler.first_name}!',
                'data': {
                    'preschooler_name': f'{preschooler.first_name} {preschooler.last_name}',
                    'weight': weight,
                    'height': height_cm,
                    'temperature': temperature,
                    'bmi': round(bmi_value, 2)
                }
            })

        except ValueError:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid number format. Please enter valid numbers.'
            })
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'An error occurred: {str(e)}'
            })

    # If GET → redirect to preschoolers
    return redirect('preschoolers')

    

    

def bmi_form(request, preschooler_id):
    preschooler = get_object_or_404(Preschooler, pk=preschooler_id)
    return render(request, 'HTML/bmi_form.html', {'preschooler': preschooler})

@csrf_exempt
def remove_preschooler(request):
    if request.method == 'POST':
        preschooler_id = request.POST.get('preschooler_id')
        try:
            preschooler = Preschooler.objects.get(pk=preschooler_id)
            preschooler.is_archived = True
            preschooler.save()
            return JsonResponse({'status': 'success'})
        except Preschooler.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Preschooler not found.'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})


def archived_preschoolers(request):

    if not request.user.is_authenticated:
        return redirect('login')
    search_query = request.GET.get('q', '').strip()

    archived = Preschooler.objects.filter(is_archived=True)

    # filter by name kung may hinanap
    if search_query:
        archived = archived.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )

    archived = archived.order_by('-date_registered')

    # pagination (10 rows per page)
    paginator = Paginator(archived, 10)
    page_number = request.GET.get('page')
    archived_page = paginator.get_page(page_number)

    return render(request, 'HTML/archived.html', {
        'archived_preschoolers': archived_page,
        'search_query': search_query
    })

@require_POST
def update_child_info(request, preschooler_id):
    preschooler = get_object_or_404(Preschooler, pk=preschooler_id)

    preschooler.place_of_birth = request.POST.get('place_of_birth') or ''
    preschooler.birth_weight = request.POST.get('birth_weight') or None
    preschooler.birth_height = request.POST.get('birth_height') or None
    preschooler.address = request.POST.get('address') or ''
    preschooler.save()

    if preschooler.parent_id:
        parent = preschooler.parent_id
        parent.mother_name = request.POST.get('mother_name') or ''
        parent.father_name = request.POST.get('father_name') or ''
        parent.save()

    messages.success(request, "Child information successfully updated.")
    return redirect('preschooler_detail', preschooler_id=preschooler.preschooler_id)





@login_required
def register_parent(request):
    if not request.user.is_authenticated:
        return redirect('login')
    """Register parent with proper barangay filtering - only allows registration in user's barangay"""
    
    if request.method == 'POST':
        try:
            # Get and validate form data
            first_name = request.POST.get('firstName', '').strip()
            middle_name = request.POST.get('middleName', '').strip()
            last_name = request.POST.get('lastName', '').strip()    
            suffix = request.POST.get('suffix', '').strip()
            email = request.POST.get('email', '').strip()
            contact_number = request.POST.get('contact_number', '').strip()
            birthdate = request.POST.get('birthdate', '').strip()
            sex = request.POST.get('sex', '').strip()

            # Basic validation
            if not all([first_name, last_name, email, contact_number, birthdate, sex]):
                messages.error(request, "All required fields must be filled.")
                return redirect('register_parent')
            
            # Validate sex field
            if sex not in ['Male', 'Female']:
                messages.error(request, "Please select a valid sex.")
                return redirect('register_parent')
            
            # Email validation
            if not email or '@' not in email:
                messages.error(request, "Please enter a valid email address.")
                return redirect('register_parent')
            
            # Address fields
            house_number = request.POST.get('houseNumber', '').strip()
            block = request.POST.get('block', '').strip()  
            lot = request.POST.get('lot', '').strip()
            phase = request.POST.get('phase', '').strip()
            street = request.POST.get('street', '').strip()
            subdivision = request.POST.get('subdivision', '').strip()
            city = request.POST.get('city', '').strip()
            province = request.POST.get('province', '').strip()

            # Define invalid values
            invalid_values = {'na', 'n/a', 'none', 'null', '--'}

            # Build complete address
            address_parts = []
            if house_number and house_number.lower() not in invalid_values:
                address_parts.append(f"{house_number}")
            if block and block.lower() not in invalid_values:
                address_parts.append(f"Block {block}")
            if lot and lot.lower() not in invalid_values:
                address_parts.append(f"Lot {lot}")
            if phase and phase.lower() not in invalid_values:
                address_parts.append(f"Phase {phase}")
            if street and street.lower() not in invalid_values:
                address_parts.append(street)
            if subdivision and subdivision.lower() not in invalid_values:
                address_parts.append(subdivision)
            if city and city.lower() not in invalid_values:
                address_parts.append(city)
            if province and province.lower() not in invalid_values:
                address_parts.append(province)

            address = ", ".join(address_parts) if address_parts else "No address provided"

            # Build full name
            name_parts = []
            if first_name:
                name_parts.append(first_name)
            if middle_name:
                name_parts.append(middle_name)
            if last_name:
                name_parts.append(last_name)
            if suffix and suffix.lower() not in ['na', 'n/a']:
                name_parts.append(suffix)
            
            full_name = " ".join(name_parts).strip()

            # Get user's barangay
            user_barangay = get_user_barangay(request.user)
            current_user_info = None
            
            logger.info(f"Registration attempt by user: {request.user.email}")
            
            if request.user.is_authenticated:
                # Try each model to find the user and their barangay
                try:
                    account = Account.objects.select_related('barangay').get(email=request.user.email)
                    current_user_info = {
                        'model': 'Account',
                        'name': account.full_name,
                        'role': account.user_role
                    }
                except Account.DoesNotExist:
                    for model_class, role_name in [
                        (BHW, 'BHW'),
                        (BNS, 'BNS'),
                        (Midwife, 'Midwife'),
                        (Nurse, 'Nurse'),
                    ]:
                        try:
                            user_obj = model_class.objects.select_related('barangay').get(email=request.user.email)
                            current_user_info = {
                                'model': role_name,
                                'name': user_obj.full_name,
                                'role': role_name
                            }
                            break
                        except model_class.DoesNotExist:
                            continue

            # Validate that user exists and has proper authorization
            if not current_user_info:
                logger.error(f"Unauthorized registration attempt by {request.user.email}")
                messages.error(request, "You are not authorized to register parents. Please contact the administrator.")
                return redirect('register_parent')

            # Validate that user has a barangay assigned
            if not user_barangay:
                logger.error(f"No barangay assigned to user {request.user.email}")
                messages.error(request, f"No barangay assigned to your {current_user_info['role']} account. Please contact the administrator.")
                return redirect('register_parent')

            # Validate user role permissions (only for Account model)
            if current_user_info['model'] == 'Account':
                allowed_roles = ['bhw', 'bns', 'barangay nutritional scholar', 'barangay nutrition scholar', 
                               'nutritional scholar', 'nutrition scholar', 'midwife', 'nurse', 'admin', 'administrator']
                if current_user_info['role'].lower() not in [role.lower() for role in allowed_roles]:
                    logger.error(f"Unauthorized role: {current_user_info['role']}")
                    messages.error(request, f"Your role '{current_user_info['role']}' is not authorized to register parents.")
                    return redirect('register_parent')

            logger.info(f"Authorization passed - Role: {current_user_info['role']}, Barangay: {user_barangay}")

            # ✅ FIXED: Check for existing records with proper orphan cleanup
            parent_exists = Parent.objects.filter(email__iexact=email).exists()
            user_exists = User.objects.filter(email__iexact=email).exists()
            account_exists = Account.objects.filter(email__iexact=email).exists()
            
            # Handle orphaned User records (User exists but no Parent/Account)
            if user_exists and not parent_exists and not account_exists:
                try:
                    User.objects.filter(email__iexact=email).delete()
                    logger.info(f"🗑️ Deleted orphaned User for email: {email}")
                except Exception as e:
                    logger.error(f"❌ Error deleting orphaned User: {e}")
                    messages.error(request, "An error occurred. Please try again.")
                    return redirect('register_parent')
            elif parent_exists:
                messages.error(request, "A parent with this email already exists.")
                return redirect('register_parent')
            elif account_exists:
                messages.error(request, "A user with this email already exists.")
                return redirect('register_parent')

            # Check if contact number already exists in same barangay
            if Parent.objects.filter(contact_number=contact_number, barangay=user_barangay).exists():
                messages.error(request, f"A parent with this contact number already exists in {user_barangay.name}.")
                return redirect('register_parent')

            # Parse and validate birthdate
            try:
                birthdate_obj = datetime.strptime(birthdate, '%Y-%m-%d').date()
                logger.info(f"Birthdate parsed successfully: {birthdate_obj}")
            except ValueError as e:
                logger.error(f"Birthdate parsing error: {e}")
                messages.error(request, "Invalid birthdate format.")
                return redirect('register_parent')

            # Generate password
            raw_password = generate_password()

            # Use database transaction for atomicity
            with transaction.atomic():
                logger.info(f"Starting database transaction for {full_name}")
                
                # Create Django User with hashed password
                logger.info(f"Creating Django User for {email}")
                user = User.objects.create_user(
                    username=email,
                    email=email,
                    password=raw_password
                )
                logger.info("Django User created successfully")

                try:
                    # Create Parent with correct field names
                    logger.info("Creating Parent record")
                    parent = Parent.objects.create(
                        first_name=first_name,
                        middle_name=middle_name,
                        suffix=suffix,
                        last_name=last_name,
                        sex=sex,
                        email=email,
                        contact_number=contact_number,
                        birthdate=birthdate_obj,
                        address=address,
                        barangay=user_barangay,
                        must_change_password=True,
                        password=raw_password,
                        created_at=timezone.now()
                    )
                    logger.info("Parent created successfully")

                    # Create Account - also assign to same barangay
                    logger.info("Creating Account record")
                    account = Account.objects.create(
                        email=email,
                        first_name=first_name,
                        middle_name=middle_name,
                        suffix=suffix,
                        last_name=last_name,
                        sex=sex,
                        contact_number=contact_number,
                        birthdate=birthdate_obj,
                        user_role='parent',
                        barangay=user_barangay,
                        is_validated=False,
                        is_rejected=False,
                        last_activity=timezone.now(),
                        password=raw_password,
                        must_change_password=True
                    )
                    logger.info("Account created successfully")
                    
                except Exception as e:
                    # If Parent or Account creation fails, delete the User to prevent orphans
                    logger.error(f"❌ Error creating Parent/Account: {e}")
                    user.delete()
                    logger.info("🗑️ Rolled back User creation due to error")
                    raise  # Re-raise to trigger outer exception handling

            # ========== Background email sending ==========
            def send_parent_registration_email():
                try:
                    subject = "PPMS Cluster 4 – Parent Registration Successful"
                    html_message = f"""
                    <html>
                    <body style='font-family: Arial, sans-serif; background-color: #f9fafb; padding: 20px;'>
                        <div style='max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
                            <div style='text-align: center; border-bottom: 3px solid #198754; padding-bottom: 20px; margin-bottom: 20px;'>
                                <h1 style='color: #198754; margin: 0;'>PPMS Cluster 4</h1>
                                <p style='color: #6b7280; margin: 5px 0 0 0;'>Imus City Healthcare Management</p>
                            </div>
                            
                            <h2 style='text-align: center; color: #1e293b;'>Registration Successful!</h2>
                            
                            <p style='font-size: 16px; color: #334155;'>Hello <strong>{full_name}</strong>,</p>
                            
                            <p style='font-size: 16px; color: #334155;'>
                                Your parent account has been successfully registered for <strong>{user_barangay.name}</strong>.
                            </p>
                            
                            <div style='background: #ecfdf5; border: 1px solid #10b981; padding: 20px; border-radius: 8px; margin: 20px 0;'>
                                <h3 style='color: #065f46; margin-top: 0;'>Login Credentials</h3>
                                <p style='margin: 10px 0; color: #065f46;'>
                                    <strong>Email:</strong> {email}<br>
                                    <strong>Password:</strong> <code style='background: #d1fae5; padding: 4px 8px; border-radius: 4px; font-size: 14px;'>{raw_password}</code>
                                </p>
                            </div>
                            
                            <div style='background: #fef3c7; border: 1px solid #f59e0b; padding: 15px; border-radius: 8px; margin: 20px 0;'>
                                <p style='margin: 0; color: #92400e;'>
                                    <strong>⚠️ Important:</strong> You must change your password on first login for security purposes.
                                </p>
                            </div>
                            
                            <div style='text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #e5e7eb;'>
                                <p style='font-size: 13px; color: #6b7280;'>
                                    This is an automated message. Please do not reply.<br>
                                    © 2025 PPMS Cluster 4. All rights reserved.
                                </p>
                            </div>
                        </div>
                    </body>
                    </html>
                    """
                    plain_message = f"""
PPMS Cluster 4 – Parent Registration Successful

Hello {full_name},

Your parent account has been successfully registered for {user_barangay.name}.

Login Credentials:
Email: {email}
Password: {raw_password}

⚠️ IMPORTANT: You must change your password on first login for security purposes.

PPMS Cluster 4
Imus City Healthcare Management

This is an automated message. Please do not reply.
© 2025 PPMS Cluster 4. All rights reserved.
                    """
                    
                    send_mail(
                        subject, 
                        plain_message, 
                        settings.DEFAULT_FROM_EMAIL, 
                        [email], 
                        html_message=html_message, 
                        fail_silently=False
                    )
                    logger.info(f"✅ Email sent successfully to {email}")
                except Exception as email_error:
                    logger.warning(f"[EMAIL ERROR]: {email_error}")

            # Start email sending in background thread
            threading.Thread(target=send_parent_registration_email).start()

            # Success message (displayed immediately without waiting for email)
            messages.success(request, f"Parent '{full_name}' registered successfully in {user_barangay.name}!\nEmail: {email}\nPassword: {raw_password}")
            return redirect('register_parent')

        except IntegrityError as e:
            logger.error(f"IntegrityError: {e}")
            messages.error(request, "Registration failed due to duplicate data. Please check email and contact number.")
            return redirect('register_parent')
        
        except Exception as e:
            logger.error(f"Unexpected error during registration: {e}")
            messages.error(request, "An unexpected error occurred during registration. Please try again.")
            return redirect('register_parent')

    # For GET request
    context = {}
    if request.user.is_authenticated:
        user_barangay = get_user_barangay(request.user)
        if user_barangay:
            context['user_barangay'] = user_barangay

        try:
            account = Account.objects.get(email=request.user.email)
            context['account'] = account
        except Account.DoesNotExist:
            context['account'] = None

    return render(request, 'HTML/register_parent.html', context)


def get_user_barangay(user):
    """Helper function to get user's barangay from any user model"""
    if not user.is_authenticated:
        return None
    
    # Try each model in order
    models_to_check = [
        (Account, 'email'),
        (BHW, 'email'), 
        (BNS, 'email'),
        (Midwife, 'email'),
        (Nurse, 'email'),
        (Parent, 'email'),
    ]
    
    for model_class, field_name in models_to_check:
        try:
            obj = model_class.objects.get(**{field_name: user.email})
            return obj.barangay
        except model_class.DoesNotExist:
            continue
    
    return None

def generate_password(length=8):
    characters = string.ascii_letters + string.digits
    return ''.join(random.choice(characters) for _ in range(length))

from django.contrib.auth import update_session_auth_hash

def change_password_first(request):
    if request.method == 'POST':
        email = request.session.get('email', '').strip()
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if not email:
            messages.error(request, "Session expired. Please log in again.")
            return redirect('login')

        if new_password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect('change_password_first')

        try:
            from .models import Account, Parent
            
            # Get the User object
            user = User.objects.get(username=email)
            user.set_password(new_password)
            user.save()

            # Try to find if it's a Parent account
            try:
                parent = Parent.objects.get(email=email)
                parent.must_change_password = False
                parent.save()
                
                # Keep user logged in after password change
                update_session_auth_hash(request, user)
                
                messages.success(request, "Password updated successfully!")
                return redirect('parent_dashboard')
                
            except Parent.DoesNotExist:
                # If not Parent, try Account (for Admin, BHW, etc.)
                try:
                    account = Account.objects.get(email=email)
                    account.password = user.password  # Update hashed password
                    account.must_change_password = False  # ✅ CRITICAL: Disable forced password change
                    account.save(update_fields=['password', 'must_change_password'])
                    
                    # Keep user logged in after password change
                    update_session_auth_hash(request, user)
                    
                    messages.success(request, "Password updated successfully!")
                    
                    # Redirect based on role
                    if account.user_role.lower() == 'admin':
                        return redirect('Admindashboard')
                    else:
                        return redirect('dashboard')
                        
                except Account.DoesNotExist:
                    messages.error(request, "Account record not found.")
                    return redirect('login')

        except User.DoesNotExist:
            messages.error(request, "User not found.")
            return redirect('change_password_first')

    return render(request, 'HTML/parent_change_password.html')





@login_required
def history(request):
    if not request.user.is_authenticated:
        return redirect('login')
    """History view with proper barangay filtering - only shows logs from user's barangay"""
    
    user_barangay = get_user_barangay(request.user)
    current_user_info = None

    if request.user.is_authenticated:
        # Try each model to find the user
        try:
            account = Account.objects.select_related('barangay').get(email=request.user.email)
            current_user_info = {
                'model': 'Account',
                'name': account.full_name,
                'role': account.user_role,
                'object': account
            }
        except Account.DoesNotExist:
            try:
                bhw = BHW.objects.select_related('barangay').get(email=request.user.email)
                current_user_info = {
                    'model': 'BHW',
                    'name': bhw.full_name,
                    'role': 'BHW',
                    'object': bhw
                }
            except BHW.DoesNotExist:
                try:
                    bns = BNS.objects.select_related('barangay').get(email=request.user.email)
                    current_user_info = {
                        'model': 'BNS',
                        'name': bns.full_name,
                        'role': 'BNS',
                        'object': bns
                    }
                except BNS.DoesNotExist:
                    try:
                        midwife = Midwife.objects.select_related('barangay').get(email=request.user.email)
                        current_user_info = {
                            'model': 'Midwife',
                            'name': midwife.full_name,
                            'role': 'Midwife',
                            'object': midwife
                        }
                    except Midwife.DoesNotExist:
                        try:
                            nurse = Nurse.objects.select_related('barangay').get(email=request.user.email)
                            current_user_info = {
                                'model': 'Nurse',
                                'name': nurse.full_name,
                                'role': 'Nurse',
                                'object': nurse
                            }
                        except Nurse.DoesNotExist:
                            try:
                                parent = Parent.objects.select_related('barangay').get(email=request.user.email)
                                current_user_info = {
                                    'model': 'Parent',
                                    'name': parent.full_name,
                                    'role': 'Parent',
                                    'object': parent
                                }
                            except Parent.DoesNotExist:
                                print("DEBUG: User not found in any model")

    if not current_user_info or not user_barangay:
        print(f"DEBUG: No user info or barangay found. User info: {current_user_info}, Barangay: {user_barangay}")
        return render(request, 'HTML/history.html', {
            'account': None,
            'parent_logs': [],
            'preschooler_logs': [],
            'user_barangay': user_barangay,
            'error_message': 'No barangay assigned to your account or user not found.'
        })

    print(f"DEBUG: History access authorized for {current_user_info['role']} in {user_barangay}")

    # 🔹 Removed automatic deletion of logs older than 1 day

    # Get logs ONLY from user's barangay
    parent_logs = ParentActivityLog.objects.filter(barangay=user_barangay).select_related('parent', 'barangay').order_by('-timestamp')
    preschooler_logs = PreschoolerActivityLog.objects.filter(barangay=user_barangay).select_related('barangay').order_by('-timestamp')

    parent_paginator = Paginator(parent_logs, 10)
    preschooler_paginator = Paginator(preschooler_logs, 10)

    parent_page_number = request.GET.get('parent_page')
    preschooler_page_number = request.GET.get('preschooler_page')

    parent_logs_page = parent_paginator.get_page(parent_page_number)
    preschooler_logs_page = preschooler_paginator.get_page(preschooler_page_number)

    return render(request, 'HTML/history.html', {
        'account': current_user_info['object'],
        'parent_logs': parent_logs_page,
        'preschooler_logs': preschooler_logs_page,
        'user_barangay': user_barangay,
        'current_user_info': current_user_info,
        'total_parent_logs': parent_logs.count(),
        'total_preschooler_logs': preschooler_logs.count(),
    })




# Additional helper function to create activity logs with barangay validation
def create_parent_activity_log(parent, activity, performed_by_user):
    """Create parent activity log with barangay validation"""
    user_barangay = get_user_barangay(performed_by_user)
    
    # Only create log if parent and user are in same barangay
    if user_barangay and parent.barangay == user_barangay:
        try:
            # Get user info for logging
            user_info = None
            try:
                account = Account.objects.get(email=performed_by_user.email)
                user_info = f"{account.full_name} ({account.user_role})"
            except Account.DoesNotExist:
                # Try other user models
                for model_class in [BHW, BNS, Midwife, Nurse]:
                    try:
                        user_obj = model_class.objects.get(email=performed_by_user.email)
                        user_info = f"{user_obj.full_name} ({model_class.__name__})"
                        break
                    except model_class.DoesNotExist:
                        continue
            
            if not user_info:
                user_info = performed_by_user.email
            
            ParentActivityLog.objects.create(
                parent=parent,
                barangay=user_barangay,
                activity=activity,
                performed_by=user_info  # Assuming you have this field
            )
        except Exception as e:  # ← THIS WAS MISSING
            print(f"DEBUG: Error creating parent activity log: {e}")

def create_preschooler_activity_log(preschooler, activity, performed_by_user):
    """Create preschooler activity log with barangay validation"""
    user_barangay = get_user_barangay(performed_by_user)
    
    # Only create log if preschooler and user are in same barangay
    if user_barangay and preschooler.barangay == user_barangay:
        try:
            # Get user info for logging
            user_info = None
            try:
                account = Account.objects.get(email=performed_by_user.email)
                user_info = f"{account.full_name} ({account.user_role})"
            except Account.DoesNotExist:
                # Try other user models
                for model_class in [BHW, BNS, Midwife, Nurse]:
                    try:
                        user_obj = model_class.objects.get(email=performed_by_user.email)
                        user_info = f"{user_obj.full_name} ({model_class.__name__})"
                        break
                    except model_class.DoesNotExist:
                        continue
            
            if not user_info:
                user_info = performed_by_user.email
            
            PreschoolerActivityLog.objects.create(
                preschooler_name=f"{preschooler.first_name} {preschooler.last_name}",
                barangay=user_barangay,
                activity=activity,
                performed_by=user_info
            )
        except Exception as e:  # ← THIS WAS MISSING
            print(f"DEBUG: Error creating preschooler activity log: {e}")

@admin_required
def admin_logs(request):
    if request.session.get('user_role') != 'admin':
        return redirect('login')

    # Removed automatic deletion of logs older than 1 day

    # Filter to show only "Transferred to" activities
    parent_logs_all = ParentActivityLog.objects.select_related('parent', 'barangay').filter(
        activity__startswith='Transferred to'
    ).order_by('-timestamp')
    
    preschooler_logs_all = PreschoolerActivityLog.objects.select_related('barangay').filter(
        activity__startswith='Transferred to'
    ).exclude(
        activity__startswith='Recently transferred'
    ).order_by('-timestamp')

    # Paginate
    parent_paginator = Paginator(parent_logs_all, 10)  # 10 per page
    preschooler_paginator = Paginator(preschooler_logs_all, 20)

    parent_page = request.GET.get('parent_page')
    preschooler_page = request.GET.get('preschooler_page')

    context = {
        'parent_logs': parent_paginator.get_page(parent_page),
        'preschooler_logs': preschooler_paginator.get_page(preschooler_page),
    }

    return render(request, 'HTML/admin_logs.html', context)






from rest_framework_simplejwt.tokens import RefreshToken





















from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.utils import timezone
from django.contrib.auth.models import User
from .models import Announcement
import cloudinary.uploader

@admin_required    
def manage_announcements(request):
    """
    View to display all announcements with options to add, edit, delete
    """
    try:
        announcements = Announcement.objects.all().order_by('-created_at')
    except Exception as e:
        messages.error(request, 'Error loading announcements. Please ensure the database is properly set up.')
        announcements = []
    
    context = {
        'announcements': announcements,
        'account': request.user,
    }
    
    return render(request, 'HTML/manage_announcements.html', context)


def add_announcement(request):
    """
    View to add a new announcement with Cloudinary image upload
    """
    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        is_active = request.POST.get('is_active') == 'on'
        image = request.FILES.get('image', None)
        
        if title and content:
            try:
                # Manual user retrieval
                if hasattr(request, 'user') and request.user.is_authenticated:
                    user_obj = User.objects.get(id=request.user.id)
                    created_by = user_obj
                else:
                    created_by = None
                
                # Create announcement instance
                announcement = Announcement(
                    title=title,
                    content=content,
                    is_active=is_active,
                    created_by=created_by,
                    created_at=timezone.now()
                )
                
                # Upload image to Cloudinary if provided
                if image:
                    upload_result = cloudinary.uploader.upload(
                        image,
                        folder='announcements/',
                        resource_type='image'
                    )
                    announcement.image = upload_result['secure_url']
                    announcement.cloudinary_public_id = upload_result['public_id']
                
                announcement.save()
                messages.success(request, 'Announcement added successfully!')
                
            except User.DoesNotExist:
                messages.error(request, 'User not found. Please log in again.')
            except Exception as e:
                messages.error(request, f'Error creating announcement: {str(e)}')
        else:
            messages.error(request, 'Title and content are required!')
    
    return redirect('manage_announcements')


def edit_announcement(request, announcement_id):
    """
    View to edit an existing announcement with Cloudinary image replacement support
    """
    announcement = get_object_or_404(Announcement, id=announcement_id)
    
    if request.method == 'POST':
        announcement.title = request.POST.get('title')
        announcement.content = request.POST.get('content')
        announcement.is_active = request.POST.get('is_active') == 'on'
        announcement.updated_at = timezone.now()
        
        # Handle image replacement - only if a new image is uploaded
        new_image = request.FILES.get('image', None)
        if new_image:
            # Delete old image from Cloudinary if it exists
            if hasattr(announcement, 'cloudinary_public_id') and announcement.cloudinary_public_id:
                try:
                    cloudinary.uploader.destroy(announcement.cloudinary_public_id)
                except Exception as e:
                    print(f"Error deleting old image from Cloudinary: {str(e)}")
            
            # Upload new image to Cloudinary
            try:
                upload_result = cloudinary.uploader.upload(
                    new_image,
                    folder='announcements/',
                    resource_type='image'
                )
                announcement.image = upload_result['secure_url']
                if hasattr(announcement, 'cloudinary_public_id'):
                    announcement.cloudinary_public_id = upload_result['public_id']
            except Exception as e:
                messages.error(request, f'Error uploading image: {str(e)}')
                return redirect('manage_announcements')
        
        # If no new image is uploaded, keep the existing image (do nothing)
        
        if announcement.title and announcement.content:
            try:
                announcement.save()
                messages.success(request, 'Announcement updated successfully!')
            except Exception as e:
                messages.error(request, f'Error updating announcement: {str(e)}')
        else:
            messages.error(request, 'Title and content are required!')
    
    return redirect('manage_announcements')


def delete_announcement(request, announcement_id):
    """
    View to delete an announcement and its Cloudinary image
    """
    if request.method == 'POST':
        try:
            announcement = get_object_or_404(Announcement, id=announcement_id)
            
            # Delete associated image from Cloudinary if it exists
            if announcement.cloudinary_public_id:
                try:
                    cloudinary.uploader.destroy(announcement.cloudinary_public_id)
                except Exception as e:
                    print(f"Error deleting image from Cloudinary: {str(e)}")
            
            announcement.delete()
            messages.success(request, 'Announcement deleted successfully!')
        except Exception as e:
            messages.error(request, f'Error deleting announcement: {str(e)}')
    
    return redirect('manage_announcements')


def get_announcement(request, announcement_id):
    """
    API view to return announcement data as JSON for editing
    """
    try:
        announcement = get_object_or_404(Announcement, id=announcement_id)
        
        data = {
            'status': 'success',
            'data': {
                'title': announcement.title,
                'content': announcement.content,
                'is_active': announcement.is_active,
                'image_url': announcement.image.url if announcement.image else None,
            }
        }
        
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)



from django.db.models import Count, Q

@admin_required
def registered_barangays(request):
    import re
    from django.db.models import Count, Q
    from django.core.paginator import Paginator

    query = request.GET.get("search", "").strip()

    barangays = Barangay.objects.annotate(
        preschooler_count=Count("preschooler", distinct=True),
        parent_count=Count("parent", distinct=True),
        bhw_bns_count=Count(
            "account",
            filter=Q(account__user_role__in=["BHW", "Barangay Nutritional Scholar"]),
            distinct=True,
        ),
    )

    if query:
        barangays = barangays.filter(
            Q(name__icontains=query)
            | Q(phone_number__icontains=query)
            | Q(hall_address__icontains=query)
        )

    # ✅ Natural sorting (handles "2-A", "2-B", "2-C", ... correctly)
    def natural_sort_key(x):
        # Split into digits and letters so numbers sort numerically
        return [
            int(part) if part.isdigit() else part.upper()
            for part in re.split(r"(\d+)", x.name)
        ]

    barangays_list = sorted(list(barangays), key=natural_sort_key)

    paginator = Paginator(barangays_list, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "HTML/barangay_list.html", {"barangays": page_obj})
    
@admin_required 
def healthcare_workers(request):
    """Improved healthcare workers view with unified pagination"""
    from django.utils import timezone
    from django.utils.timesince import timesince
    from django.db.models import Q
    from datetime import timedelta
    
    print("\n=== HEALTHCARE WORKERS VIEW DEBUG ===")
    
    # Get filter parameters from URL
    role_filter = request.GET.get('role', 'all')
    barangay_filter = request.GET.get('barangay', 'all')
    page_number = request.GET.get('page', 1)
    
    print(f"Filters: role={role_filter}, barangay={barangay_filter}, page={page_number}")
    
    # Get all barangays for the filter dropdown
    barangays = Barangay.objects.all().order_by('name')
    
    # ===== COLLECT ALL WORKERS =====
    all_workers = []
    
    # ===== BHW DATA =====
    print("\n--- Fetching BHW Data ---")
    bhw_accounts = Account.objects.filter(
        Q(user_role__iexact='healthworker') | Q(user_role__iexact='BHW'),
        is_validated=True
    ).select_related('barangay')
    
    print(f"Found {bhw_accounts.count()} BHW accounts")
    
    for bhw in bhw_accounts:
        try:
            bhw_profile = BHW.objects.filter(email=bhw.email).first()
        except Exception as e:
            print(f"Error getting BHW profile for {bhw.email}: {str(e)}")
            bhw_profile = None
        
        set_activity_status(bhw)
        
        # Determine barangay
        barangay_name = 'No Barangay'
        barangay_lower = 'no barangay'
        if bhw.barangay:
            barangay_name = bhw.barangay.name
            barangay_lower = bhw.barangay.name.lower()
        elif bhw_profile and bhw_profile.barangay:
            barangay_name = bhw_profile.barangay.name
            barangay_lower = bhw_profile.barangay.name.lower()
        
        all_workers.append({
            'type': 'bhw',
            'full_name': bhw.full_name,
            'role': 'Health Worker (BHW)',
            'role_class': 'bhw',
            'barangay': barangay_name,
            'barangay_lower': barangay_lower,
            'contact_number': bhw.contact_number or 'N/A',
            'email': bhw.email,
            'last_activity': bhw.last_activity_display,
            'account_id': bhw.account_id,
            'pk': bhw.pk,
            'remove_url': 'remove_bhw',
            'account': bhw,
            'profile': bhw_profile
        })
    
    # ===== BNS DATA =====
    print("\n--- Fetching BNS Data ---")
    
    # Simplified BNS query - try all common variations
    bns_accounts = Account.objects.filter(
        Q(user_role__iexact='bns') |
        Q(user_role__iexact='BNS') |
        Q(user_role__icontains='nutritional') |
        Q(user_role__icontains='Nutritional') |
        Q(user_role__icontains='scholar') |
        Q(user_role__icontains='Scholar'),
        is_validated=True
    ).select_related('barangay').distinct()
    
    print(f"Found {bns_accounts.count()} BNS accounts")
    
    for bns in bns_accounts:
        try:
            # Try to find BNS profile by email first, then by name
            bns_profile = BNS.objects.filter(email=bns.email).first()
            if not bns_profile:
                bns_profile = BNS.objects.filter(full_name__iexact=bns.full_name).first()
        except Exception as e:
            print(f"Error getting BNS profile for {bns.email}: {str(e)}")
            bns_profile = None
        
        set_activity_status(bns)
        
        # Determine barangay
        barangay_name = 'No Barangay'
        barangay_lower = 'no barangay'
        if bns.barangay:
            barangay_name = bns.barangay.name
            barangay_lower = bns.barangay.name.lower()
        elif bns_profile and bns_profile.barangay:
            barangay_name = bns_profile.barangay.name
            barangay_lower = bns_profile.barangay.name.lower()
        
        all_workers.append({
            'type': 'bns',
            'full_name': bns.full_name,
            'role': 'BNS',
            'role_class': 'bns',
            'barangay': barangay_name,
            'barangay_lower': barangay_lower,
            'contact_number': bns.contact_number or 'N/A',
            'email': bns.email,
            'last_activity': bns.last_activity_display,
            'account_id': bns.account_id,
            'pk': bns.pk,
            'remove_url': 'remove_bns',
            'account': bns,
            'profile': bns_profile
        })
    
    # ===== MIDWIFE DATA =====
    print("\n--- Fetching Midwife Data ---")
    midwife_accounts = Account.objects.filter(
        Q(user_role__iexact='midwife') | Q(user_role__iexact='Midwife'),
        is_validated=True
    ).select_related('barangay')
    
    print(f"Found {midwife_accounts.count()} Midwife accounts")
    
    for midwife in midwife_accounts:
        try:
            midwife_profile = Midwife.objects.filter(email=midwife.email).first()
        except Exception as e:
            print(f"Error getting Midwife profile for {midwife.email}: {str(e)}")
            midwife_profile = None
        
        set_activity_status(midwife)
        
        # Determine barangay
        barangay_name = 'No Barangay'
        barangay_lower = 'no barangay'
        if midwife.barangay:
            barangay_name = midwife.barangay.name
            barangay_lower = midwife.barangay.name.lower()
        elif midwife_profile and midwife_profile.barangay:
            barangay_name = midwife_profile.barangay.name
            barangay_lower = midwife_profile.barangay.name.lower()
        
        all_workers.append({
            'type': 'midwife',
            'full_name': midwife.full_name,
            'role': 'Midwife',
            'role_class': 'midwife',
            'barangay': barangay_name,
            'barangay_lower': barangay_lower,
            'contact_number': midwife.contact_number or 'N/A',
            'email': midwife.email,
            'last_activity': midwife.last_activity_display,
            'account_id': midwife.account_id,
            'pk': midwife.pk,
            'remove_url': 'remove_midwife',
            'account': midwife,
            'profile': midwife_profile
        })
    
    # ===== NURSE DATA =====
    print("\n--- Fetching Nurse Data ---")
    nurse_accounts = Account.objects.filter(
        Q(user_role__iexact='nurse') | Q(user_role__iexact='Nurse'),
        is_validated=True
    ).select_related('barangay')
    
    print(f"Found {nurse_accounts.count()} Nurse accounts")
    
    for nurse in nurse_accounts:
        try:
            nurse_profile = Nurse.objects.filter(email=nurse.email).first()
        except Exception as e:
            print(f"Error getting Nurse profile for {nurse.email}: {str(e)}")
            nurse_profile = None
        
        set_activity_status(nurse)
        
        # Determine barangay
        barangay_name = 'No Barangay'
        barangay_lower = 'no barangay'
        if nurse.barangay:
            barangay_name = nurse.barangay.name
            barangay_lower = nurse.barangay.name.lower()
        elif nurse_profile and nurse_profile.barangay:
            barangay_name = nurse_profile.barangay.name
            barangay_lower = nurse_profile.barangay.name.lower()
        
        all_workers.append({
            'type': 'nurse',
            'full_name': nurse.full_name,
            'role': 'Nurse',
            'role_class': 'nurse',
            'barangay': barangay_name,
            'barangay_lower': barangay_lower,
            'contact_number': nurse.contact_number or 'N/A',
            'email': nurse.email,
            'last_activity': nurse.last_activity_display,
            'account_id': nurse.account_id,
            'pk': nurse.pk,
            'remove_url': 'remove_nurse',
            'account': nurse,
            'profile': nurse_profile
        })
    
    print(f"\nTotal workers collected: {len(all_workers)}")
    
    # ===== APPLY FILTERS =====
    filtered_workers = all_workers
    
    # Role filter
    if role_filter != 'all':
        filtered_workers = [w for w in filtered_workers if w['type'] == role_filter]
        print(f"After role filter '{role_filter}': {len(filtered_workers)} workers")
    
    # Barangay filter
    if barangay_filter != 'all':
        if barangay_filter == 'no barangay':
            filtered_workers = [w for w in filtered_workers if w['barangay_lower'] == 'no barangay']
        else:
            filtered_workers = [w for w in filtered_workers if barangay_filter.lower() in w['barangay_lower']]
        print(f"After barangay filter '{barangay_filter}': {len(filtered_workers)} workers")
    
    # ===== PAGINATION - 10 items per page =====
    paginator = Paginator(filtered_workers, 10)
    
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    print(f"Pagination: Page {page_obj.number} of {paginator.num_pages}")
    print(f"Showing {len(page_obj.object_list)} workers on this page")
    
    # Count by role for filter badges
    bhw_count = len([w for w in all_workers if w['type'] == 'bhw'])
    bns_count = len([w for w in all_workers if w['type'] == 'bns'])
    midwife_count = len([w for w in all_workers if w['type'] == 'midwife'])
    nurse_count = len([w for w in all_workers if w['type'] == 'nurse'])
    
    print(f"\nCounts - BHW: {bhw_count}, BNS: {bns_count}, Midwife: {midwife_count}, Nurse: {nurse_count}")
    
    context = {
        'page_obj': page_obj,
        'workers': page_obj.object_list,
        'barangays': barangays,
        'current_role': role_filter,
        'current_barangay': barangay_filter,
        'bhw_count': bhw_count,
        'bns_count': bns_count,
        'midwife_count': midwife_count,
        'nurse_count': nurse_count,
        'total_workers': len(all_workers),
        # Legacy support - in case template still references these
        'bhws': bhw_accounts,
        'bnss': bns_accounts,
        'midwives': midwife_accounts,
        'nurses': nurse_accounts,
    }
    
    return render(request, 'HTML/healthcare_workers.html', context)




def set_activity_status(user):
    """Helper function to set activity status for any user"""
    from django.utils import timezone
    from django.utils.timesince import timesince
    from datetime import timedelta
    
    if hasattr(user, 'last_activity') and user.last_activity:
        if timezone.now() - user.last_activity <= timedelta(minutes=1):
            user.last_activity_display = "🟢 Online"
        else:
            time_diff = timesince(user.last_activity, timezone.now())
            user.last_activity_display = f"{time_diff} ago"
    else:
        user.last_activity_display = "No activity"


@csrf_exempt
def get_announcement_data(request, announcement_id):
    """
    API endpoint to get announcement data for editing (AJAX)
    """
    if request.method == 'GET':
        try:
            announcement = get_object_or_404(Announcement, id=announcement_id)
            data = {
                'id': announcement.id,
                'title': announcement.title,
                'content': announcement.content,
                'priority': announcement.priority,
                'is_active': announcement.is_active,
                'image_url': announcement.image.url if announcement.image else None,
            }
            return JsonResponse({'status': 'success', 'data': data})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

def get_latest_weight(request):
    """Return the latest weight data from hardware"""
    try:
        
        latest_weight = 0.0  # Get from your hardware source
        
        return JsonResponse({'weight': latest_weight})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_latest_temp(request):
    """Return the latest temperature data from hardware"""
    try:
        # Replace this with your actual logic to get temperature from hardware/database
        latest_temp = 0.0  # Get from your hardware source
        
        return JsonResponse({'temperature': latest_temp})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_latest_distance(request):
    """Return the latest distance data from hardware"""
    try:
        # Replace this with your actual logic to get distance from hardware/database
        latest_distance = 0.0  # Get from your hardware source
        
        return JsonResponse({'distance': latest_distance})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    









import traceback
@csrf_exempt
def generate_report(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    print("=== GENERATE REPORT DEBUG ===")
    print("User: {request.user}")
    print("Method: {request.method}")
    
    try:
        # Check if WeasyPrint is available
        try:
            from weasyprint import HTML
            print("✓ WeasyPrint imported successfully")
        except ImportError as e:
            print("✗ WeasyPrint import failed: {e}")
            messages.error(request, "WeasyPrint is not installed. Please install it with: pip install weasyprint")
            return redirect('Admindashboard')
        
        # Handle hardcoded admin account
        user = request.user
        account = None
        
        if user.is_authenticated:
            try:
                account = Account.objects.select_related('barangay').get(email=user.email)
                print("✓ Account found from authenticated user: {account.email}")
            except Account.DoesNotExist:
                print("✗ Account not found for authenticated user: {user.email}")
        
        if not account:
            class MockAccount:
                def __init__(self):
                    self.first_name = "Admin"
                    self.last_name = "User"
                    self.email = "admin@ppms.com"
                    self.barangay = None
                
                @property
                def full_name(self):
                    return "{self.first_name} {self.last_name}"
            
            account = MockAccount()
            print("✓ Using hardcoded admin account")

        # Get all barangays and their data
        barangays = Barangay.objects.all()
        print("✓ Found {barangays.count()} barangays")
        
        # Overall summary counters - matching template expectations
        overall_summary = {
            'Severely_Underweight': 0,
            'Underweight': 0,
            'Normal': 0,
            'Overweight': 0,
            'Obese': 0,
        }

        # Barangay detail data for the template
        barangay_details = []
        total_preschoolers = 0
        total_at_risk = 0
        highest_barangay = None
        highest_count = 0
        
        today = date.today()
        
        for brgy in barangays:
            preschoolers = Preschooler.objects.filter(
                barangay=brgy,
                is_archived=False
            ).prefetch_related('bmi_set')

            nutritional_counts = {
                'severely_underweight': 0,
                'underweight': 0,
                'normal': 0,
                'risk_of_overweight': 0,
                'overweight': 0,
                'obese': 0,
            }

            preschoolers_with_bmi = 0
            preschooler_count = preschoolers.count()
            total_preschoolers += preschooler_count

            for p in preschoolers:
                latest_bmi = p.bmi_set.order_by('-date_recorded').first()
                if latest_bmi:
                    try:
                        # Calculate age in months
                        birth_date = p.birth_date
                        age_years = today.year - birth_date.year
                        age_months = today.month - birth_date.month
                        if today.day < birth_date.day:
                            age_months -= 1
                        if age_months < 0:
                            age_years -= 1
                            age_months += 12
                        total_age_months = age_years * 12 + age_months

                        # Calculate BMI and classify
                        bmi_value = calculate_bmi(latest_bmi.weight, latest_bmi.height)
                        z = bmi_zscore(p.sex, total_age_months, bmi_value)
                        category = classify_bmi_for_age(z)

                        preschoolers_with_bmi += 1
                        
                        # Map categories to match template expectations
                        if category == "Severely Wasted":
                            nutritional_counts['severely_underweight'] += 1
                            overall_summary['Severely_Underweight'] += 1
                        elif category == "Wasted":
                            nutritional_counts['underweight'] += 1
                            overall_summary['Underweight'] += 1
                        elif category == "Normal":
                            nutritional_counts['normal'] += 1
                            overall_summary['Normal'] += 1
                        elif category in ["Risk of Overweight", "Overweight"]:
                            nutritional_counts['overweight'] += 1
                            overall_summary['Overweight'] += 1
                        elif category == "Obese":
                            nutritional_counts['obese'] += 1
                            overall_summary['Obese'] += 1

                    except Exception as e:
                        print("⚠️ BMI classification error for preschooler {p.id}: {e}")

            # Calculate at-risk percentage (severely underweight + underweight + obese)
            at_risk_count = (nutritional_counts['severely_underweight'] + 
                           nutritional_counts['underweight'] + 
                           nutritional_counts['obese'])
            at_risk_percentage = round((at_risk_count / preschooler_count * 100), 1) if preschooler_count > 0 else 0
            total_at_risk += at_risk_count

            # Track highest count barangay
            if preschooler_count > highest_count:
                highest_count = preschooler_count
                highest_barangay = brgy

            barangay_details.append({
                'name': brgy.name,
                'total_preschoolers': preschooler_count,
                'severely_underweight': nutritional_counts['severely_underweight'],
                'underweight': nutritional_counts['underweight'],
                'normal': nutritional_counts['normal'],
                'overweight': nutritional_counts['overweight'],
                'obese': nutritional_counts['obese'],
                'at_risk_percentage': at_risk_percentage,
            })

        print("✓ Processed {len(barangay_details)} barangays")
        print("✓ Overall summary: {overall_summary}")

        # Total barangays count
        total_barangays = barangays.count()

        # Context for template - matching reportTemplate.html expectations
        context = {
            'account': account,
            'report_date': date.today().strftime('%B %d, %Y'),
            'total_preschoolers': total_preschoolers,
            'total_barangays': total_barangays,
            'total_at_risk': total_at_risk,
            'highest_barangay': highest_barangay,
            'barangay_details': barangay_details,
            'overall_summary': overall_summary,
        }

        print("✓ Context prepared")

        # Use the new comprehensive template
        template_path = 'HTML/reportTemplate.html'  # Updated to use your new template
        try:
            html_string = render_to_string(template_path, context)
            print("✓ Template rendered successfully, length: {len(html_string)}")
        except Exception as e:
            print("✗ Template rendering failed: {e}")
            # Fall back to the admin comprehensive report
            template_path = 'HTML/admin_comprehensive_report.html'
            try:
                html_string = render_to_string(template_path, context)
                print("✓ Fallback template rendered, length: {len(html_string)}")
            except Exception as e2:
                print("✗ Fallback template also failed: {e2}")
                messages.error(request, "Template rendering failed: {e}")
                return redirect('Admindashboard')
        
        # Generate PDF with better error handling
        try:
            print("→ Starting PDF generation...")
            html = HTML(string=html_string, base_url=request.build_absolute_uri())
            print("→ HTML object created")
            
            # Add CSS for better PDF rendering
            pdf = html.write_pdf(stylesheets=[])
            print("✓ PDF generated successfully, size: {len(pdf)} bytes")
        except Exception as e:
            print("✗ PDF generation failed: {e}")
            print("Error type: {type(e)}")
            print("Traceback: {traceback.format_exc()}")
            
            # Try to provide more specific error info
            if "font" in str(e).lower():
                messages.error(request, "PDF generation failed due to font issues. Please check server fonts.")
            elif "css" in str(e).lower():
                messages.error(request, "PDF generation failed due to CSS issues.")
            else:
                messages.error(request, "PDF generation failed: {e}")
            return redirect('Admindashboard')

        # Create response
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = "PPMS-Cluster4-Overall-Barangay-Report-{date.today().strftime('%Y%m%d')}.pd"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        print("✓ Response created with filename: {filename}")
        print("=== DEBUG COMPLETE ===")
        return response

    except Exception as e:
        print("✗ Unexpected error: {str(e)}")
        print("Traceback: {traceback.format_exc()}")
        logger.error("Error generating report: {str(e)}")
        messages.error(request, "Error generating report: {str(e)}")
        return redirect('Admindashboard')
    
from openpyxl import Workbook
from io import BytesIO
from openpyxl.utils import get_column_letter


def generate_nutrition_excel(request):

    if not request.user.is_authenticated:
        return redirect('login')
    """Generate Excel file with nutrition report data - Direct download with barangay filtering"""
    month = request.GET.get("month")
    
    # Get user's barangay using consistent logic
    user_barangay = get_user_barangay(request.user)
    current_user_info = None
    
    print("DEBUG: Nutrition Excel - Current user email: {request.user.email}")
    
    if request.user.is_authenticated:
        # Try each model to find the user
        try:
            # Try Account model first
            account = Account.objects.select_related('barangay').get(email=request.user.email)
            current_user_info = {
                'model': 'Account',
                'name': account.full_name,
                'role': account.user_role,
                'object': account
            }
            print("DEBUG: Found in Account: {account.email}, Barangay: {user_barangay}")
        except Account.DoesNotExist:
            try:
                # Try BHW model
                bhw = BHW.objects.select_related('barangay').get(email=request.user.email)
                current_user_info = {
                    'model': 'BHW',
                    'name': bhw.full_name,
                    'role': 'BHW',
                    'object': bhw
                }
                print("DEBUG: Found in BHW: {bhw.email}, Barangay: {user_barangay}")
            except BHW.DoesNotExist:
                try:
                    # Try BNS model
                    bns = BNS.objects.select_related('barangay').get(email=request.user.email)
                    current_user_info = {
                        'model': 'BNS',
                        'name': bns.full_name,
                        'role': 'BNS',
                        'object': bns
                    }
                    print("DEBUG: Found in BNS: {bns.email}, Barangay: {user_barangay}")
                except BNS.DoesNotExist:
                    try:
                        # Try Midwife model
                        midwife = Midwife.objects.select_related('barangay').get(email=request.user.email)
                        current_user_info = {
                            'model': 'Midwife',
                            'name': midwife.full_name,
                            'role': 'Midwife',
                            'object': midwife
                        }
                        print("DEBUG: Found in Midwife: {midwife.email}, Barangay: {user_barangay}")
                    except Midwife.DoesNotExist:
                        try:
                            # Try Nurse model
                            nurse = Nurse.objects.select_related('barangay').get(email=request.user.email)
                            current_user_info = {
                                'model': 'Nurse',
                                'name': nurse.full_name,
                                'role': 'Nurse',
                                'object': nurse
                            }
                            print("DEBUG: Found in Nurse: {nurse.email}, Barangay: {user_barangay}")
                        except Nurse.DoesNotExist:
                            try:
                                # Try Parent model
                                parent = Parent.objects.select_related('barangay').get(email=request.user.email)
                                current_user_info = {
                                    'model': 'Parent',
                                    'name': parent.full_name,
                                    'role': 'Parent',
                                    'object': parent
                                }
                                print("DEBUG: Found in Parent: {parent.email}, Barangay: {user_barangay}")
                            except Parent.DoesNotExist:
                                print("DEBUG: User not found in any model")

    # Validate user and barangay
    if not current_user_info or not user_barangay:
        # Return error response or redirect
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Access denied: No barangay assigned to your account or user not found.")

    print("DEBUG: Excel generation authorized for {current_user_info['role']} in {user_barangay}")
    
    # Parse month if provided
    selected_month = None
    selected_year = None
    month_name = "All Time"
    
    if month:
        try:
            selected_year, selected_month = map(int, month.split('-'))
            month_name = f"{calendar.month_name[selected_month]} {selected_year}"
        except (ValueError, IndexError):
            month = None
    
    # Filter preschoolers ONLY from user's barangay
    preschoolers_query = Preschooler.objects.filter(
        is_archived=False,
        barangay=user_barangay  # Only preschoolers from user's barangay
    )
    
    print("DEBUG: Base query found {preschoolers_query.count()} preschoolers in {user_barangay}")
    
    if month and selected_month and selected_year:
        preschoolers_query = preschoolers_query.filter(
            bmi__date_recorded__year=selected_year,
            bmi__date_recorded__month=selected_month
        ).distinct()
        print("DEBUG: After month filter: {preschoolers_query.count()} preschoolers")
    
    # Use correct relationship with barangay filtering
    preschoolers = preschoolers_query.prefetch_related('bmi_set').select_related('parent_id', 'barangay')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Nutrition Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFFFF")  # White

    header_fill = PatternFill(start_color="007b9e", end_color="007b9e", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    # Add title with barangay information
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f"Nutrition Report - {user_barangay.name} - {month_name}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = center_alignment
    
    # Add generation info with user details
    ws.merge_cells('A2:J2')
    info_cell = ws['A2']
    info_cell.value = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')} by {current_user_info['name']} ({current_user_info['role']})"
    info_cell.font = Font(italic=True)
    info_cell.alignment = center_alignment
    
    # Add barangay info
    ws.merge_cells('A3:J3')
    barangay_cell = ws['A3']
    barangay_cell.value = f"Barangay: {user_barangay.name}"
    barangay_cell.font = Font(bold=True)
    barangay_cell.alignment = center_alignment
    
    # Add headers (row 5, shifted down due to barangay info)
    headers = [
        "Child Seq.",
        "Address or Location of Child's Residence\nPurok, Block#, Area or Location in the Barangay",
        "Name of Mother or Caregiver\n(Surname, First Name)",
        "Full Name of Child\n(Surname, First Name)",
        "Belongs to IP Group?\nYES/NO",
        "Sex\nM/",
        "Date of Birth",
        "Date Measured",
        "Weight\n(kg)",
        "Height\n(cm)"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Add data
    row_num = 6
    processed_count = 0
    
    for preschooler in preschoolers:
        # Double-check barangay (security measure)
        if preschooler.barangay != user_barangay:
            print("DEBUG: Skipping preschooler {preschooler.first_name} {preschooler.last_name} - wrong barangay")
            continue
            
        bmi_query = preschooler.bmi_set.all()
        
        if month and selected_month and selected_year:
            bmi_query = bmi_query.filter(
                date_recorded__year=selected_year,
                date_recorded__month=selected_month
            )
        
        latest_bmi = bmi_query.order_by('-date_recorded').first()
        
        if not latest_bmi:
            continue
            
        if not all([latest_bmi.weight, latest_bmi.height, latest_bmi.date_recorded]):
            continue
        
        try:
            # Handle parent relationship using ForeignKey
            if preschooler.parent_id:
                parent = preschooler.parent_id
                # Verify parent is also in same barangay
                if parent.barangay != user_barangay:
                    print("DEBUG: Skipping preschooler {preschooler.first_name} {preschooler.last_name} - parent in different barangay")
                    continue
                    
                if hasattr(parent, 'mother_name') and parent.mother_name:
                    mother_name = parent.mother_name
                elif hasattr(parent, 'full_name') and parent.full_name:
                    mother_name = parent.full_name
                else:
                    mother_name = 'N/A'
            else:
                mother_name = 'N/A'
            
            data = [
                processed_count + 1,  # Child sequence number
                preschooler.address or "{user_barangay.name}",  # Include barangay in address
                mother_name,
                f"{preschooler.last_name}, {preschooler.first_name}",
                'NO',  # IP Group
                preschooler.sex,
                preschooler.birth_date.strftime('%b-%d-%Y') if preschooler.birth_date else 'N/A',
                latest_bmi.date_recorded.strftime('%b-%d-%Y'),
                latest_bmi.weight,
                latest_bmi.height
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                cell.alignment = center_alignment
                cell.border = border
                
                # Alternate row coloring
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="f9f9f9", end_color="f9f9f9", fill_type="solid")
            
            row_num += 1
            processed_count += 1
            
        except Exception as e:
            print("Error processing preschooler {preschooler.preschooler_id} for Excel: {e}")
            continue
    
    print("DEBUG: Processed {processed_count} preschoolers for Excel export")
    
    # Add empty rows to match template (up to row 482)
    for empty_row in range(row_num, 487):  # Adjusted for shifted rows
        for col in range(1, 11):
            cell = ws.cell(row=empty_row, column=col)
            cell.value = processed_count + (empty_row - row_num) + 1 if col == 1 else ""  # Only sequence number
            cell.alignment = center_alignment
            cell.border = border
            if empty_row % 2 == 0:
                cell.fill = PatternFill(start_color="f9f9f9", end_color="f9f9f9", fill_type="solid")
    
    # Auto-adjust column widths
    column_widths = [10, 25, 25, 25, 15, 8, 15, 15, 10, 10]
    for col, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = width
    
    # Set row heights (adjusted for shifted rows)
    for row in range(5, 487):
        ws.row_dimensions[row].height = 20
    
    # Save workbook to buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create HTTP response with barangay in filename
    filename = f"Nutrition-Report-{user_barangay.name}-{month_name.replace(' ', '-')}.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    print("DEBUG: Generated Excel file: {filename} with {processed_count} preschoolers")
    return response



from .models import FCMToken

@csrf_exempt
def save_fcm_token(request):
    """Enhanced FCM token saving with nutrition service support"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get data from POST or JSON
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            email = data.get('email')
            fcm_token = data.get('token')
            source = data.get('source', 'unknown')
            supports_nutrition = data.get('supports_nutrition_notifications', True)
        else:
            email = request.POST.get('email')
            fcm_token = request.POST.get('token')
            source = request.POST.get('source', 'form')
            supports_nutrition = True

        logger.info(f"🔥 Attempting to save FCM token for nutrition services")
        logger.info(f"📧 Email: {email}")
        logger.info(f"🔑 FCM Token: {fcm_token[:20] if fcm_token else 'None'}...")
        logger.info(f"🍎 Source: {source}, Supports Nutrition: {supports_nutrition}")

        if not email or not fcm_token:
            logger.warning("❌ Missing email or FCM token")
            return JsonResponse({'success': False, 'error': 'Email and token required'})

        # Find and update account
        try:
            from .models import Account, FCMToken  # Adjust import path as needed
            
            account = Account.objects.get(email=email)
            logger.info(f"✅ Found account: {account.email}")
            
            # Update account FCM token
            old_token = account.fcm_token
            account.fcm_token = fcm_token
            account.save(update_fields=['fcm_token'])
            logger.info(f"🔄 Updated token from {old_token[:20] if old_token else 'None'}... to {fcm_token[:20]}...")

            # Also create/update FCMToken record if you have this model
            try:
                fcm_token_obj, created = FCMToken.objects.update_or_create(
                    account=account,
                    defaults={
                        'token': fcm_token,
                        'device_type': 'android',
                        'is_active': True,
                        'updated_at': timezone.now(),
                        'supports_nutrition_notifications': supports_nutrition,
                        'registration_source': source
                    }
                )
                logger.info(f"✅ FCMToken record {'created' if created else 'updated'}")
            except Exception as fcm_model_error:
                logger.warning(f"⚠️ FCMToken model update failed (this may be OK): {fcm_model_error}")

            logger.info(f"✅ FCM token saved successfully for {email}")
            
            # Send test notification specifically for nutrition services
            test_result = PushNotificationService.send_push_notification(
                token=fcm_token,
                title="🍎 Nutrition Services Notifications Enabled",
                body=f"Your device is now registered for nutrition service notifications!",
                data={
                    "type": "registration_success",
                    "email": email,
                    "supports_nutrition": str(supports_nutrition),
                    "timestamp": str(timezone.now())
                }
            )
            
            return JsonResponse({
                'success': True, 
                'message': f'FCM token saved for {email} - Nutrition notifications enabled',
                'email': email,
                'token_preview': f"{fcm_token[:20]}...",
                'supports_nutrition': supports_nutrition,
                'test_notification': test_result
            })

        except Account.DoesNotExist:
            logger.error(f"❌ Account not found: {email}")
            return JsonResponse({'success': False, 'error': f'Account not found for {email}'})

    except json.JSONDecodeError as e:
        logger.error(f"❌ JSON decode error: {e}")
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"❌ FCM save error: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
@require_POST
def register_fcm_token(request):
    """Enhanced FCM token registration with nutrition service support"""
    try:
        data = json.loads(request.body)
        token = data.get('token')
        device_type = data.get('device_type', 'android')
        supports_nutrition = data.get('supports_nutrition_notifications', True)
        source = data.get('source', 'app_startup')

        logger.info(f"🔥 Enhanced FCM token registration attempt")
        logger.info(f"🔑 Token: {token[:20] if token else 'None'}...")
        logger.info(f"📱 Device: {device_type}")
        logger.info(f"🍎 Supports Nutrition: {supports_nutrition}")
        logger.info(f"📍 Source: {source}")

        if not token:
            return JsonResponse({'success': False, 'message': 'FCM token required'})

        # Store token temporarily for later association
        # This is useful when the token arrives before login
        try:
            from .models import FCMToken
            
            # Check if token already exists
            existing_token = FCMToken.objects.filter(token=token).first()
            if existing_token:
                existing_token.device_type = device_type
                existing_token.supports_nutrition_notifications = supports_nutrition
                existing_token.registration_source = source
                existing_token.updated_at = timezone.now()
                existing_token.is_active = True
                existing_token.save()
                logger.info(f"✅ Updated existing FCM token record")
            else:
                # Create temporary token record without account association
                FCMToken.objects.create(
                    token=token,
                    device_type=device_type,
                    supports_nutrition_notifications=supports_nutrition,
                    registration_source=source,
                    is_active=True,
                    created_at=timezone.now(),
                    updated_at=timezone.now()
                    # account will be null until login
                )
                logger.info(f"✅ Created temporary FCM token record")
                
        except Exception as model_error:
            logger.warning(f"⚠️ Could not create FCMToken model (this may be OK): {model_error}")
        
        logger.info(f"✅ FCM token received and processed: {token[:20]}...")
        
        return JsonResponse({
            'success': True, 
            'message': 'FCM token registered successfully for nutrition services',
            'token_preview': f"{token[:20]}...",
            'supports_nutrition': supports_nutrition,
            'device_type': device_type
        })
        
    except Exception as e:
        logger.error(f"❌ FCM token registration error: {e}")
        return JsonResponse({'success': False, 'message': str(e)})
    
import requests

@login_required
def check_notification_status(request):
    """Debug endpoint to check notification configuration and status"""
    try:
        from .models import Preschooler, Account, Parent
        from django.conf import settings
        import os
        
        # Check FCM configuration
        fcm_configured = False
        try:
            firebase_key_path = getattr(settings, 'FIREBASE_KEY_PATH', None)
            if firebase_key_path and os.path.exists(firebase_key_path):
                fcm_configured = True
        except:
            pass
        
        # Check email configuration
        email_configured = bool(
            getattr(settings, 'EMAIL_HOST', None) and 
            getattr(settings, 'DEFAULT_FROM_EMAIL', None)
        )
        
        # Count parents and tokens
        total_parents = Parent.objects.count()
        parents_with_tokens = Account.objects.filter(
            fcm_token__isnull=False
        ).exclude(fcm_token='').count()
        
        # Sample FCM tokens (first 20 chars for security)
        sample_tokens = list(Account.objects.filter(
            fcm_token__isnull=False
        ).exclude(fcm_token='').values_list('email', 'fcm_token')[:5])
        
        sample_tokens_safe = [
            {
                'email': token[0], 
                'token_preview': token[1][:20] + '...' if token[1] else 'None'
            } 
            for token in sample_tokens
        ]
        
        return JsonResponse({
            'success': True,
            'fcm_configured': fcm_configured,
            'email_configured': email_configured,
            'total_parents': total_parents,
            'parents_with_tokens': parents_with_tokens,
            'sample_tokens': sample_tokens_safe,
            'firebase_key_exists': os.path.exists(getattr(settings, 'FIREBASE_KEY_PATH', '')) if getattr(settings, 'FIREBASE_KEY_PATH', None) else False
        })
        
    except Exception as e:
        logger.error("Error checking notification status: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })



@login_required
@require_POST
def test_push_notification(request):
    """Test endpoint to send a push notification to verify setup"""
    try:
        from .models import Account
        from .services import PushNotificationService  # Adjust import path
        
        # Get email from request
        email = request.POST.get('email') or request.user.email
        
        if not email:
            return JsonResponse({
                'success': False, 
                'error': 'Email required'
            })
        
        # Find account with FCM token
        account = Account.objects.filter(email=email, fcm_token__isnull=False).exclude(fcm_token='').first()
        
        if not account:
            return JsonResponse({
                'success': False, 
                'error': f'No FCM token found for {email}'
            })
        
        logger.info(f"🧪 Testing push notification for {email}")
        logger.info(f"🔑 Using FCM token: {account.fcm_token[:20]}...")
        
        # Send test notification
        result = PushNotificationService.send_push_notification(
            token=account.fcm_token,
            title="🧪 Test Notification",
            body="This is a test push notification from PPMS system. If you receive this, notifications are working!",
            data={
                "type": "test_notification",
                "timestamp": str(timezone.now()),
                "test": "true"
            }
        )
        
        logger.info(f"🧪 Test notification result: {result}")
        
        return JsonResponse({
            'success': result.get('success', False),
            'message': 'Test notification sent!' if result.get('success') else 'Test notification failed',
            'result': result,
            'email': email,
            'token_preview': account.fcm_token[:20] + '...'
        })
        
    except Exception as e:
        logger.error(f"❌ Test notification error: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })





@csrf_protect
@login_required
@require_http_methods(["POST"])
def save_temperature(request):
    try:
        preschooler_id = request.POST.get('preschooler_id')
        temperature_value = request.POST.get('temperature_value')
        
        if not preschooler_id or not temperature_value:
            return JsonResponse({
                'status': 'error',
                'message': 'Missing required fields'
            })
        
        # Convert and validate temperature
        try:
            temp_float = float(temperature_value)
            if temp_float < 30 or temp_float > 45:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Temperature must be between 30°C and 45°C'
                })
        except (ValueError, TypeError):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid temperature value'
            })
        
        # Get the preschooler
        try:
            preschooler = Preschooler.objects.get(preschooler_id=preschooler_id, is_archived=False)
        except Preschooler.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Preschooler not found'
            })
        
        # Get BHW record - adjust this based on your actual model relationships
        bhw = None
        try:
            # Option 1: If BHW model has a direct user field
            if hasattr(BHW, 'user'):
                bhw = BHW.objects.get(user=request.user)
            
            # Option 2: If you need to find BHW by username or other field
            # elif hasattr(BHW, 'username'):
            #     bhw = BHW.objects.get(username=request.user.username)
            
            # Option 3: If BHW has an email field
            # elif hasattr(BHW, 'email'):
            #     bhw = BHW.objects.get(email=request.user.email)
            
            # Option 4: If you need to find by a custom field
            # else:
            #     bhw = BHW.objects.filter(/* your custom condition */).first()
                
        except BHW.DoesNotExist:
            # Create temperature record without BHW if not found
            pass
        except Exception as e:
            print("Error finding BHW: {e}")
        
        # Create temperature record
        temperature_record = Temperature.objects.create(
            preschooler_id=preschooler,
            temperature_value=temp_float,
            recorded_by=bhw  # This can be None if BHW not found
        )
        
        return JsonResponse({
            'status': 'success',
            'message': f'Temperature {temp_float}°C recorded successfully',
            'temperature_id': temperature_record.temperature_id,
            'date_recorded': temperature_record.date_recorded.strftime('%Y-%m-%d')
        })
        
    except Exception as e:
        print("Temperature save error: {e}")  # For debugging
        return JsonResponse({
            'status': 'error',
            'message': 'An unexpected error occurred while saving temperature'
        })

#10/6/2025

@csrf_exempt
@require_http_methods(["POST"])
def announce_device(request):
    """
    ESP32 announces itself as online when it connects to WiFi
    """
    global DEVICE_STATUS

    try:
        data = json.loads(request.body)
        device_id = data.get("device_id")
        device_name = data.get("device_name", device_id)

        if not device_id:
            return JsonResponse({
                "status": "error",
                "message": "device_id is required"
            }, status=400)

        # Create or update device status
        DEVICE_STATUS[device_id] = {
            "device_name": device_name,
            "last_seen": timezone.now(),
            "is_online": True,
            "measurements_today": DEVICE_STATUS.get(device_id, {}).get("measurements_today", 0),
            "last_measurement_type": DEVICE_STATUS.get(device_id, {}).get("last_measurement_type", None),
        }

        return JsonResponse({
            "status": "success",
            "message": f"Device {device_id} announced online",
            "device_id": device_id,
            "device_name": device_name
        })

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)



def get_pending_validation_count(request):
    count = Account.objects.filter(
        is_validated=False
    ).exclude(user_role="parent").count()  # Changed "Parent" to "parent"
    return JsonResponse({'pending_count': count})



def generate_random_password(length=12):
    """Generate a random password with letters and numbers only (no special characters)"""
    characters = string.ascii_letters + string.digits
    password = [
        random.choice(string.ascii_uppercase),
        random.choice(string.ascii_lowercase),
        random.choice(string.digits)
    ]
    password += [random.choice(characters) for _ in range(length - 3)]
    random.shuffle(password)
    return ''.join(password)

@require_http_methods(["POST"])
def create_admin_view(request):
    """View to handle admin account creation via AJAX"""
    
    # Check if user is logged in as admin via session
    user_role = request.session.get('user_role', '').lower()
    
    if not (request.user.is_authenticated or user_role == 'admin'):
        return JsonResponse({
            'success': False,
            'error': 'You must be logged in to perform this action'
        }, status=401)
    
    # Check if user is admin/staff
    if not (request.user.is_staff or request.user.is_superuser or user_role == 'admin'):
        return JsonResponse({
            'success': False,
            'error': 'You do not have permission to create admin accounts'
        }, status=403)
    
    try:
        # Parse JSON data
        data = json.loads(request.body)
        first_name = data.get('first_name', '').strip()
        last_name = data.get('last_name', '').strip()
        email = data.get('email', '').strip().lower()
        
        # Validation
        if not all([first_name, last_name, email]):
            return JsonResponse({
                'success': False,
                'error': 'All fields are required'
            }, status=400)
        
        # Import models
        from .models import Admin, Account
        
        # Check if email already exists
        if Admin.objects.filter(email=email).exists():
            return JsonResponse({
                'success': False,
                'error': 'An admin account with this email already exists'
            }, status=400)
        
        if Account.objects.filter(email=email).exists():
            return JsonResponse({
                'success': False,
                'error': 'An account with this email already exists'
            }, status=400)
        
        if User.objects.filter(email=email).exists() or User.objects.filter(username=email).exists():
            return JsonResponse({
                'success': False,
                'error': 'A user with this email already exists'
            }, status=400)
        
        # Generate random password
        password = generate_random_password()
        
        # Create admin account within a transaction
        with transaction.atomic():
            # Step 1: Create Django User (this is the most important for login)
            user = User.objects.create_user(
                username=email,
                email=email,
                password=password,  # Django hashes this automatically
                first_name=first_name,
                last_name=last_name,
                is_staff=True,
                is_superuser=False,
                is_active=True
            )
            
            # Step 2: Create Account entry
            full_name = f"{first_name} {last_name}".strip()
            account = Account.objects.create(
                user=user,
                email=email,
                first_name=first_name,
                last_name=last_name,
                full_name=full_name,
                contact_number='',
                user_role='admin',
                is_validated=True,
                password=user.password,  # Copy hashed password from User
                sex='Male',
                must_change_password=True,  # Force password change on first login
                created_at=timezone.now(),
                last_activity=timezone.now()
            )
            
            # Step 3: Create Admin model entry
            admin = Admin.objects.create(
                username=email,
                email=email,
                password=user.password  # Copy hashed password from User
            )
        
        # ========== Background email sending ==========
        def send_admin_registration_email():
            try:
                subject = "PPMS Cluster 4 – Admin Account Created"
                html_message = f"""
                <html>
                <body style='font-family: Arial, sans-serif; background-color: #f9fafb; padding: 20px;'>
                    <div style='max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
                        <div style='text-align: center; border-bottom: 3px solid #1565c0; padding-bottom: 20px; margin-bottom: 20px;'>
                            <h1 style='color: #1565c0; margin: 0;'>PPMS Cluster 4</h1>
                            <p style='color: #6b7280; margin: 5px 0 0 0;'>Imus City Healthcare Management</p>
                        </div>
                        
                        <h2 style='text-align: center; color: #1e293b;'>Admin Account Created!</h2>
                        
                        <p style='font-size: 16px; color: #334155;'>Hello <strong>{full_name}</strong>,</p>
                        
                        <p style='font-size: 16px; color: #334155;'>
                            Your administrator account has been successfully created for the PPMS Cluster 4 system.
                        </p>
                        
                        <div style='background: #eff6ff; border: 1px solid #3b82f6; padding: 20px; border-radius: 8px; margin: 20px 0;'>
                            <h3 style='color: #1e40af; margin-top: 0;'> Login Credentials</h3>
                            <p style='margin: 10px 0; color: #1e40af;'>
                                <strong>Email:</strong> {email}<br>
                                <strong>Temporary Password:</strong> <code style='background: #dbeafe; padding: 4px 8px; border-radius: 4px; font-size: 14px; font-weight: bold;'>{password}</code>
                            </p>
                        </div>
                        
                        <div style='background: #fef3c7; border: 1px solid #f59e0b; padding: 15px; border-radius: 8px; margin: 20px 0;'>
                            <p style='margin: 0; color: #92400e;'>
                                <strong>⚠️ IMPORTANT SECURITY NOTICE:</strong><br>
                                • You <strong>MUST</strong> change your password immediately upon first login<br>
                                • Do not share this password with anyone<br>
                                • This is a one-time use password for initial access
                            </p>
                        </div>
                        
                        <div style='background: #f3f4f6; border-left: 4px solid #5bdab3; padding: 15px; border-radius: 4px; margin: 20px 0;'>
                            <h4 style='margin-top: 0; color: #374151;'> Your Admin Privileges Include:</h4>
                            <ul style='color: #4b5563; line-height: 1.8; margin: 0; padding-left: 20px;'>
                                <li>Managing healthcare workers (BHW, BNS, Midwives, Nurses)</li>
                                <li>Managing parent and preschooler records</li>
                                <li>Validating new account registrations</li>
                                <li>Generating reports and analytics</li>
                                <li>Managing barangay assignments</li>
                                <li>System-wide administrative controls</li>
                            </ul>
                        </div>
                        
                        <div style='text-align: center; margin: 30px 0;'>
                            <a href='{request.build_absolute_uri("/")}' style='display: inline-block; background: linear-gradient(135deg, #5bdab3 0%, #1565c0 100%); color: white; padding: 14px 32px; text-decoration: none; border-radius: 8px; font-weight: 600; box-shadow: 0 4px 12px rgba(91, 218, 179, 0.3);'>
                                Login to Dashboard
                            </a>
                        </div>
                        
                        <div style='background: #fef2f2; border: 1px solid #ef4444; padding: 15px; border-radius: 8px; margin: 20px 0;'>
                            <p style='margin: 0; color: #991b1b; font-size: 14px;'>
                                <strong> Security Reminder:</strong> If you did not expect to receive this email or believe this is an error, please contact the system administrator immediately.
                            </p>
                        </div>
                        
                        <div style='text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #e5e7eb;'>
                            <p style='font-size: 13px; color: #6b7280;'>
                                Need help? Contact your system administrator<br>
                                This is an automated message. Please do not reply.<br>
                                © 2025 PPMS Cluster 4. All rights reserved.
                            </p>
                        </div>
                    </div>
                </body>
                </html>
                """
                
                plain_message = f"""
PPMS Cluster 4 – Admin Account Created

Hello {full_name},

Your administrator account has been successfully created for the PPMS Cluster 4 system.

 LOGIN CREDENTIALS:
Email: {email}
Temporary Password: {password}

 IMPORTANT SECURITY NOTICE:
• You MUST change your password immediately upon first login
• Do not share this password with anyone
• This is a one-time use password for initial access

 Your Admin Privileges Include:
• Managing healthcare workers (BHW, BNS, Midwives, Nurses)
• Managing parent and preschooler records
• Validating new account registrations
• Generating reports and analytics
• Managing barangay assignments
• System-wide administrative controls

 Security Reminder: If you did not expect to receive this email or believe this is an error, please contact the system administrator immediately.

PPMS Cluster 4
Imus City Healthcare Management

Need help? Contact your system administrator
This is an automated message. Please do not reply.
© 2025 PPMS Cluster 4. All rights reserved.
                """
                
                send_mail(
                    subject, 
                    plain_message, 
                    settings.DEFAULT_FROM_EMAIL, 
                    [email], 
                    html_message=html_message, 
                    fail_silently=False
                )
                logger.info(f"✅ Admin registration email sent successfully to {email}")
            except Exception as email_error:
                logger.warning(f"[EMAIL ERROR]: Failed to send admin registration email - {email_error}")

        # Start email sending in background thread
        threading.Thread(target=send_admin_registration_email).start()
        
        # Return success
        return JsonResponse({
            'success': True,
            'message': 'Admin account created successfully',
            'name': full_name,
            'email': email,
            'password': password,  # Plain text password for display
            'admin_id': admin.admin_id,
            'account_id': account.account_id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.error(f"Error creating admin account: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=500)










































































































